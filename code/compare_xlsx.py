#!/usr/bin/env python3
"""
compare_xlsx.py
===============
Regression-test helper for XLSX outputs.

Two modes:

  Bless (create / update expected snapshots):
    python code/compare_xlsx.py --bless actual.xlsx expected_dir/

  Compare (diff actual against snapshots — exits non-zero on mismatch):
    python code/compare_xlsx.py actual.xlsx expected_dir/

Each sheet is dumped to <expected_dir>/<sheet_name>.csv.
Values are normalised (None → "", floats stripped of trailing zeros)
so formatting changes don't produce false positives.

Typical regression-test workflow
---------------------------------
1. Run the generator once to produce known-good output.
2. Bless it:
     python code/compare_xlsx.py --bless tests/hiv_teams/out/kobo/delphi_w1_hiv_kobo_programa_a.xlsx \\
         tests/hiv_teams/expected/kobo/delphi_w1_hiv_kobo_programa_a/
3. Commit the expected/ CSVs alongside your fixture.
4. On future runs, compare:
     python code/compare_xlsx.py tests/hiv_teams/out/kobo/delphi_w1_hiv_kobo_programa_a.xlsx \\
         tests/hiv_teams/expected/kobo/delphi_w1_hiv_kobo_programa_a/
"""

import sys
import csv
import pathlib

try:
    import openpyxl
except ImportError:
    sys.exit("Install openpyxl:  pip install openpyxl")


def _require_existing_file(path, label):
    """Validate that path exists and is a file, else exit with a clear message."""
    p = pathlib.Path(path)
    if not p.exists():
        sys.exit(
            f"Error: {label} not found: {p}\n"
            f"Hint: run the generator first, or check the path."
        )
    if not p.is_file():
        sys.exit(f"Error: {label} must be a file, got directory: {p}")
    return p


def _require_existing_dir(path, label):
    """Validate that path exists and is a directory, else exit with a clear message."""
    p = pathlib.Path(path)
    if not p.exists():
        sys.exit(
            f"Error: {label} not found: {p}\n"
            f"Hint: run with --bless first to create expected snapshots."
        )
    if not p.is_dir():
        sys.exit(f"Error: {label} must be a directory, got file: {p}")
    return p


def _safe_load_workbook(xlsx_path):
    """Load workbook with friendly errors for missing/invalid inputs."""
    try:
        return openpyxl.load_workbook(xlsx_path, data_only=True)
    except FileNotFoundError:
        sys.exit(
            f"Error: XLSX not found: {xlsx_path}\n"
            f"Hint: run the generator first, or check the path."
        )
    except Exception as exc:
        sys.exit(f"Error: failed to read XLSX '{xlsx_path}': {exc}")


# ── Value normalisation ───────────────────────────────────────────────────────

def _norm(value):
    """Return a stable string representation of a cell value."""
    if value is None:
        return ""
    if isinstance(value, float):
        # Strip trailing zeros: 1.0 → "1", 3.14 → "3.14"
        s = f"{value:.10f}".rstrip("0").rstrip(".")
        return s
    return str(value)


# ── Sheet → list-of-rows ──────────────────────────────────────────────────────

def sheet_to_rows(ws):
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([_norm(v) for v in row])
    # Drop trailing empty rows
    while rows and all(v == "" for v in rows[-1]):
        rows.pop()
    return rows


# ── Write a sheet to a CSV file ───────────────────────────────────────────────

def write_csv(rows, path):
    path = pathlib.Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(rows)


# ── Read a CSV back to list-of-rows ──────────────────────────────────────────

def read_csv(path):
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.reader(f))


# ── Core operations ───────────────────────────────────────────────────────────

def bless(xlsx_path, expected_dir):
    """Dump all sheets of xlsx_path as CSVs into expected_dir."""
    # We snapshot as CSV per sheet because binary .xlsx diffs are noisy and
    # not review-friendly in regression workflows.
    xlsx_path = _require_existing_file(xlsx_path, "input XLSX")
    expected_dir = pathlib.Path(expected_dir)
    wb = _safe_load_workbook(xlsx_path)
    for name in wb.sheetnames:
        rows = sheet_to_rows(wb[name])
        safe_name = name.replace("/", "_").replace("\\", "_")
        out = expected_dir / f"{safe_name}.csv"
        write_csv(rows, out)
        print(f"  blessed: {out}  ({len(rows)} rows)")
    print(f"Blessed {len(wb.sheetnames)} sheet(s) from {xlsx_path}")


def compare(xlsx_path, expected_dir):
    """Compare all sheets against CSVs in expected_dir. Returns number of mismatches."""
    xlsx_path = _require_existing_file(xlsx_path, "input XLSX")
    expected_dir = _require_existing_dir(expected_dir, "expected snapshot directory")
    wb = _safe_load_workbook(xlsx_path)

    mismatches = 0
    checked = 0

    # Check every expected CSV exists in actual workbook
    expected_csvs = {p.stem: p for p in expected_dir.glob("*.csv")}
    if not expected_csvs:
        sys.exit(
            f"Error: no CSV snapshots found in: {expected_dir}\n"
            f"Hint: run --bless to create snapshots first."
        )
    for sheet_stem, csv_path in sorted(expected_csvs.items()):
        # Find the workbook sheet whose sanitised name matches
        match = next(
            (n for n in wb.sheetnames if n.replace("/", "_").replace("\\", "_") == sheet_stem),
            None,
        )
        if match is None:
            print(f"  MISSING sheet: '{sheet_stem}' (expected from {csv_path.name})")
            mismatches += 1
            continue

        actual_rows = sheet_to_rows(wb[match])
        expected_rows = read_csv(csv_path)

        if actual_rows == expected_rows:
            print(f"  OK   sheet '{match}'  ({len(actual_rows)} rows)")
            checked += 1
            continue

        # Report differences row by row
        mismatches += 1
        n_actual   = len(actual_rows)
        n_expected = len(expected_rows)
        if n_actual != n_expected:
            print(f"  DIFF sheet '{match}': {n_expected} expected rows vs {n_actual} actual rows")
        else:
            print(f"  DIFF sheet '{match}': same row count ({n_actual}) but values differ")
        max_rows = max(n_actual, n_expected)
        diff_count = 0
        for r in range(max_rows):
            a_row = actual_rows[r]   if r < n_actual   else []
            e_row = expected_rows[r] if r < n_expected else []
            if a_row != e_row:
                diff_count += 1
                if diff_count <= 10:   # limit noise
                    # Find first differing column
                    max_cols = max(len(a_row), len(e_row))
                    for c in range(max_cols):
                        av = a_row[c]   if c < len(a_row)   else ""
                        ev = e_row[c] if c < len(e_row) else ""
                        if av != ev:
                            print(f"    row {r+1} col {c+1}: expected={ev!r}  actual={av!r}")
                            break
        if diff_count > 10:
            print(f"    ... and {diff_count - 10} more differing rows")

    # Warn about sheets present in workbook but not in expected (new, unblessed sheets)
    blessed_stems = set(expected_csvs.keys())
    for name in wb.sheetnames:
        stem = name.replace("/", "_").replace("\\", "_")
        if stem not in blessed_stems:
            print(f"  NEW  sheet '{name}' not in expected — run --bless to add it")

    if mismatches == 0:
        print(f"All {checked} sheet(s) match.")
    else:
        print(f"{mismatches} sheet(s) differ.")
    return mismatches


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    args = sys.argv[1:]
    if not args or args[0] in ("-h", "--help"):
        print(__doc__)
        sys.exit(0)

    if args[0] == "--bless":
        if len(args) < 3:
            sys.exit("Usage: compare_xlsx.py --bless <actual.xlsx> <expected_dir/>")
        bless(args[1], args[2])
    else:
        if len(args) < 2:
            sys.exit("Usage: compare_xlsx.py <actual.xlsx> <expected_dir/>")
        n = compare(args[0], args[1])
        sys.exit(1 if n else 0)


if __name__ == "__main__":
    main()
