#!/usr/bin/env python3
"""
Gerador de Relatório Delphi W1
================================
Uso:
    python gerar_relatorio_delphi_w1.py <resultados.xlsx> [dicionario.xlsx]
      [--output-dir DIR]
      [--exclude-experts ESP1,ESP2,...]
      [--exclude-file ficheiro.txt]
      [--config config.env]
      [--simple-sections | --compact-report]

Argumentos:
    resultados.xlsx   Ficheiro com os resultados agregados (obrigatório).
                      Folha usada: 'Responses' (tall-skinny, uma linha por
                      expert_code × intervention).
    dicionario.xlsx   Ficheiro dicionário com os metadados das intervenções
                      (opcional mas recomendado).
                      Folha usada: primeira cujo nome começa por 'Catalogo'.
    --output-dir DIR  Directório de saída (opcional, padrão: ../reports/).
    --exclude-experts Lista de códigos de especialistas a excluir (separados por vírgula,
              ponto-e-vírgula ou espaço). Exemplo: --exclude-experts 001PM,001XX
    --exclude-file    Ficheiro com códigos de especialistas a excluir (um por linha,
              com suporte a comentários iniciados por '#').
    --config          Ficheiro de configuração opcional para exclusões.
              Chaves suportadas: REPORT_EXCLUDE_EXPERTS, EXCLUDE_EXPERTS,
              REPORT_EXCLUDE_FILE, EXCLUDE_FILE.
    --simple-sections Omite secções visuais extensas: XY plot (2g),
              "Pontuação e Metodologia", diagrama aluvial e tabela de pontuação.
    --compact-report  Alias para --simple-sections.

Metadados das intervenções (código, nome, componente, URL):
    1. Dicionário externo (se fornecido) — folha Catalogo_*
    2. Folha Catalogo_* dentro do próprio ficheiro de resultados
    3. Colunas url_* na folha Submissions do ficheiro de resultados
       (nome e componente ficam em branco neste último caso)
    Se nenhuma fonte estiver disponível, o script usa os códigos brutos.

Colunas obrigatórias na folha Responses:
    expert_code, intervention, gate
Colunas opcionais:
    dup, which_dup, which_dup_other, intg, which_intg, which_intg_other,
    res, oth, oth_reason, impact, cmt, exp, modality, group

Saída:
    delphi_w1_relatorio_<timestamp>.html  no directório de trabalho actual
"""

import sys
import os
import math
from collections import defaultdict, OrderedDict
from datetime import datetime
import statistics
from pathlib import Path

try:
    import pandas as pd
    import openpyxl
except ImportError:
    sys.exit("Erro: dependências em falta. Execute: pip install pandas openpyxl")

# ─────────────────────────────────────────────────────────────────────────────
# EXTRACÇÃO DE METADADOS
# ─────────────────────────────────────────────────────────────────────────────

def _find_catalogo_sheet(wb):
    """Return the first sheet whose name starts with 'Catalogo' (case-insensitive)."""
    for name in wb.sheetnames:
        if name.lower().startswith("catalogo"):
            return wb[name]
    return None

def _parse_catalogo_sheet(ws):
    """
    Parse a Catalogo_* sheet.
    Expects row 1 as section headers (ignored), row 2 as column names,
    data from row 3 onward.
    Returns list of dicts with keys: code, label, component, url
    """
    # Find header row: look for a row containing 'Código'
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if any(str(c).strip() == "Código" for c in row if c):
            header_row = i
            headers = [str(c).strip() if c else "" for c in row]
            break
    if header_row is None:
        return []

    col = {h: i for i, h in enumerate(headers)}
    interventions = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        code = row[col.get("Código", -1)] if col.get("Código") is not None else None
        if not code or not str(code).strip():
            continue
        code = str(code).strip()
        label     = row[col["Intervenção"]]     if "Intervenção"  in col else None
        component = row[col["Componente"]]      if "Componente"   in col else None
        url       = row[col["URL da Ficha"]]    if "URL da Ficha" in col else None
        interventions.append({
            "code":      code,
            "label":     str(label).strip()     if label     else code,
            "component": str(component).strip() if component else "",
            "url":       str(url).strip()        if url       else "",
        })
    return interventions

def _parse_submissions_urls(wb):
    """
    Fallback: extract URLs from url_* columns in the Submissions sheet.
    Returns list of dicts with code and url only (label/component empty).
    """
    if "Submissions" not in wb.sheetnames:
        return []
    ws = wb["Submissions"]
    headers = [c.value for c in ws[1]]
    url_cols = [(i, h) for i, h in enumerate(headers)
                if h and str(h).lower().startswith("url_")]
    seen = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        for i, h in url_cols:
            code = h[4:]  # strip "url_"
            if code not in seen and row[i]:
                seen[code] = str(row[i]).strip()
    return [{"code": c, "label": c, "component": "", "url": u}
            for c, u in sorted(seen.items())]

def _parse_responses_codes(df):
    """Last-resort fallback: just use the codes found in the data."""
    codes = sorted(df["intervention"].dropna().unique())
    return [{"code": c, "label": c, "component": "", "url": ""} for c in codes]

def load_metadata(results_path, dict_path=None):
    """
    Load intervention metadata with the following priority:
      1. External dictionary file (dict_path), Catalogo_* sheet
      2. Catalogo_* sheet inside results file
      3. url_* columns in Submissions sheet of results file
      4. Raw codes from Responses sheet
    Returns ordered list of dicts: code, label, component, url
    """
    interventions = []

    # 1. External dictionary
    if dict_path and os.path.exists(dict_path):
        wb = openpyxl.load_workbook(dict_path, data_only=True)
        ws = _find_catalogo_sheet(wb)
        if ws:
            interventions = _parse_catalogo_sheet(ws)
            if interventions:
                print(f"  Metadados: dicionário externo '{os.path.basename(dict_path)}' "
                      f"({len(interventions)} intervenções)")
                return interventions

    # 2. Catalogo_* in results file
    wb_res = openpyxl.load_workbook(results_path, data_only=True)
    ws = _find_catalogo_sheet(wb_res)
    if ws:
        interventions = _parse_catalogo_sheet(ws)
        if interventions:
            print(f"  Metadados: folha Catalogo_* no ficheiro de resultados "
                  f"({len(interventions)} intervenções)")
            return interventions

    # 3. url_* columns in Submissions
    interventions = _parse_submissions_urls(wb_res)
    if interventions:
        print(f"  Metadados: colunas url_* na folha Submissions "
              f"({len(interventions)} intervenções, sem nomes)")
        return interventions

    # 4. Raw codes
    return None  # signal to caller to build from df after loading

# ─────────────────────────────────────────────────────────────────────────────
# NORMALIZAÇÃO
# ─────────────────────────────────────────────────────────────────────────────

GATE_NORM = {
    "sim_def": "sim_def", "sim definitivamente": "sim_def",
    "sim,definitivamente": "sim_def", "sim": "sim_def",
    "possivelmente": "possivelmente",
    "nao": "nao", "não": "nao", "no": "nao",
}
YN_NORM = {"sim": "sim", "yes": "sim", "nao": "nao", "não": "nao", "no": "nao"}

def norm_gate(v):
    if pd.isna(v): return None
    return GATE_NORM.get(str(v).strip().lower(), None)

def norm_yn(v):
    if pd.isna(v): return None
    return YN_NORM.get(str(v).strip().lower(), None)

def parse_multi(v):
    if pd.isna(v) or str(v).strip() == "": return []
    return [x.strip() for x in str(v).replace(",", " ").split() if x.strip()]

def safe_int(v):
    try: return int(float(str(v)))
    except: return None

# ─────────────────────────────────────────────────────────────────────────────
# CARREGAMENTO DE DADOS
# ─────────────────────────────────────────────────────────────────────────────

def load_data(path, sheet="Responses"):
    df = pd.read_excel(path, sheet_name=sheet, dtype=str)
    df.columns = [c.strip().lower() for c in df.columns]
    missing = {"expert_code", "intervention", "gate"} - set(df.columns)
    if missing:
        sys.exit(f"Erro: colunas obrigatórias em falta: {missing}")
    return df


def load_expected_experts(experts_file=None):
    """Load unique expected experts from experts.txt, ignoring comments/blank lines."""
    if experts_file is None:
        experts_file = Path(__file__).resolve().parents[1] / "experts.txt"

    experts = set()
    try:
        with open(experts_file, "r", encoding="utf-8") as f:
            for line in f:
                raw_line = line.strip()
                lower_line = raw_line.lower()
                if (
                    not raw_line
                    or raw_line.startswith("#")
                    or "# test" in lower_line
                    or "# ignore" in lower_line
                ):
                    continue
                # Support optional inline comments after '#'
                entry = raw_line.split("#", 1)[0].strip().lower()
                if entry:
                    experts.add(entry)
    except FileNotFoundError:
        return []

    return sorted(experts)


def _split_codes(value):
    """Split exclusion codes by comma/semicolon/whitespace."""
    if not value:
        return []
    import re
    return [tok.strip() for tok in re.split(r"[,;\s]+", str(value)) if tok.strip()]


def _load_simple_config(config_path=None):
    """Load key=value config from config.env-style files."""
    candidates = []
    if config_path:
        candidates.append(Path(config_path))
    else:
        # Prefer cwd config, then repo root (parent of /code)
        candidates.append(Path.cwd() / "config.env")
        candidates.append(Path(__file__).resolve().parents[1] / "config.env")

    cfg_file = next((p for p in candidates if p.exists()), None)
    if not cfg_file:
        return {}

    cfg = {}
    with open(cfg_file, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, _, v = line.partition("=")
            cfg[k.strip()] = v.strip()
    return cfg


def load_excluded_codes(cli_codes=None, exclude_file=None, config_path=None):
    """Collect exclusion expert codes from CLI, file, and config file."""
    excluded = set()

    # 1) Config values
    cfg = _load_simple_config(config_path)
    cfg_codes = cfg.get("REPORT_EXCLUDE_EXPERTS", cfg.get("EXCLUDE_EXPERTS", ""))
    excluded.update(_split_codes(cfg_codes))

    cfg_file = cfg.get("REPORT_EXCLUDE_FILE", cfg.get("EXCLUDE_FILE", "")).strip()
    if cfg_file and not exclude_file:
        exclude_file = cfg_file

    # 2) Exclude file (if provided by config or CLI)
    if exclude_file:
        if not os.path.exists(exclude_file):
            print(f"Aviso: ficheiro de exclusões não encontrado: {exclude_file}")
        else:
            with open(exclude_file, "r", encoding="utf-8") as f:
                for line in f:
                    raw = line.strip()
                    if not raw or raw.startswith("#"):
                        continue
                    # Inline comments allowed: CODE_X # reason
                    raw = raw.split("#", 1)[0].strip()
                    excluded.update(_split_codes(raw))

    # 3) CLI list has highest precedence (adds more exclusions)
    excluded.update(_split_codes(cli_codes or ""))

    # Normalize and de-duplicate while preserving intent
    return sorted({c.strip() for c in excluded if c and c.strip()})

# ─────────────────────────────────────────────────────────────────────────────
# AGREGAÇÃO
# ─────────────────────────────────────────────────────────────────────────────

def aggregate(df, interventions):
    """Aggregate responses per intervention. Returns dict keyed by code."""
    def is_placeholder_comment(value):
        txt = str(value).strip().lower()
        txt = txt.strip('"\'`.,;:!?()[]{}')
        return txt in {
            "sem comentarios",
            "sem comentário",
            "sem comentario",
            "sem comentários",
            "sem comment",
            "no comments",
        }

    inv_codes  = [r["code"] for r in interventions]
    inv_labels = {r["code"]: r["label"] for r in interventions}
    results = {}

    for inv in interventions:
        code = inv["code"]
        sub  = df[df["intervention"] == code].copy()

        # Per expert: prefer row with gate filled
        deduped = []
        for _, grp in sub.groupby("expert_code", sort=False):
            filled = grp[grp["gate"].notna() & (grp["gate"].str.strip() != "")]
            deduped.append(filled.iloc[0] if len(filled) > 0 else grp.iloc[0])

        n_sim = n_poss = n_nao = n_missing = 0
        impacts = []
        dup_yes = intg_yes = res_yes = 0
        which_dup_counts  = defaultdict(int)
        which_intg_counts = defaultdict(int)
        comments = []
        # For scoring diagnostics
        gate_scores = []   # (gate_score, impact, exp_weight) per respondent
        gate_only_scores = []  # gate_score only (unweighted), includes nao=0 when gate answered

        GATE_SCORE = {"sim_def": 1.0, "possivelmente": 0.5, "nao": 0.0}

        for row in deduped:
            g = norm_gate(row.get("gate", None))
            if   g == "sim_def":       n_sim  += 1
            elif g == "possivelmente": n_poss += 1
            elif g == "nao":           n_nao  += 1
            else:                      n_missing += 1

            if g is not None:
                gs  = GATE_SCORE.get(g, 0.0)
                exp = safe_int(row.get("exp", None))
                exp = exp if exp and 1 <= exp <= 3 else 1  # neutral weight if missing
                gate_only_scores.append(gs)

            if g in ("sim_def", "possivelmente"):
                imp = safe_int(row.get("impact", None))
                if imp and 1 <= imp <= 3:
                    impacts.append(imp)
                    gate_scores.append((gs, imp, exp))

                if norm_yn(row.get("dup", None)) == "sim":
                    dup_yes += 1
                    for t in parse_multi(row.get("which_dup", None)):
                        if t in inv_codes and t != code:
                            which_dup_counts[t] += 1

                if norm_yn(row.get("intg", None)) == "sim":
                    intg_yes += 1
                    for t in parse_multi(row.get("which_intg", None)):
                        if t in inv_codes and t != code:
                            which_intg_counts[t] += 1

                if norm_yn(row.get("res", None)) == "sim":
                    res_yes += 1

                cmt = row.get("cmt", None)
                if cmt and not pd.isna(cmt):
                  cmt_txt = str(cmt).strip()
                  if cmt_txt not in ("", ".", "None") and not is_placeholder_comment(cmt_txt):
                    comments.append(cmt_txt)

        n_resp = n_sim + n_poss + n_nao

        top_dup  = sorted(which_dup_counts,  key=which_dup_counts.get,  reverse=True)[:3]
        top_intg = sorted(which_intg_counts, key=which_intg_counts.get, reverse=True)[:3]

        avg_impact = round(sum(impacts) / len(impacts), 2) if impacts else 0

        # S_base = mean(gate_score) × mean(impact)  [range 0–3]
        if gate_scores:
            gs_vals  = [x[0] for x in gate_scores]
            imp_vals = [x[1] for x in gate_scores]
            exp_vals = [x[2] for x in gate_scores]
            s_base = round(sum(gs_vals) / len(gs_vals) * sum(imp_vals) / len(imp_vals), 3)
            # S_pond = Σ(gs × imp × exp) / Σ(exp)  [range 0–3]
            numerator = sum(gs * imp * exp for gs, imp, exp in gate_scores)
            denominator = sum(x[2] for x in gate_scores)
            s_pond = round(numerator / denominator, 3) if denominator else 0
            exp_mean = round(sum(exp_vals) / len(exp_vals), 2)
        else:
            s_base = s_pond = exp_mean = 0.0

        gate_mean = round(sum(gate_only_scores) / len(gate_only_scores), 3) if gate_only_scores else 0.0

        results[code] = {
            "code":      code,
            "label":     inv["label"],
            "url":       inv["url"],
            "component": inv["component"],
            "n_total":   n_resp,
            "n_missing": n_missing,
            "n_sim":     n_sim,
            "n_poss":    n_poss,
            "n_nao":     n_nao,
            "pct_optimizable": round((n_sim + n_poss) / n_resp * 100) if n_resp else 0,
            "pct_definitely":  round(n_sim / n_resp * 100)            if n_resp else 0,
            "avg_impact": avg_impact,
            "dup_pct":   round(dup_yes  / n_resp * 100) if n_resp else 0,
            "intg_pct":  round(intg_yes / n_resp * 100) if n_resp else 0,
            "res_pct":   round(res_yes  / n_resp * 100) if n_resp else 0,
            "top_dup":   top_dup,
            "top_intg":  top_intg,
            "comments":  comments,
            "composite": round((n_sim / n_resp) * avg_impact, 3) if n_resp and avg_impact else 0,
            "gate_mean": gate_mean,
            "s_base":    s_base,
            "s_pond":    s_pond,
            "exp_mean":  exp_mean,
        }

    return results

# ─────────────────────────────────────────────────────────────────────────────
# ESTATÍSTICAS SUMÁRIAS
# ─────────────────────────────────────────────────────────────────────────────

def summary_stats(results, df, n_experts_expected=None):
    n_experts_observed = df["expert_code"].nunique()
    n_experts = n_experts_expected if n_experts_expected and n_experts_expected > 0 else n_experts_observed
    n_inv_80    = sum(1 for r in results.values() if r["pct_optimizable"] >= 80)
    n_imp_high  = sum(1 for r in results.values() if r["avg_impact"] >= 2.5)
    n_unanimous = sum(1 for r in results.values() if r["pct_optimizable"] == 100)

    # Response rate per intervention: n_total / n_experts (% of experts who answered)
    n_experts_int = n_experts if n_experts > 0 else 1
    rr = [r["n_total"] / n_experts_int * 100 for r in results.values() if r["n_total"] > 0]
    rr_median = round(statistics.median(rr), 1) if rr else 0
    rr_min    = round(min(rr), 1)               if rr else 0
    rr_max    = round(max(rr), 1)               if rr else 0

    # Per-intervention breakdown for the response rate table
    rr_detail = sorted(
        [{"label": r["label"], "n": r["n_total"], "pct": round(r["n_total"] / n_experts_int * 100, 1)}
         for r in results.values()],
        key=lambda x: x["pct"], reverse=True
    )

    return {
        "n_experts":   n_experts,
      "n_experts_observed": n_experts_observed,
        "n_inv":       len(results),
        "n_inv_80":    n_inv_80,
        "n_imp_high":  n_imp_high,
        "n_unanimous": n_unanimous,
        "rr_median":   rr_median,
        "rr_min":      rr_min,
        "rr_max":      rr_max,
        "rr_detail":   rr_detail,
    }

# ─────────────────────────────────────────────────────────────────────────────
# PONTUAÇÃO E RANKS
# ─────────────────────────────────────────────────────────────────────────────

def compute_ranks(results):
    """Add ranks and normalized scores used in scoring diagnostics."""
    items = list(results.values())
    # Normalise to 0-100
    max_gate = max((r.get("gate_mean", 0) for r in items), default=1) or 1
    max_base = max((r["s_base"] for r in items), default=1) or 1
    max_pond = max((r["s_pond"] for r in items), default=1) or 1
    for r in items:
        r["score_gate_n"] = round(r.get("gate_mean", 0) / max_gate * 100, 1)
        r["score_base_n"] = round(r["s_base"] / max_base * 100, 1)
        r["score_wtd_n"]  = round(r["s_pond"] / max_pond * 100, 1)

    sorted_gate = sorted(items, key=lambda r: r.get("gate_mean", 0), reverse=True)
    sorted_base = sorted(items, key=lambda r: r["s_base"], reverse=True)
    sorted_pond = sorted(items, key=lambda r: r["s_pond"], reverse=True)
    rank_gate = {r["code"]: i + 1 for i, r in enumerate(sorted_gate)}
    rank_base = {r["code"]: i + 1 for i, r in enumerate(sorted_base)}
    rank_pond = {r["code"]: i + 1 for i, r in enumerate(sorted_pond)}
    for r in items:
        r["rank_gate"]  = rank_gate[r["code"]]
        r["rank_base"]  = rank_base[r["code"]]
        r["rank_wtd"]   = rank_pond[r["code"]]
        r["rank_delta"] = rank_base[r["code"]] - rank_pond[r["code"]]

# ─────────────────────────────────────────────────────────────────────────────
# SVG HELPERS (sem dependências externas)
# ─────────────────────────────────────────────────────────────────────────────

def svg_hbar_stacked(rows, width=460, bar_h=14, gap=4,
                     colors=("#1a6b3a","#d4a017","#e0e0e0"),
                     labels=("Sim def.","Possiv.","Não"),
                     label_w=220):
    """Horizontal stacked bar chart. rows = list of (label, [v1,v2,v3])."""
    h = len(rows) * (bar_h + gap) + 4
    out = [f'<svg width="{width}" height="{h}" style="display:block;overflow:visible">']
    for i, (lbl, vals) in enumerate(rows):
        y = i * (bar_h + gap)
        total = sum(vals) or 1
        out.append(f'<text x="{label_w-4}" y="{y+bar_h-2}" text-anchor="end" '
                   f'font-size="10" fill="#6b7280">{esc(lbl[:38])}</text>')
        x = label_w
        bar_total_w = width - label_w - 40
        for v, col in zip(vals, colors):
            w = round(v / total * bar_total_w)
            if w > 0:
                out.append(f'<rect x="{x}" y="{y}" width="{w}" height="{bar_h}" fill="{col}"/>')
            x += w
        pct = round(sum(vals[:2]) / total * 100) if total else 0
        out.append(f'<text x="{label_w + bar_total_w + 4}" y="{y+bar_h-2}" '
                   f'font-size="10" fill="#374151" font-weight="600">{pct}%</text>')
    out.append("</svg>")
    return "".join(out)

def svg_hbar_single(rows, width=460, bar_h=14, gap=4, color="#1a6b3a", label_w=220, fmt=".1f"):
    """Single horizontal bar per row. rows = list of (label, value, max_value)."""
    h = len(rows) * (bar_h + gap) + 4
    out = [f'<svg width="{width}" height="{h}" style="display:block;overflow:visible">']
    for i, (lbl, val, max_val) in enumerate(rows):
        y = i * (bar_h + gap)
        bar_w = width - label_w - 50
        w = round(val / max_val * bar_w) if max_val else 0
        out.append(f'<text x="{label_w-4}" y="{y+bar_h-2}" text-anchor="end" '
                   f'font-size="10" fill="#6b7280">{esc(lbl[:38])}</text>')
        out.append(f'<rect x="{label_w}" y="{y}" width="{bar_w}" height="{bar_h}" fill="#e5e7eb" rx="1"/>')
        if w > 0:
            out.append(f'<rect x="{label_w}" y="{y}" width="{w}" height="{bar_h}" fill="{color}" rx="1"/>')
        out.append(f'<text x="{label_w + bar_w + 4}" y="{y+bar_h-2}" '
                   f'font-size="10" fill="#374151" font-weight="600">'
                   f'{val:{fmt}}</text>')
    out.append("</svg>")
    return "".join(out)

def svg_donut(counts_dict, colors, size=100, label=""):
    """Simple donut chart. counts_dict = {label: count}."""
    total = sum(counts_dict.values()) or 1
    cx = cy = size / 2
    r_out, r_in = size * 0.4, size * 0.22
    import math
    angle = -math.pi / 2
    paths = []
    legend = []
    for (lbl, cnt), col in zip(counts_dict.items(), colors):
        sweep = cnt / total * 2 * math.pi
        if sweep < 0.001:
            angle += sweep
            continue
        x1 = cx + r_out * math.cos(angle)
        y1 = cy + r_out * math.sin(angle)
        x2 = cx + r_out * math.cos(angle + sweep)
        y2 = cy + r_out * math.sin(angle + sweep)
        xi1 = cx + r_in * math.cos(angle)
        yi1 = cy + r_in * math.sin(angle)
        xi2 = cx + r_in * math.cos(angle + sweep)
        yi2 = cy + r_in * math.sin(angle + sweep)
        large = 1 if sweep > math.pi else 0
        pct = round(cnt / total * 100)
        paths.append(
            f'<path d="M{xi1:.1f},{yi1:.1f} L{x1:.1f},{y1:.1f} '
            f'A{r_out},{r_out} 0 {large},1 {x2:.1f},{y2:.1f} '
            f'L{xi2:.1f},{yi2:.1f} A{r_in},{r_in} 0 {large},0 {xi1:.1f},{yi1:.1f} Z" '
            f'fill="{col}"><title>{esc(lbl)}: {cnt} ({pct}%)</title></path>'
        )
        legend.append((lbl, cnt, pct, col))
        angle += sweep
    legend_h = len(legend) * 14 + 4
    total_h = max(size, legend_h)
    svg_parts = [f'<svg width="{size + 140}" height="{total_h}" style="display:inline-block;vertical-align:middle">']
    svg_parts += paths
    if label:
        svg_parts.append(f'<text x="{cx}" y="{cy+4}" text-anchor="middle" '
                         f'font-size="9" fill="#6b7280">{esc(label)}</text>')
    for j, (lbl, cnt, pct, col) in enumerate(legend):
        ly = j * 14 + 10
        svg_parts.append(f'<rect x="{size+4}" y="{ly-7}" width="9" height="9" fill="{col}" rx="1"/>')
        svg_parts.append(f'<text x="{size+16}" y="{ly+1}" font-size="10" fill="#374151">'
                         f'{esc(lbl)} <tspan font-weight="600">{pct}%</tspan></text>')
    svg_parts.append("</svg>")
    return "".join(svg_parts)

def svg_scatter_optim_impact_exp(items, width=1060, height=420):
  """
  Single real-scale scatter with overlays:
  - Axes: real x/y values, viewport adjusted to observed ranges
  - Background shaded zones for current dual-cutoff rationale
  Encoding: size = experience mean, color = % reduction resources.
  """
  if not items:
    return '<div style="font-size:12px;color:#6b7280">Sem dados para gráfico de dispersão.</div>'

  # Color/size helpers
  def res_color(res_pct):
    p = max(0.0, min(100.0, float(res_pct or 0.0)))
    t = p / 100.0
    c1 = (226, 232, 240)
    c2 = (91, 94, 166)
    r = round(c1[0] + (c2[0] - c1[0]) * t)
    g = round(c1[1] + (c2[1] - c1[1]) * t)
    b = round(c1[2] + (c2[2] - c1[2]) * t)
    return f"rgb({r},{g},{b})"

  def exp_radius(exp_mean):
    e = max(1.0, min(3.0, float(exp_mean or 1.0)))
    return 4 + (e - 1.0) / 2.0 * 7

  gate_vals = [float(r.get("gate_mean", 0) or 0) for r in items]
  imp_vals = [float(r.get("avg_impact", 1.0) or 1.0) for r in items]
  gmin, gmax = min(gate_vals), max(gate_vals)
  imin, imax = min(imp_vals), max(imp_vals)
  gspan_obs = (gmax - gmin) if (gmax - gmin) > 1e-12 else 1.0
  ispan_obs = (imax - imin) if (imax - imin) > 1e-12 else 1.0

  # Keep plot in real units but zoom to observed range with a small padding.
  x_min = max(0.0, gmin - 0.08 * gspan_obs)
  x_max = min(1.0, gmax + 0.08 * gspan_obs)
  y_min = max(1.0, imin - 0.08 * ispan_obs)
  y_max = min(3.0, imax + 0.08 * ispan_obs)
  x_span = (x_max - x_min) if (x_max - x_min) > 1e-12 else 1.0
  y_span = (y_max - y_min) if (y_max - y_min) > 1e-12 else 1.0

  panel_w = 700
  panel_h = height
  left = 56
  right = 20
  top = 34
  bottom = 50
  w = panel_w - left - right
  h = panel_h - top - bottom

  def x_px(v):
    v = max(x_min, min(x_max, float(v)))
    return left + ((v - x_min) / x_span) * w

  def y_px(v):
    v = max(y_min, min(y_max, float(v)))
    return top + (1.0 - ((v - y_min) / y_span)) * h

  x_cut_actual = statistics.median(gate_vals)
  y_cut_actual = statistics.median(imp_vals)
  x_cut = max(x_min, min(x_max, x_cut_actual))
  y_cut = max(y_min, min(y_max, y_cut_actual))

  out = [f'<svg width="{panel_w}" height="{panel_h}" style="display:block;overflow:visible">']
  out.append(f'<text x="{left}" y="18" font-size="11" fill="#334155" font-weight="600">Escala real (janela min-max observada)</text>')

  # Shaded areas for current dual-cutoff rationale (3 recommendation structures)
  xl = x_px(x_min)
  xm = x_px(x_cut)
  xr = x_px(x_max)
  yb = y_px(y_min)
  ym = y_px(y_cut)
  yt = y_px(y_max)
  out.append(f'<rect x="{xl:.1f}" y="{ym:.1f}" width="{(xm-xl):.1f}" height="{(yb-ym):.1f}" fill="#fee2e2" fill-opacity="0.35"/>')
  out.append(f'<rect x="{xm:.1f}" y="{ym:.1f}" width="{(xr-xm):.1f}" height="{(yb-ym):.1f}" fill="#fef3c7" fill-opacity="0.30"/>')
  out.append(f'<rect x="{xl:.1f}" y="{yt:.1f}" width="{(xm-xl):.1f}" height="{(ym-yt):.1f}" fill="#fef3c7" fill-opacity="0.30"/>')
  out.append(f'<rect x="{xm:.1f}" y="{yt:.1f}" width="{(xr-xm):.1f}" height="{(ym-yt):.1f}" fill="#dcfce7" fill-opacity="0.30"/>')

  # Axes
  out.append(f'<line x1="{left}" y1="{top + h}" x2="{left + w}" y2="{top + h}" stroke="#94a3b8" stroke-width="1"/>')
  out.append(f'<line x1="{left}" y1="{top}" x2="{left}" y2="{top + h}" stroke="#94a3b8" stroke-width="1"/>')

  # Grid and ticks in real-value axes
  x_ticks = [x_min + i * (x_span / 4.0) for i in range(5)]
  y_ticks = [y_min + i * (y_span / 4.0) for i in range(5)]
  for xv in x_ticks:
    xp = x_px(xv)
    out.append(f'<line x1="{xp:.1f}" y1="{top}" x2="{xp:.1f}" y2="{top + h}" stroke="#e5e7eb" stroke-width="1"/>')
    out.append(f'<text x="{xp:.1f}" y="{top + h + 16}" text-anchor="middle" font-size="10" fill="#64748b">{xv:.2f}</text>')
  for yv in y_ticks:
    yp = y_px(yv)
    out.append(f'<line x1="{left}" y1="{yp:.1f}" x2="{left + w}" y2="{yp:.1f}" stroke="#e5e7eb" stroke-width="1"/>')
    out.append(f'<text x="{left - 8}" y="{yp + 3:.1f}" text-anchor="end" font-size="10" fill="#64748b">{yv:.2f}</text>')

  # Current dual cutoffs in real-value axes (clipped to viewport)
  out.append(f'<line x1="{x_px(x_cut):.1f}" y1="{top}" x2="{x_px(x_cut):.1f}" y2="{top + h}" stroke="#64748b" stroke-width="1" stroke-dasharray="4 3"/>')
  out.append(f'<line x1="{left}" y1="{y_px(y_cut):.1f}" x2="{left + w}" y2="{y_px(y_cut):.1f}" stroke="#64748b" stroke-width="1" stroke-dasharray="4 3"/>')
  out.append(f'<text x="{left + 6}" y="{top + 14}" font-size="9" fill="#64748b">cutoff x&gt;mediana={x_cut_actual:.3f}</text>')
  out.append(f'<text x="{left + 6}" y="{top + 26}" font-size="9" fill="#64748b">cutoff y&gt;mediana={y_cut_actual:.3f}</text>')

  # Points
  for i, r in enumerate(sorted(items, key=lambda z: z.get("composite", 0), reverse=True), 1):
    idx = int(r.get("display_idx", i))
    x = x_px(float(r.get("gate_mean", 0) or 0))
    y = y_px(float(r.get("avg_impact", 1.0) or 1.0))
    fill = res_color(r.get("res_pct", 0))
    rad = exp_radius(r.get("exp_mean", 1.0))
    tip = (f'{idx}. {esc(r.get("label", ""))} | '
           f'optimizabilidade: {r.get("gate_mean", 0):.3f} | impacto: {r.get("avg_impact", 0):.2f} | '
           f'S_base: {r.get("s_base", 0):.3f} | S_pond: {r.get("s_pond", 0):.3f}')
    out.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="{rad:.1f}" fill="{fill}" fill-opacity="0.92" stroke="#0f172a" stroke-width="0.7">'
               f'<title>{tip}</title></circle>')
    out.append(f'<text x="{x:.1f}" y="{y + 3:.1f}" text-anchor="middle" font-size="8" fill="#0f172a" font-weight="700" pointer-events="none">{idx}</text>')

  # Axis labels
  out.append(f'<text x="{left + w / 2:.1f}" y="{panel_h - 8}" text-anchor="middle" font-size="11" fill="#334155">Score de optimizabilidade</text>')
  out.append(f'<text x="14" y="{top + h / 2:.1f}" transform="rotate(-90 14 {top + h / 2:.1f})" text-anchor="middle" font-size="11" fill="#334155">Impacto esperado (1-3)</text>')

  out.append('</svg>')
  scatter_svg = ''.join(out)

  # Shared legend block
  legend = (
    '<div style="margin-top:6px;display:flex;gap:16px;align-items:flex-start;flex-wrap:wrap">'
    '<div style="font-size:10px;color:#64748b">'
    '<div style="margin-bottom:4px;font-weight:600">Cor = % redução recursos</div>'
    '<div style="width:150px;height:10px;background:linear-gradient(90deg,rgb(226,232,240),rgb(91,94,166));border:1px solid #94a3b8"></div>'
    '<div style="display:flex;justify-content:space-between;width:150px"><span>0%</span><span>50%</span><span>100%</span></div>'
    '</div>'
    '<div style="font-size:10px;color:#64748b">'
    '<div style="margin-bottom:4px;font-weight:600">Tamanho = experiência média</div>'
    '<div style="display:flex;align-items:center;gap:8px"><span style="display:inline-block;width:8px;height:8px;border-radius:50%;background:#cbd5e1;border:1px solid #64748b"></span><span>1.0</span>'
    '<span style="display:inline-block;width:14px;height:14px;border-radius:50%;background:#cbd5e1;border:1px solid #64748b"></span><span>2.0</span>'
    '<span style="display:inline-block;width:20px;height:20px;border-radius:50%;background:#cbd5e1;border:1px solid #64748b"></span><span>3.0</span></div>'
    '</div>'
    '<div style="font-size:10px;color:#64748b">'
    '<div style="margin-bottom:4px;font-weight:600">Overlays de cutoff</div>'
    '<div style="display:flex;gap:10px;align-items:center;flex-wrap:wrap">'
    '<span style="display:inline-block;width:12px;height:12px;background:#dcfce7;border:1px solid #86efac"></span><span>Alta (caixa atual)</span>'
    '<span style="display:inline-block;width:12px;height:12px;background:#fef3c7;border:1px solid #fcd34d"></span><span>Média (caixas atuais)</span>'
    '<span style="display:inline-block;width:12px;height:12px;background:#fee2e2;border:1px solid #fca5a5"></span><span>Baixa (caixa atual)</span>'
    '</div></div>'
    '</div>'
  )

  return (
    f'{scatter_svg}'
    + legend
  )

def svg_alluvial_weighting(items, width=1140, row_h=24, node_w=14, label_w=240):
    """
    Draw an alluvial-like SVG linking rank positions across 3 stages:
    optimizabilidade (left) -> S_base (middle) -> S_pond (right).
    """
    if not items:
      return '<div style="font-size:12px;color:#6b7280">Sem dados para diagrama aluvial.</div>'

    left_x = label_w
    mid_x = round(width / 2) - round(node_w / 2)
    right_x = width - label_w - node_w
    top_pad = 42
    bottom_pad = 24
    n = len(items)
    height = top_pad + bottom_pad + max(1, n) * row_h

    left_order = sorted(items, key=lambda r: (r.get("rank_gate", 10**9), r.get("label", "")))
    mid_order = sorted(items, key=lambda r: (r.get("rank_base", 10**9), r.get("label", "")))
    right_order = sorted(items, key=lambda r: (r.get("rank_wtd", 10**9), r.get("label", "")))

    left_y = {r["code"]: top_pad + i * row_h + row_h / 2 for i, r in enumerate(left_order)}
    mid_y = {r["code"]: top_pad + i * row_h + row_h / 2 for i, r in enumerate(mid_order)}
    right_y = {r["code"]: top_pad + i * row_h + row_h / 2 for i, r in enumerate(right_order)}

    def flow_color(delta):
        if delta > 0:
            return "#1a6b3a"  # moved up with weighting
        if delta < 0:
            return "#c0392b"  # moved down with weighting
        return "#1a5276"      # unchanged

    out = [
      f'<svg width="{width}" height="{height}" style="display:block;max-width:100%;overflow:visible">',
      f'<text x="{left_x - 8}" y="20" text-anchor="end" font-size="11" fill="#6b7280" '
      f'style="letter-spacing:.08em;text-transform:uppercase">Rank optimizabilidade</text>',
      f'<text x="{mid_x + node_w / 2}" y="20" text-anchor="middle" font-size="11" fill="#6b7280" '
      f'style="letter-spacing:.08em;text-transform:uppercase">Rank S_base</text>',
      f'<text x="{right_x + node_w + 8}" y="20" text-anchor="start" font-size="11" fill="#6b7280" '
      f'style="letter-spacing:.08em;text-transform:uppercase">Rank S_pond</text>',
      f'<line x1="{left_x + node_w}" y1="30" x2="{right_x}" y2="30" stroke="#e5e7eb" stroke-width="1"/>'
    ]

    # Stage 1: optimizabilidade -> S_base (impact weighting effect)
    for r in left_order:
      code = r["code"]
      y1 = left_y[code]
      y2 = mid_y[code]
      c1x = left_x + node_w + (mid_x - (left_x + node_w)) * 0.35
      c2x = left_x + node_w + (mid_x - (left_x + node_w)) * 0.65
      left_delta = r.get("rank_gate", 0) - r.get("rank_base", 0)
      col = flow_color(left_delta)
      out.append(
        f'<path d="M {left_x + node_w:.1f},{y1:.1f} C {c1x:.1f},{y1:.1f} {c2x:.1f},{y2:.1f} {mid_x:.1f},{y2:.1f}" '
        f'stroke="{col}" stroke-opacity="0.45" stroke-width="6" fill="none">'
        f'<title>#{r.get("display_idx", "-")} - {esc(r["label"])} | Rank optimizabilidade: {r.get("rank_gate", "-")} -> Rank S_base: {r.get("rank_base", "-")} '
        f'| Δrank: {left_delta}</title></path>'
      )

    # Stage 2: S_base -> S_pond (experience weighting effect)
    for r in mid_order:
      code = r["code"]
      y1 = mid_y[code]
      y2 = right_y[code]
      c1x = mid_x + node_w + (right_x - (mid_x + node_w)) * 0.35
      c2x = mid_x + node_w + (right_x - (mid_x + node_w)) * 0.65
      col = flow_color(r.get("rank_delta", 0))
      out.append(
        f'<path d="M {mid_x + node_w:.1f},{y1:.1f} C {c1x:.1f},{y1:.1f} {c2x:.1f},{y2:.1f} {right_x:.1f},{y2:.1f}" '
        f'stroke="{col}" stroke-opacity="0.45" stroke-width="6" fill="none">'
        f'<title>#{r.get("display_idx", "-")} - {esc(r["label"])} | Rank S_base: {r.get("rank_base", "-")} -> Rank S_pond: {r.get("rank_wtd", "-")} '
        f'| Δrank: {r.get("rank_delta", 0)}</title></path>'
      )

    for r in left_order:
      code = r["code"]
      y = left_y[code] - 6
      out.append(f'<rect x="{left_x}" y="{y:.1f}" width="{node_w}" height="12" fill="#7a5c00" rx="2"/>')
      out.append(
        f'<text x="{left_x - 6}" y="{y + 9:.1f}" text-anchor="end" font-size="10" fill="#111827">'
        f'{r.get("display_idx", "-")}. {esc(r["label"][:32])}</text>'
      )

    for r in mid_order:
      code = r["code"]
      y = mid_y[code] - 6
      out.append(f'<rect x="{mid_x}" y="{y:.1f}" width="{node_w}" height="12" fill="#1a6b3a" rx="2"/>')
      out.append(
        f'<text x="{mid_x + node_w / 2}" y="{y - 2:.1f}" text-anchor="middle" font-size="9" fill="#4b5563">'
        f'{r.get("rank_base", "-")}</text>'
      )

    for r in right_order:
      code = r["code"]
      y = right_y[code] - 6
      out.append(f'<rect x="{right_x}" y="{y:.1f}" width="{node_w}" height="12" fill="#1a5276" rx="2"/>')
      delta = r.get("rank_delta", 0)
      delta_str = f'+{delta}' if delta > 0 else str(delta)
      out.append(
        f'<text x="{right_x + node_w + 6}" y="{y + 9:.1f}" text-anchor="start" font-size="10" fill="#111827">'
        f'{r.get("display_idx", "-")} (Δ{delta_str})</text>'
      )

    out.append('</svg>')
    return ''.join(out)

# ─────────────────────────────────────────────────────────────────────────────
# ANÁLISE UNIVARIADA
# ─────────────────────────────────────────────────────────────────────────────

def univariate_analysis(results, df):
    """Compute cross-intervention univariate distributions for the report."""
    items = list(results.values())

    # Gate aggregate
    total_sim  = sum(r["n_sim"]  for r in items)
    total_poss = sum(r["n_poss"] for r in items)
    total_nao  = sum(r["n_nao"]  for r in items)

    # Impact distribution
    imp_counts = {1: 0, 2: 0, 3: 0}
    for r in items:
        for v, n in [(1, r["n_nao"]), (2, r["n_poss"]), (3, r["n_sim"])]:
            pass  # we need per-respondent data; use avg_impact × n_total as proxy
    # Use per-row impact from df for accurate distribution
    imp_vals_all = []
    for v in df["impact"].dropna():
        i = safe_int(v)
        if i and 1 <= i <= 3:
            imp_vals_all.append(i)
            imp_counts[i] = imp_counts.get(i, 0) + 1

    # Expertise distribution
    exp_counts = {1: 0, 2: 0, 3: 0}
    for v in df["exp"].dropna() if "exp" in df.columns else []:
        e = safe_int(v)
        if e and 1 <= e <= 3:
            exp_counts[e] = exp_counts.get(e, 0) + 1

    # Per-intervention sorted lists for bar charts
    sorted_gate   = sorted(items, key=lambda r: r["pct_optimizable"], reverse=True)
    sorted_impact = sorted(items, key=lambda r: r["avg_impact"], reverse=True)
    sorted_dup    = sorted(items, key=lambda r: r["dup_pct"],  reverse=True)
    sorted_intg   = sorted(items, key=lambda r: r["intg_pct"], reverse=True)
    sorted_res    = sorted(items, key=lambda r: r["res_pct"],  reverse=True)

    return {
        "gate_agg":      {"Sim def.": total_sim, "Possiv.": total_poss, "Não": total_nao},
        "imp_counts":    imp_counts,
        "exp_counts":    exp_counts,
        "sorted_gate":   sorted_gate,
        "sorted_impact": sorted_impact,
        "sorted_dup":    sorted_dup,
        "sorted_intg":   sorted_intg,
        "sorted_res":    sorted_res,
        "imp_vals_all":  imp_vals_all,
    }

# ─────────────────────────────────────────────────────────────────────────────
# GERAÇÃO DE HTML
# ─────────────────────────────────────────────────────────────────────────────

def esc(s):
    return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;").replace('"','&quot;')

def render_html(results, stats, interventions, source_file, univariate=None,
                include_xyplot=True, include_scoring=True):
    inv_label = {r["code"]: r["label"] for r in interventions}
    sorted_inv = sorted(results.values(), key=lambda r: r["composite"], reverse=True)
    for i, r in enumerate(sorted_inv, 1):
        r["display_idx"] = i
    now = datetime.now().strftime("%d/%m/%Y às %H:%M")
    med_gate = statistics.median([r.get("gate_mean", 0) for r in sorted_inv]) if sorted_inv else 0
    med_imp = statistics.median([r.get("avg_impact", 0) for r in sorted_inv]) if sorted_inv else 0

    def tier(r):
        high_gate = r.get("gate_mean", 0) > med_gate
        high_imp = r.get("avg_impact", 0) > med_imp
        if high_gate and high_imp: return "alta"
        if high_gate or high_imp: return "media"
        return "baixa"

    alta  = [f'{r.get("display_idx", "-")}. {r["label"]}' for r in sorted_inv if tier(r) == "alta"]
    media = [f'{r.get("display_idx", "-")}. {r["label"]}' for r in sorted_inv if tier(r) == "media"]
    baixa = [f'{r.get("display_idx", "-")}. {r["label"]}' for r in sorted_inv if tier(r) == "baixa"]

    # ── Response rate section ──────────────────────────────────────────────
    rr_rows = ""
    n_experts = stats["n_experts"]
    n_experts_observed = stats.get("n_experts_observed", n_experts)
    for item in stats["rr_detail"]:
        bar_w = round(item["pct"])
        color = "#1a6b3a" if item["pct"] >= 80 else ("#d4a017" if item["pct"] >= 60 else "#e57373")
        rr_rows += f"""<tr>
          <td style="max-width:320px;font-size:12px">{esc(item["label"])}</td>
          <td style="text-align:center;font-weight:600">{item["n"]}/{n_experts}</td>
          <td style="width:200px">
            <div style="display:flex;align-items:center;gap:8px">
              <div style="flex:1;height:8px;background:#e5e7eb;border-radius:2px;overflow:hidden">
                <div style="width:{bar_w}%;height:100%;background:{color}"></div>
              </div>
              <span style="font-size:11px;font-weight:600;color:{color};white-space:nowrap">{item["pct"]}%</span>
            </div>
          </td>
        </tr>"""

    # ── Ranking table rows ──────────────────────────────────────────────────
    rank_rows = ""
    for i, r in enumerate(sorted_inv, 1):
        sim_w  = round(r["n_sim"]  / r["n_total"] * 100) if r["n_total"] else 0
        poss_w = round(r["n_poss"] / r["n_total"] * 100) if r["n_total"] else 0
        nao_w  = round(r["n_nao"]  / r["n_total"] * 100) if r["n_total"] else 0
        imp    = r["avg_impact"]
        imp_cls = "imp-high" if imp >= 2.5 else ("imp-med" if imp >= 1.8 else "imp-low")
        imp_lbl = "Alto" if imp >= 2.5 else ("Médio" if imp >= 1.8 else "Baixo")
        label_with_idx = f'{r.get("display_idx", i)}. {r["label"]}'
        name_cell = (f'<a href="{esc(r["url"])}" target="_blank" class="inv-link">{esc(label_with_idx)}</a>'
               if r["url"] else esc(label_with_idx))
        rank_rows += f"""<tr>
          <td class="rank-num">{i}</td>
          <td style="font-weight:500">{name_cell}</td>
          <td><span class="comp-tag">{esc(r["component"])}</span></td>
          <td>
            <div class="mini-bar-wrap">
              <div class="mini-bar-bg">
                <div style="display:flex;height:100%">
                  <div style="width:{sim_w}%;background:var(--sim)"></div>
                  <div style="width:{poss_w}%;background:#d4a017"></div>
                  <div style="width:{nao_w}%;background:#e5e7eb"></div>
                </div>
              </div>
              <span class="bar-counts">{r["n_sim"]}S·{r["n_poss"]}P·{r["n_nao"]}N</span>
            </div>
          </td>
          <td><strong style="font-size:14px">{r["pct_optimizable"]}%</strong></td>
          <td><span class="impact-badge {imp_cls}">{imp:.1f} — {imp_lbl}</span></td>
          <td style="color:var(--muted)">{r["dup_pct"]}%</td>
          <td style="color:var(--muted)">{r["intg_pct"]}%</td>
        </tr>"""

    # ── Detail cards grouped by component ──────────────────────────────────
    by_comp = OrderedDict()
    for inv in interventions:
        comp = inv["component"] or "Outros"
        if comp not in by_comp:
            by_comp[comp] = []
        if inv["code"] in results:
            by_comp[comp].append(results[inv["code"]])

    cards_html = ""
    cmt_toggle_seq = 0
    for comp, items in by_comp.items():
        if not items:
            continue
        cards_html += f'<div class="comp-label">{esc(comp)}</div>\n<div class="priority-grid">\n'
        for r in items:
            sim_p  = round(r["n_sim"]  / r["n_total"] * 100) if r["n_total"] else 0
            poss_p = round(r["n_poss"] / r["n_total"] * 100) if r["n_total"] else 0
            nao_p  = round(r["n_nao"]  / r["n_total"] * 100) if r["n_total"] else 0
            imp    = r["avg_impact"]
            imp_cls = "mc-impact-high" if imp >= 2.5 else ("mc-impact-med" if imp >= 1.8 else "mc-impact-low")

            def make_tag(code, cls):
                lbl = inv_label.get(code, code)
                idx_ref = results.get(code, {}).get("display_idx", "?") if isinstance(results.get(code, {}), dict) else "?"
                return f'<span class="{cls}">{esc(str(idx_ref) + ". " + lbl)}</span>'

            dup_tags  = "".join(make_tag(t, "dup-tag")  for t in r["top_dup"])
            intg_tags = "".join(make_tag(t, "intg-tag") for t in r["top_intg"])
            dup_row  = (f'<div class="tag-row"><span class="tag-label dup-label">Duplicação com:</span>'
                        f'{dup_tags}</div>') if r["top_dup"]  else ""
            intg_row = (f'<div class="tag-row"><span class="tag-label">Integrar com:</span>'
                        f'{intg_tags}</div>') if r["top_intg"] else ""

            cmt_html = ""
            if r["comments"]:
                cmt_html = '<div class="comments-section"><div class="cmt-label">Sugestões dos especialistas (amostra)</div>'
                for c in r["comments"][:2]:
                    cmt_html += f'<div class="cmt-item">"{esc(c)}"</div>'
                if len(r["comments"]) > 2:
                    cmt_toggle_seq += 1
                    more_id = f'cmt-more-{cmt_toggle_seq}'
                    hidden_comments = "".join(
                        f'<div class="cmt-item">"{esc(c)}"</div>' for c in r["comments"][2:]
                    )
                    cmt_html += (
                        f'<button type="button" class="cmt-toggle" '
                        f'data-target="{more_id}" data-open="false">'
                        f'+{len(r["comments"]) - 2} comentário(s) adicionais</button>'
                        f'<div id="{more_id}" class="cmt-hidden">{hidden_comments}</div>'
                    )
                cmt_html += '</div>'

            missing_note = (f'<div class="missing-note">⚠ {r["n_missing"]} resposta(s) em falta / '
                            f'não aplicável</div>') if r["n_missing"] else ""

            label_with_idx = f'{r.get("display_idx", "-")}. {r["label"]}'
            name_cell = (f'<a href="{esc(r["url"])}" target="_blank" class="inv-link">{esc(label_with_idx)}</a>'
                         if r["url"] else esc(label_with_idx))

            cards_html += f"""<div class="inv-card">
              <div class="card-top">
                <span class="comp-tag">{esc(r["component"])}</span>
              </div>
              <h3>{name_cell}</h3>
              <div class="gate-label-sm">Precisa de optimização? (n={r["n_total"]})</div>
              <div class="gate-bar-wrap">
                <div class="gate-seg seg-sim"  style="width:{sim_p}%"></div>
                <div class="gate-seg seg-poss" style="width:{poss_p}%"></div>
                <div class="gate-seg seg-nao"  style="width:{nao_p}%"></div>
              </div>
              <div class="gate-counts">
                <div class="gc"><div class="dot" style="background:var(--sim)"></div><span class="val">{r["n_sim"]}</span><span class="gc-lbl">Sim def. ({sim_p}%)</span></div>
                <div class="gc"><div class="dot" style="background:#d4a017"></div><span class="val">{r["n_poss"]}</span><span class="gc-lbl">Possiv. ({poss_p}%)</span></div>
                <div class="gc"><div class="dot" style="background:#e0e0e0"></div><span class="val">{r["n_nao"]}</span><span class="gc-lbl">Não ({nao_p}%)</span></div>
              </div>
              <div class="metrics-row">
                <div class="metric-chip"><div class="mc-val {imp_cls}">{imp:.1f}<span class="mc-denom">/3</span></div><div class="mc-lbl">Impacto esperado</div></div>
                <div class="metric-chip"><div class="mc-val" style="color:var(--ink)">{r["dup_pct"]}%</div><div class="mc-lbl">Duplicação</div></div>
                <div class="metric-chip"><div class="mc-val" style="color:var(--accent2)">{r["intg_pct"]}%</div><div class="mc-lbl">Integração</div></div>
                <div class="metric-chip"><div class="mc-val" style="color:var(--muted)">{r["res_pct"]}%</div><div class="mc-lbl">↓ Recursos</div></div>
              </div>
              {dup_row}{intg_row}
              {cmt_html}
              {missing_note}
            </div>\n"""
        cards_html += "</div>\n"

    alta_str  = " · ".join(alta)  or "—"
    media_str = " · ".join(media) or "—"
    baixa_str = " · ".join(baixa) or "—"

    # ── Univariate analysis section ────────────────────────────────────────
    univ_html = ""
    scoring_html = ""
    if univariate:
        u = univariate
        GATE_COLORS = ("#1a6b3a", "#d4a017", "#e0e0e0")
        IMP_COLORS  = ("#e8f5ee", "#fef9e7", "#1a6b3a")
        EXP_COLORS  = ("#c7d2fe", "#818cf8", "#4f46e5")

        gate_donut = svg_donut(u["gate_agg"], GATE_COLORS, size=110, label="Gate")
        imp_donut  = svg_donut(
            {"Baixo (1)": u["imp_counts"].get(1,0),
             "Médio (2)": u["imp_counts"].get(2,0),
             "Alto (3)":  u["imp_counts"].get(3,0)},
            IMP_COLORS, size=110, label="Impacto"
        )
        exp_donut = svg_donut(
            {"General": u["exp_counts"].get(1,0),
             "Intermediate": u["exp_counts"].get(2,0),
             "Specialist": u["exp_counts"].get(3,0)},
            EXP_COLORS, size=110, label="Experiência"
        ) if any(u["exp_counts"].values()) else ""

        gate_bars = svg_hbar_stacked(
          [(f'{r.get("display_idx", "-")}. {r["label"]}'[:52], [r["n_sim"], r["n_poss"], r["n_nao"]])
             for r in u["sorted_gate"]],
            width=600, bar_h=14, gap=5, colors=GATE_COLORS, label_w=240
        )
        impact_bars = svg_hbar_single(
          [(f'{r.get("display_idx", "-")}. {r["label"]}'[:52], r["avg_impact"], 3.0) for r in u["sorted_impact"]],
            width=600, bar_h=14, gap=5, color="#1a6b3a", label_w=240, fmt=".2f"
        )
        dup_bars  = svg_hbar_single(
          [(f'{r.get("display_idx", "-")}. {r["label"]}'[:52], r["dup_pct"],  100) for r in u["sorted_dup"]],
            width=600, bar_h=14, gap=5, color="#9b2226", label_w=240, fmt=".0f"
        )
        intg_bars = svg_hbar_single(
          [(f'{r.get("display_idx", "-")}. {r["label"]}'[:52], r["intg_pct"], 100) for r in u["sorted_intg"]],
            width=600, bar_h=14, gap=5, color="#1a5276", label_w=240, fmt=".0f"
        )
        res_bars  = svg_hbar_single(
          [(f'{r.get("display_idx", "-")}. {r["label"]}'[:52], r["res_pct"],  100) for r in u["sorted_res"]],
            width=600, bar_h=14, gap=5, color="#5b5ea6", label_w=240, fmt=".0f"
        )
        scatter_oi = svg_scatter_optim_impact_exp(list(results.values())) if include_xyplot else ""
        xy_html = (f'<div class="univ-sub-header" style="margin-top:28px">2g · Dispersão em escala real (janela observada) com zonas da regra actual</div>'
             f'<div class="chart-wrap">{scatter_oi}</div>') if include_xyplot else ""

        univ_html = f"""
  <div class="section-header">
    <h2>Análise Univariada</h2>
    <span class="subtitle">Distribuições globais e por intervenção</span>
  </div>
  <div class="univ-panel">
    <div class="univ-row two-col">
      <div class="univ-col">
        <div class="univ-sub-header">2a · Distribuição da pergunta de triagem (Gate)</div>
        <div class="donut-row">{gate_donut}</div>
        <div class="chart-wrap">{gate_bars}</div>
      </div>
      <div class="univ-col">
        <div class="univ-sub-header">2b · Impacto esperado (escala 1–3)</div>
        <div class="donut-row">{imp_donut}</div>
        <div class="chart-wrap">{impact_bars}</div>
      </div>
    </div>

    {'<div class="univ-sub-header" style="margin-top:28px">2c · Experiência declarada dos respondentes</div><div class="donut-row">' + exp_donut + '</div>' if exp_donut else ''}

    <div class="univ-row two-col" style="margin-top:28px">
      <div class="univ-col">
        <div class="univ-sub-header">2d · % que identificou duplicação (por intervenção)</div>
        <div class="chart-wrap">{dup_bars}</div>
      </div>
      <div class="univ-col">
        <div class="univ-sub-header">2e · % que identificou potencial de integração</div>
        <div class="chart-wrap">{intg_bars}</div>
      </div>
    </div>

    <div class="univ-sub-header" style="margin-top:28px">2f · % que identificou possibilidade de redução de recursos</div>
    <div class="chart-wrap">{res_bars}</div>

    {xy_html}
  </div>
"""

        # ── Scoring section ──────────────────────────────────────────────────
        sorted_base = sorted(results.values(), key=lambda r: r["s_base"], reverse=True)
        sorted_pond = sorted(results.values(), key=lambda r: r["s_pond"], reverse=True)
        max_base = max((r["s_base"] for r in results.values()), default=1) or 1
        max_pond = max((r["s_pond"] for r in results.values()), default=1) or 1

        score_base_bars = svg_hbar_single(
            [(r["label"][:40], r["s_base"], max_base) for r in sorted_base],
            width=480, bar_h=14, gap=5, color="#1a6b3a", label_w=240, fmt=".3f"
        )
        score_pond_bars = svg_hbar_single(
            [(r["label"][:40], r["s_pond"], max_pond) for r in sorted_pond],
            width=480, bar_h=14, gap=5, color="#1a5276", label_w=240, fmt=".3f"
        )

        # Keep the flow readable by limiting to top interventions by S_base rank.
        _alluvial_cutoff = 20
        _all_by_base = sorted(results.values(), key=lambda r: r.get("rank_base", 10**9))
        alluvial_items = _all_by_base[:_alluvial_cutoff]
        alluvial_svg = svg_alluvial_weighting(alluvial_items)
        if len(results) > _alluvial_cutoff:
            _excl = ", ".join(f'{r.get("display_idx", "-")}. {esc(r["label"])}' for r in _all_by_base[_alluvial_cutoff:])
            alluvial_note = (f'<div style="font-size:11px;color:var(--muted);margin-top:4px">'
                            f'<em>Diagrama mostra as {_alluvial_cutoff} intervenções com maior S<sub>base</sub>'
                            f' (de {len(results)} no total). Exclu&#237;das: {_excl}.</em></div>')
        else:
            alluvial_note = ""

        score_table_rows = ""
        for r in sorted_base:
            delta = r.get("rank_delta", 0)
            delta_str = f"+{delta}" if delta > 0 else str(delta)
            delta_col = "#1a6b3a" if delta > 0 else ("#c0392b" if delta < 0 else "#6b7280")
            score_table_rows += f"""<tr>
              <td class="rank-num">{r.get("display_idx","")}</td>
              <td style="font-weight:500;font-size:12px">{esc(str(r.get("display_idx", "-")) + ". " + r["label"])}</td>
              <td style="text-align:center;font-weight:600;color:#7a5c00">{r.get("gate_mean", 0):.3f}</td>
              <td style="text-align:center;font-weight:600">{r["s_base"]:.3f}</td>
              <td style="text-align:center;font-weight:600;color:#1a5276">{r["s_pond"]:.3f}</td>
              <td style="text-align:center;font-weight:600;color:{delta_col}">{delta_str}</td>
              <td style="text-align:center;color:var(--muted)">{r["avg_impact"]:.2f}</td>
              <td style="text-align:center;color:var(--muted)">{r["exp_mean"]:.2f}</td>
            </tr>"""

        if include_scoring:
            scoring_html = f"""
  <div class="section-header">
    <h2>Pontuação e Metodologia</h2>
    <span class="subtitle">Duas métricas complementares para priorização</span>
  </div>
  <div class="formula-box">
    <div class="formula-title">Fórmulas de Pontuação</div>
    <div class="formula-grid">
      <div>
        <div class="formula-label">S<sub>base</sub> — Pontuação não ponderada (intervalo: 0–3)</div>
        <div class="formula-code">S<sub>base</sub> = <span class="f-mean">mean</span>(gate_score<sub>i</sub>) × <span class="f-mean">mean</span>(impact<sub>i</sub>)</div>
        <div class="formula-note">gate_score: sim_def = 1,0 · possivelmente = 0,5 · não = 0,0</div>
      </div>
      <div>
        <div class="formula-label">S<sub>pond</sub> — Pontuação ponderada por experiência (intervalo: 0–3)</div>
        <div class="formula-code">S<sub>pond</sub> = Σ(gate<sub>i</sub> × impact<sub>i</sub> × exp<sub>i</sub>) / Σ(exp<sub>i</sub>)</div>
        <div class="formula-note">exp<sub>i</sub> ∈ {{1, 2, 3}} — nível de experiência declarada (omisso = 1)</div>
      </div>
    </div>
  </div>
  <div class="score-charts-row">
    <div class="score-chart-col">
      <div class="score-chart-title" style="color:#1a6b3a">S<sub>base</sub> — não ponderada</div>
      {score_base_bars}
    </div>
    <div class="score-chart-col">
      <div class="score-chart-title" style="color:#1a5276">S<sub>pond</sub> — ponderada por experiência</div>
      {score_pond_bars}
    </div>
  </div>
  <div class="score-chart-col" style="margin-bottom:16px">
    <div class="score-chart-title" style="color:#0f1923">Diagrama Aluvial — optimizabilidade -> S_base -> S_pond</div>
    <div style="font-size:11px;color:var(--muted);margin-bottom:8px">
      1ª coluna: score de optimizabilidade (não ponderado). 2ª coluna: S<sub>base</sub>
      (optimizabilidade × impacto médio). 3ª coluna: S<sub>pond</sub>
      (ponderada por experiência declarada).
      Verde/vermelho/azul = subida/descida/sem mudança em ambas as transições.
    </div>
    <div class="chart-wrap">{alluvial_svg}</div>
    {alluvial_note}
  </div>
  <table id="score-table" class="rank-table" style="margin-top:16px">
    <thead>
      <tr>
        <th data-sort="num" style="cursor:pointer"># <span class="sa"></span></th>
        <th data-sort="text" style="cursor:pointer">Intervenção <span class="sa"></span></th>
        <th data-sort="num" style="text-align:center;cursor:pointer">Optimizabilidade <span class="sa"></span></th>
        <th data-sort="num" style="text-align:center;cursor:pointer">S<sub>base</sub> <span class="sa"></span></th>
        <th data-sort="num" style="text-align:center;cursor:pointer">S<sub>pond</sub> <span class="sa"></span></th>
        <th data-sort="num" style="text-align:center;cursor:pointer">Δ rank <span class="sa"></span></th>
        <th data-sort="num" style="text-align:center;cursor:pointer">Impacto médio <span class="sa"></span></th>
        <th data-sort="num" style="text-align:center;cursor:pointer">Exp. média <span class="sa"></span></th>
      </tr>
    </thead>
    <tbody>{score_table_rows}</tbody>
  </table>
"""
    else:
      scoring_html = ""

    html = f"""<!DOCTYPE html>
<html lang="pt">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Delphi W1 — Resultados Agregados</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Serif+Display:ital@0;1&family=DM+Sans:wght@300;400;500;600&display=swap" rel="stylesheet">
<style>
:root {{
  --ink:#0f1923; --paper:#f5f2ed; --accent:#c0392b; --accent2:#1a5276;
  --gold:#b7860b; --muted:#6b7280; --border:#d5cfc6;
  --sim:#1a6b3a; --sim-bg:#e8f5ee; --poss:#7a5c00; --poss-bg:#fef9e7;
  --card-bg:#ffffff;
}}
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'DM Sans',sans-serif;background:var(--paper);color:var(--ink);line-height:1.5;font-size:14px}}
.site-header{{background:var(--ink);color:white;padding:0 40px;display:flex;align-items:stretch;border-bottom:4px solid var(--accent)}}
.header-brand{{padding:24px 0;border-right:1px solid rgba(255,255,255,.12);padding-right:32px;margin-right:32px}}
.header-brand .eyebrow{{font-size:10px;letter-spacing:.2em;text-transform:uppercase;color:rgba(255,255,255,.5);margin-bottom:4px}}
.header-brand h1{{font-family:'DM Serif Display',serif;font-size:22px;font-weight:400;color:white;line-height:1.2}}
.header-meta{{padding:24px 0;display:flex;flex-direction:column;justify-content:center;gap:4px}}
.wave-badge{{display:inline-flex;align-items:center;gap:8px;background:var(--accent);color:white;font-size:11px;font-weight:600;letter-spacing:.08em;padding:3px 10px;border-radius:2px;width:fit-content;margin-bottom:6px}}
.header-meta p{{font-size:12px;color:rgba(255,255,255,.6)}}
.header-meta strong{{color:rgba(255,255,255,.9)}}
.intro-band{{background:var(--accent2);color:white;padding:20px 40px;display:flex;align-items:center;gap:20px}}
.intro-band .icon{{font-size:28px;flex-shrink:0}}
.intro-band p{{font-size:13px;line-height:1.6;color:rgba(255,255,255,.88)}}
.intro-band strong{{color:white}}
.summary-bar{{background:white;border-bottom:1px solid var(--border);padding:20px 40px;display:flex;gap:32px;align-items:center;flex-wrap:wrap}}
.stat-pill{{display:flex;flex-direction:column;align-items:center;gap:2px}}
.stat-pill .num{{font-family:'DM Serif Display',serif;font-size:32px;line-height:1;color:var(--accent)}}
.stat-pill .lbl{{font-size:10px;text-transform:uppercase;letter-spacing:.12em;color:var(--muted);text-align:center;max-width:80px}}
.stat-divider{{width:1px;height:40px;background:var(--border)}}
.legend-box{{margin-left:auto;display:flex;gap:16px;align-items:center;flex-wrap:wrap}}
.legend-item{{display:flex;align-items:center;gap:6px;font-size:11px;color:var(--muted)}}
.legend-dot{{width:10px;height:10px;border-radius:50%}}
.dot-sim{{background:var(--sim)}}.dot-poss{{background:#d4a017}}.dot-nao{{background:#d1d5db}}
.main{{padding:32px 40px;max-width:1300px;margin:0 auto}}
.section-header{{display:flex;align-items:baseline;gap:12px;margin-bottom:20px;padding-bottom:10px;border-bottom:2px solid var(--ink)}}
.section-header h2{{font-family:'DM Serif Display',serif;font-size:20px;font-weight:400}}
.section-header .subtitle{{font-size:12px;color:var(--muted);letter-spacing:.04em}}
.note-box{{background:#fffbeb;border:1px solid #fcd34d;border-left:4px solid var(--gold);border-radius:3px;padding:14px 18px;margin-bottom:28px;font-size:12px;color:#78350f;line-height:1.6}}
.note-box strong{{color:#92400e}}
/* ── Response rate section ── */
.rr-panel{{background:white;border:1px solid var(--border);border-radius:4px;margin-bottom:40px;overflow:hidden}}
.rr-summary{{display:flex;gap:0;border-bottom:1px solid var(--border)}}
.rr-stat{{flex:1;padding:16px 20px;text-align:center;border-right:1px solid var(--border)}}
.rr-stat:last-child{{border-right:none}}
.rr-stat .rr-num{{font-family:'DM Serif Display',serif;font-size:28px;line-height:1;color:var(--accent2)}}
.rr-stat .rr-lbl{{font-size:10px;text-transform:uppercase;letter-spacing:.1em;color:var(--muted);margin-top:2px}}
.rr-table{{width:100%;border-collapse:collapse;font-size:12px}}
.rr-table td{{padding:7px 16px;border-bottom:1px solid #f0ede8;vertical-align:middle}}
.rr-table tr:last-child td{{border-bottom:none}}
.rr-table tr:nth-child(even) td{{background:#fafaf8}}
.rr-details{{margin-top:12px}}
.rr-details summary{{cursor:pointer;padding:10px 16px;background:#f7f9fb;border-top:1px solid var(--border);font-size:12px;font-weight:600;color:var(--accent2);user-select:none;display:flex;align-items:center;gap:8px}}
.rr-details summary:hover{{background:#eef2f7}}
.rr-details summary::marker{{content:''}}
.rr-details summary::before{{content:'▶';font-size:10px;transition:transform 0.2s}}
.rr-details[open] summary::before{{transform:rotate(90deg)}}
/* ── Ranking table ── */
.rank-table{{width:100%;border-collapse:collapse;margin-bottom:40px;background:white;border:1px solid var(--border);border-radius:4px;overflow:hidden;font-size:12px}}
.rank-table thead tr{{background:var(--ink);color:white}}
.rank-table th{{padding:10px 14px;font-size:10px;text-transform:uppercase;letter-spacing:.1em;font-weight:600;text-align:left}}
.rank-table td{{padding:9px 14px;border-bottom:1px solid var(--border);vertical-align:middle}}
.rank-table tr:last-child td{{border-bottom:none}}
.rank-table tr:nth-child(even) td{{background:#fafafa}}
.rank-num{{font-family:'DM Serif Display',serif;font-size:18px;color:var(--muted);width:32px}}
.comp-tag{{font-size:10px;background:var(--paper);border:1px solid var(--border);color:var(--muted);padding:2px 6px;border-radius:2px;white-space:nowrap}}
.inv-link{{color:inherit;text-decoration:none;border-bottom:1px dotted var(--muted)}}
.inv-link:hover{{color:var(--accent2);border-bottom-color:var(--accent2)}}
.mini-bar-wrap{{display:flex;align-items:center;gap:8px}}
.mini-bar-bg{{flex:1;height:8px;background:#e5e7eb;border-radius:2px;overflow:hidden}}
.bar-counts{{font-size:10px;color:var(--muted);white-space:nowrap}}
.impact-badge{{display:inline-flex;align-items:center;gap:4px;font-size:11px;font-weight:600;padding:3px 8px;border-radius:2px}}
.imp-high{{background:var(--sim-bg);color:var(--sim)}}
.imp-med{{background:var(--poss-bg);color:var(--poss)}}
.imp-low{{background:#f3f4f6;color:var(--muted)}}
/* ── Cards ── */
.comp-label{{font-size:11px;font-weight:700;letter-spacing:.12em;text-transform:uppercase;color:var(--accent2);padding:6px 0;border-bottom:1px solid var(--border);margin-bottom:14px;margin-top:28px}}
.priority-grid{{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:16px}}
.inv-card{{background:var(--card-bg);border:1px solid var(--border);border-radius:4px;padding:16px 18px}}
.inv-card:hover{{box-shadow:0 4px 16px rgba(0,0,0,.08)}}
.card-top{{display:flex;align-items:flex-start;justify-content:flex-end;gap:8px;margin-bottom:10px}}
.inv-card h3{{font-size:13px;font-weight:600;line-height:1.3;margin-bottom:12px}}
.gate-label-sm{{font-size:10px;text-transform:uppercase;letter-spacing:.1em;color:var(--muted);margin-bottom:5px}}
.gate-bar-wrap{{display:flex;height:12px;border-radius:2px;overflow:hidden;gap:1px;margin-bottom:4px}}
.gate-seg{{height:100%}}
.seg-sim{{background:var(--sim)}}.seg-poss{{background:#d4a017}}.seg-nao{{background:#e0e0e0}}
.gate-counts{{display:flex;gap:12px;margin-bottom:2px}}
.gc{{display:flex;align-items:center;gap:4px;font-size:10px}}
.gc .dot{{width:7px;height:7px;border-radius:50%}}
.gc .val{{font-weight:600}}
.gc-lbl{{color:var(--muted)}}
.metrics-row{{display:flex;gap:8px;margin-top:12px;padding-top:12px;border-top:1px solid var(--border)}}
.metric-chip{{flex:1;background:var(--paper);border-radius:3px;padding:6px 8px;text-align:center}}
.mc-val{{font-family:'DM Serif Display',serif;font-size:16px;line-height:1;margin-bottom:2px}}
.mc-denom{{font-size:10px;font-family:'DM Sans',sans-serif}}
.mc-lbl{{font-size:9px;text-transform:uppercase;letter-spacing:.08em;color:var(--muted)}}
.mc-impact-high{{color:var(--sim)}}.mc-impact-med{{color:var(--poss)}}.mc-impact-low{{color:#9ca3af}}
.tag-row{{margin-top:8px;display:flex;flex-wrap:wrap;gap:4px;align-items:center}}
.tag-label{{font-size:9px;text-transform:uppercase;letter-spacing:.08em;color:var(--muted);margin-right:2px}}
.dup-label{{color:#9b2226}}
.intg-tag{{font-size:10px;font-weight:600;background:#e8f0fe;color:#1a3a8f;padding:2px 7px;border-radius:2px;white-space:normal;line-height:1.4}}
.dup-tag{{font-size:10px;font-weight:600;background:#fde8e8;color:#9b2226;padding:2px 7px;border-radius:2px;white-space:normal;line-height:1.4}}
.comments-section{{margin-top:10px;padding-top:10px;border-top:1px solid var(--border)}}
.cmt-label{{font-size:9px;text-transform:uppercase;letter-spacing:.08em;color:var(--muted);margin-bottom:5px}}
.cmt-item{{font-size:11px;color:var(--ink);font-style:italic;line-height:1.5;margin-bottom:4px;padding-left:8px;border-left:2px solid var(--border)}}
.cmt-toggle{{margin-top:3px;padding:0;border:0;background:none;font-size:10px;color:var(--muted);text-decoration:underline;cursor:pointer}}
.cmt-toggle:hover{{color:var(--ink)}}
.cmt-hidden{{display:none;margin-top:4px}}
.cmt-hidden.is-open{{display:block}}
.missing-note{{font-size:10px;color:#b45309;margin-top:6px;background:#fffbeb;padding:3px 6px;border-radius:2px}}
/* ── Next steps ── */
.next-box{{background:white;border:1px solid var(--border);border-radius:4px;padding:24px;margin-bottom:40px}}
.next-box p{{font-size:13px;line-height:1.8}}
.tier-panels{{margin-top:16px;padding-top:16px;border-top:1px solid var(--border);display:flex;gap:16px;flex-wrap:wrap}}
.tier-panel{{flex:1;min-width:180px;border-radius:3px;padding:12px 16px}}
.tier-alta{{background:var(--sim-bg)}}.tier-media{{background:var(--poss-bg)}}.tier-baixa{{background:#f3f4f6}}
.tier-heading{{font-size:10px;text-transform:uppercase;letter-spacing:.1em;font-weight:700;margin-bottom:6px}}
.tier-alta .tier-heading{{color:var(--sim)}}.tier-media .tier-heading{{color:var(--poss)}}.tier-baixa .tier-heading{{color:var(--muted)}}
.tier-panel p{{font-size:12px;line-height:1.8}}
.report-footer{{background:var(--ink);color:rgba(255,255,255,.5);padding:20px 40px;font-size:11px;display:flex;justify-content:space-between;align-items:center;margin-top:40px;flex-wrap:wrap;gap:8px}}
.report-footer strong{{color:rgba(255,255,255,.8)}}
@media print{{body{{background:white}}.site-header,.intro-band{{-webkit-print-color-adjust:exact;print-color-adjust:exact}}.inv-card{{break-inside:avoid}}}}
/* ── Univariate ── */
.univ-panel{{background:white;border:1px solid var(--border);border-radius:4px;padding:24px;margin-bottom:40px}}
.univ-sub-header{{font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.1em;color:var(--accent2);margin-bottom:10px}}
.univ-row.two-col{{display:grid;grid-template-columns:1fr 1fr;gap:24px}}
.univ-col{{min-width:0}}
.donut-row{{margin-bottom:12px}}
.chart-wrap{{overflow-x:auto;padding-bottom:4px}}
/* ── Scoring ── */
.formula-box{{background:#0f1923;color:white;border-radius:4px;padding:20px 24px;margin-bottom:20px}}
.formula-title{{font-size:11px;text-transform:uppercase;letter-spacing:.12em;color:rgba(255,255,255,.5);margin-bottom:12px}}
.formula-grid{{display:grid;grid-template-columns:1fr 1fr;gap:24px}}
.formula-label{{font-size:11px;color:rgba(255,255,255,.6);margin-bottom:6px}}
.formula-code{{font-family:monospace;font-size:14px;color:#86efac;margin-bottom:4px}}
.formula-note{{font-size:10px;color:rgba(255,255,255,.4)}}
.score-charts-row{{display:grid;grid-template-columns:1fr 1fr;gap:24px;margin-bottom:16px}}
.score-chart-col{{background:white;border:1px solid var(--border);border-radius:4px;padding:16px;overflow-x:auto}}
.score-chart-title{{font-size:12px;font-weight:600;margin-bottom:10px}}
@media (max-width: 980px){{
  .univ-row.two-col{{grid-template-columns:1fr}}
}}
</style>
</head>
<body>

<header class="site-header">
  <div class="header-brand">
    <div class="eyebrow">Programa Nacional de Controlo da Malária</div>
    <h1>Exercício de Optimização<br>de Intervenções</h1>
  </div>
  <div class="header-meta">
    <div class="wave-badge">⬤ &nbsp;ONDA 1 — RESULTADOS AGREGADOS</div>
    <p>Gerado em: <strong>{now}</strong> &nbsp;|&nbsp; Fonte: <strong>{esc(os.path.basename(source_file))}</strong></p>
    <p>Base esperada: <strong>{stats["n_experts"]} especialistas</strong> (experts.txt) &nbsp;|&nbsp; Respondentes observados: <strong>{n_experts_observed}</strong> &nbsp;|&nbsp; Intervenções avaliadas: <strong>{stats["n_inv"]}</strong></p>
    <p>Este documento é <strong>confidencial</strong> — para uso exclusivo dos participantes da oficina Delphi</p>
  </div>
</header>

<div class="intro-band">
  <div class="icon">📋</div>
  <p>Este relatório apresenta os <strong>resultados anónimos e agregados</strong> da primeira onda (W1).
  Os dados reflectem as avaliações colectivas dos especialistas sobre o potencial de optimização de cada intervenção.
  <strong>Nenhuma resposta individual é identificável.</strong>
  Com base nestes resultados, o grupo seleccionará as intervenções prioritárias para a Onda 2 (W2).</p>
</div>

<div class="summary-bar">
  <div class="stat-pill"><span class="num">{stats["n_experts"]}</span><span class="lbl">Especialistas esperados</span></div>
  <div class="stat-divider"></div>
  <div class="stat-pill"><span class="num">{n_experts_observed}</span><span class="lbl">Especialistas respondentes</span></div>
  <div class="stat-divider"></div>
  <div class="stat-pill"><span class="num">{stats["n_inv"]}</span><span class="lbl">Intervenções avaliadas</span></div>
  <div class="stat-divider"></div>
  <div class="stat-pill"><span class="num">{stats["n_inv_80"]}</span><span class="lbl">Com ≥80% a favor de optimização</span></div>
  <div class="stat-divider"></div>
  <div class="stat-pill"><span class="num">{stats["n_imp_high"]}</span><span class="lbl">Com impacto esperado alto (≥2,5)</span></div>
  <div class="stat-divider"></div>
  <div class="stat-pill"><span class="num">{stats["n_unanimous"]}</span><span class="lbl">Com consenso unânime (100%)</span></div>
  <div class="legend-box">
    <div class="legend-item"><div class="legend-dot dot-sim"></div>Sim, definitivamente</div>
    <div class="legend-item"><div class="legend-dot dot-poss"></div>Possivelmente</div>
    <div class="legend-item"><div class="legend-dot dot-nao"></div>Não</div>
  </div>
</div>

<div class="main">

  <div class="note-box">
    <strong>Como ler este relatório:</strong> Para cada intervenção são apresentados:
    (1) a distribuição de respostas à pergunta de triagem ("precisa de optimização?"),
    (2) o impacto médio esperado se optimizada (escala 1–3),
    e (3) as percentagens de especialistas que identificaram duplicação, potencial de integração
    e possibilidade de redução de recursos.
    As intervenções estão ordenadas por <em>pontuação composta</em> (% "Sim definitivamente" × impacto médio).
    S = Sim definitivamente &nbsp;|&nbsp; P = Possivelmente &nbsp;|&nbsp; N = Não.
    Os nomes das intervenções são hiperligações para as respectivas fichas de descrição.
  </div>

  <!-- ═══ RESPONSE RATE ═══════════════════════════════════════════════════ -->
  <div class="section-header">
    <h2>Taxa de Resposta por Intervenção</h2>
    <span class="subtitle">Proporção de especialistas que avaliaram cada intervenção</span>
  </div>

  <div class="rr-panel">
    <div class="rr-summary">
      <div class="rr-stat">
        <div class="rr-num">{stats["rr_median"]}%</div>
        <div class="rr-lbl">Mediana</div>
      </div>
      <div class="rr-stat">
        <div class="rr-num">{stats["rr_min"]}%</div>
        <div class="rr-lbl">Mínimo</div>
      </div>
      <div class="rr-stat">
        <div class="rr-num">{stats["rr_max"]}%</div>
        <div class="rr-lbl">Máximo</div>
      </div>
      <div class="rr-stat" style="flex:2;text-align:left;padding-left:24px">
        <div style="font-size:12px;color:var(--muted);line-height:1.6">
          A taxa de resposta por intervenção varia porque o formulário W1 está dividido
          em grupos de intervenções, e nem todos os especialistas completaram todos os grupos.
          Intervenções com taxa abaixo de 60% devem ser interpretadas com cautela.
        </div>
      </div>
    </div>
    <details class="rr-details">
      <summary>Ver detalhes por intervenção ({stats["n_inv"]} intervenções)</summary>
      <table class="rr-table">
        <tbody>{rr_rows}</tbody>
      </table>
    </details>
  </div>

  {univ_html}
  {scoring_html}

  <!-- ═══ RANKING TABLE ════════════════════════════════════════════════════ -->
  <div class="section-header">
    <h2>Tabela de Prioridade — Todas as Intervenções</h2>
    <span class="subtitle">Ordenadas por pontuação composta</span>
  </div>

  <table class="rank-table">
    <thead>
      <tr>
        <th>#</th><th>Intervenção</th><th>Componente</th>
        <th style="width:200px">Distribuição de respostas</th>
        <th>% Optimizável</th><th>Impacto esperado</th>
        <th>% Duplicação</th><th>% Integração</th>
      </tr>
    </thead>
    <tbody>{rank_rows}</tbody>
  </table>

  <!-- ═══ DETAIL CARDS ════════════════════════════════════════════════════ -->
  <div class="section-header">
    <h2>Detalhe por Intervenção</h2>
    <span class="subtitle">Agrupado por componente programático</span>
  </div>

  {cards_html}

  <!-- ═══ NEXT STEPS ══════════════════════════════════════════════════════ -->
  <div class="section-header">
    <h2>Próximos Passos — Onda 2</h2>
  </div>
  <div class="next-box">
    <p>Com base nos resultados da W1, o grupo irá <strong>seleccionar colectivamente as intervenções a focar na W2</strong>.
    Recomenda-se priorizar intervenções estritamente acima da mediana simultaneamente em optimizabilidade e impacto esperado
    (cutoffs dinâmicos nesta ronda: gate_mean &gt; {med_gate:.3f}; impacto &gt; {med_imp:.3f}).
    As intervenções seleccionadas serão distribuídas por grupos de trabalho temáticos que irão desenvolver <strong>propostas concretas de optimização</strong>.
    Essas propostas serão submetidas à avaliação anónima colectiva na Onda 3 (W3).</p>
    <div class="tier-panels">
      <div class="tier-panel tier-alta">
        <div class="tier-heading">Candidatas de Alta Prioridade (N={len(alta)})</div>
        <p>{esc(alta_str)}</p>
      </div>
      <div class="tier-panel tier-media">
        <div class="tier-heading">Candidatas de Prioridade Média (N={len(media)})</div>
        <p>{esc(media_str)}</p>
      </div>
      <div class="tier-panel tier-baixa">
        <div class="tier-heading">Baixa Prioridade / Manter (N={len(baixa)})</div>
        <p>{esc(baixa_str)}</p>
      </div>
    </div>
  </div>

</div>

<script>
// Sortable table
(function() {{
  var tbl = document.getElementById('score-table');
  if (!tbl) return;
  var sortCol = -1, sortAsc = true;
  var ths = tbl.querySelectorAll('thead th');
  ths.forEach(function(th, idx) {{
    th.addEventListener('click', function() {{
      if (sortCol === idx) {{ sortAsc = !sortAsc; }}
      else {{ sortCol = idx; sortAsc = (th.getAttribute('data-sort') !== 'text'); }}
      var tbody = tbl.querySelector('tbody');
      var rows = Array.from(tbody.rows);
      rows.sort(function(a, b) {{
        var ta = a.cells[idx].textContent.trim();
        var tb = b.cells[idx].textContent.trim();
        var va, vb;
        if (th.getAttribute('data-sort') === 'text') {{
          va = ta.toLowerCase(); vb = tb.toLowerCase();
        }} else {{
          va = parseFloat(ta.replace(/^\+/, '')) || 0;
          vb = parseFloat(tb.replace(/^\+/, '')) || 0;
        }}
        if (va < vb) return sortAsc ? -1 : 1;
        if (va > vb) return sortAsc ? 1 : -1;
        return 0;
      }});
      rows.forEach(function(r) {{ tbody.appendChild(r); }});
      ths.forEach(function(h) {{ var s = h.querySelector('.sa'); if (s) s.textContent = ''; }});
      var sa = th.querySelector('.sa');
      if (sa) sa.textContent = sortAsc ? ' ▲' : ' ▼';
    }});
  }});
}})();

document.addEventListener('click', function(evt) {{
  const btn = evt.target.closest('.cmt-toggle');
  if (!btn) return;
  const targetId = btn.getAttribute('data-target');
  if (!targetId) return;
  const panel = document.getElementById(targetId);
  if (!panel) return;

  const isOpen = btn.getAttribute('data-open') === 'true';
  panel.classList.toggle('is-open', !isOpen);
  btn.setAttribute('data-open', isOpen ? 'false' : 'true');

  if (isOpen) {{
    btn.textContent = btn.textContent.replace(/^−[ \t]*/, '+');
  }} else {{
    btn.textContent = btn.textContent.replace(/^[+][ \t]*/, '− ');
  }}
}});
</script>

<div class="report-footer">
  <span>Programa Nacional de Controlo da Malária — INS &nbsp;|&nbsp; Exercício Delphi 2026 &nbsp;|&nbsp; <strong>CONFIDENCIAL — Uso interno</strong></span>
  <span>Resultados anónimos. Nenhuma resposta individual é identificável. &nbsp;|&nbsp; Gerado em {now}</span>
</div>

</body>
</html>"""
    return html

# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(0)

    results_path = sys.argv[1]
    dict_path    = None
    output_dir   = None
    exclude_codes_arg = None
    exclude_file = None
    config_path = None
    simple_sections = False

    # Parse optional arguments
    i = 2
    while i < len(sys.argv):
      arg = sys.argv[i]

      if arg in ("--output-dir", "-o"):
        if i + 1 >= len(sys.argv):
          sys.exit("Erro: --output-dir requer um valor.")
        output_dir = sys.argv[i + 1]
        i += 2
        continue
      if arg.startswith("--output-dir="):
        output_dir = arg.split("=", 1)[1]
        i += 1
        continue

      if arg in ("--exclude-experts", "-x"):
        if i + 1 >= len(sys.argv):
          sys.exit("Erro: --exclude-experts requer um valor.")
        exclude_codes_arg = sys.argv[i + 1]
        i += 2
        continue
      if arg.startswith("--exclude-experts="):
        exclude_codes_arg = arg.split("=", 1)[1]
        i += 1
        continue

      if arg in ("--exclude-file", "-X"):
        if i + 1 >= len(sys.argv):
          sys.exit("Erro: --exclude-file requer um caminho.")
        exclude_file = sys.argv[i + 1]
        i += 2
        continue
      if arg.startswith("--exclude-file="):
        exclude_file = arg.split("=", 1)[1]
        i += 1
        continue

      if arg == "--config":
        if i + 1 >= len(sys.argv):
          sys.exit("Erro: --config requer um caminho.")
        config_path = sys.argv[i + 1]
        i += 2
        continue
      if arg.startswith("--config="):
        config_path = arg.split("=", 1)[1]
        i += 1
        continue

      if arg in ("--simple-sections", "--compact-report"):
        simple_sections = True
        i += 1
        continue

      if arg.startswith("-"):
        print(f"Aviso: argumento não reconhecido (ignorado): {arg}")
        i += 1
        continue

      if dict_path is None:
        dict_path = arg
      else:
        print(f"Aviso: argumento posicional extra ignorado: {arg}")
      i += 1

    # Default output directory: ./reports relative to current working directory
    if output_dir is None:
        output_dir = "reports"
    output_dir = os.path.abspath(output_dir)

    if not os.path.exists(results_path):
        sys.exit(f"Erro: ficheiro de resultados não encontrado: {results_path}")
    if dict_path and not os.path.exists(dict_path):
        print(f"Aviso: dicionário não encontrado: {dict_path} — continuando sem ele.")
        dict_path = None

    expected_experts = load_expected_experts()
    n_expected_experts = len(expected_experts)

    # Ensure output directory exists
    os.makedirs(output_dir, exist_ok=True)

    print(f"A carregar resultados: '{results_path}'...")
    df = load_data(results_path, "Responses")
    print(f"  {len(df)} linhas · {df['expert_code'].nunique()} especialistas únicos")
    if n_expected_experts > 0:
      print(f"  Base esperada (experts.txt): {n_expected_experts} especialistas")
    else:
      print("  Aviso: experts.txt não encontrado ou vazio; usando especialistas observados.")

    print("A carregar metadados das intervenções...")
    interventions = load_metadata(results_path, dict_path)
    if interventions is None:
        df_codes = sorted(df["intervention"].dropna().unique())
        interventions = [{"code": c, "label": c, "component": "", "url": ""} for c in df_codes]
        print(f"  Metadados: códigos brutos da folha Responses ({len(interventions)} intervenções)")

    excluded_codes = load_excluded_codes(
      cli_codes=exclude_codes_arg,
      exclude_file=exclude_file,
      config_path=config_path,
    )
    if excluded_codes:
      excluded_lc = {c.lower() for c in excluded_codes}

      before_rows = len(df)
      before_experts = df["expert_code"].nunique()
      before_expert_codes = {
        str(v).strip().lower()
        for v in df["expert_code"].dropna().astype(str).tolist()
        if str(v).strip()
      }

      df = df[
        ~df["expert_code"].fillna("").astype(str).str.strip().str.lower().isin(excluded_lc)
      ].copy()
      after_experts = df["expert_code"].nunique()

      matched = sorted(c for c in excluded_codes if c.lower() in before_expert_codes)
      unmatched = sorted(c for c in excluded_codes if c.lower() not in before_expert_codes)

      print(f"  Exclusões activas: {len(excluded_codes)} especialista(s)")
      print(f"    Removidas {before_rows - len(df)} linha(s) de respostas")
      print(f"    Removidos {before_experts - after_experts} especialista(s)")
      if matched:
        print("    Especialistas excluídos encontrados: " + ", ".join(matched))
      if unmatched:
        print("    Aviso: especialistas não encontrados (ignorados): " + ", ".join(unmatched))

    if not interventions:
      sys.exit("Erro: não restaram intervenções após aplicar as exclusões.")

    print("A agregar respostas...")
    results = aggregate(df, interventions)
    compute_ranks(results)
    stats   = summary_stats(results, df, n_expected_experts)

    print(f"  {stats['n_experts']} especialistas · {stats['n_inv']} intervenções")
    print(f"  Taxa de resposta: mediana={stats['rr_median']}% "
          f"intervalo=[{stats['rr_min']}%–{stats['rr_max']}%]")

    # Console ranking table
    sorted_base = sorted(results.values(), key=lambda r: r["s_base"], reverse=True)
    print(f"\n  {'Rank':>4}  {'S_base':>7}  {'S_pond':>7}  {'Δrank':>6}  Intervenção")
    print(f"  {'----':>4}  {'-------':>7}  {'-------':>7}  {'------':>6}  -----------")
    for r in sorted_base:
        delta = r.get("rank_delta", 0)
        delta_str = f"+{delta}" if delta > 0 else str(delta)
        print(f"  {r['rank_base']:>4}  {r['s_base']:>7.3f}  {r['s_pond']:>7.3f}  {delta_str:>6}  {r['label'][:52]}")

    print("\nA calcular estatísticas univariadas...")
    univariate = univariate_analysis(results, df)

    print("A gerar relatório HTML...")
    html = render_html(
      results,
      stats,
      interventions,
      results_path,
      univariate,
      include_xyplot=not simple_sections,
      include_scoring=not simple_sections,
    )

    ts       = datetime.now().strftime("%Y%m%d_%H%M")
    out_path = os.path.join(output_dir, f"delphi_w1_relatorio_{ts}.html")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"\n✓ Relatório guardado em: {out_path}")

if __name__ == "__main__":
    main()
