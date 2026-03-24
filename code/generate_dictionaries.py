#!/usr/bin/env python3
"""
generate_dictionaries.py
========================
Generates Delphi W1 data dictionaries for TB, HIV/SIDA, and SMI programs.
Also supports Malária for completeness.

Each dictionary is a 4-sheet Excel file:
  Catalogo_<Program>    — enriched catalogue (code, URL, group, all fields)
  Dicionario_Perguntas  — KoboToolbox questionnaire template (identical across programs)
  Listas_de_Opcoes      — option lists (expertise, gate, impact, codes, interventions)
  README                — structure documentation

Usage:
    # Generate all programs:
    python generate_dictionaries.py --all

    # Generate one program:
    python generate_dictionaries.py --program tb     --catalog path/to/TB.xlsx
    python generate_dictionaries.py --program hiv    --catalog path/to/HIV.xlsx
    python generate_dictionaries.py --program smi    --catalog path/to/SMI.xlsx
    python generate_dictionaries.py --program malaria --catalog path/to/Malaria.xlsx

    # Optional: set output directory
    python generate_dictionaries.py --all --output-dir ./dicionarios

Notes on catalog structure differences
---------------------------------------
TB     : Columns identical to Malária. Row 3 contains a template example row
         (Área = "Exemplo - HIV") which is skipped automatically.
         14 real interventions.

HIV    : Different structure — no "Área" column; uses "Actividade" instead of
         "Intervenção"; has an "Implementador" column; duplicate "Gastos" columns
         (two sub-columns + "Total Gastos 2024"). Programme = col B (not col A).
         177 rows covering activities × implementers across 13 programme areas.
         The dictionary preserves this granularity; the intervention label is
         "Actividade" and Implementador is included as an additional field.

SMI    : Column structure identical to Malária except "Descrição" (with accent)
         instead of "Descricao". 89 interventions across 14 programme areas.
"""

import sys
import os
import math
import pathlib
import unicodedata
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Install openpyxl:  pip install openpyxl")

# ─────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────

def load_config(path="config.env"):
    cfg = {}
    cwd_path   = pathlib.Path.cwd() / path
    script_dir = pathlib.Path(__file__).parent / path
    config_path = cwd_path if cwd_path.exists() else (script_dir if script_dir.exists() else None)
    if not config_path:
        return cfg
    with open(config_path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, _, v = line.partition("=")
            cfg[k.strip()] = v.strip()
    return cfg

_cfg = load_config()

def _get(key, default=""):
    return _cfg.get(key, default)

EXPERT_CODE_SUFFIX = _get("EXPERT_CODE_SUFFIX", "PM")
EXPERT_CODE_COUNT  = int(_get("EXPERT_CODE_COUNT", "25") or "25")
ADMIN_CODE_SUFFIX  = _get("ADMIN_CODE_SUFFIX",  "XX")
ADMIN_CODE_COUNT   = int(_get("ADMIN_CODE_COUNT",  "15") or "15")

# ─────────────────────────────────────────────────────────────────
# PROGRAM CONFIGURATIONS
# ─────────────────────────────────────────────────────────────────

GROUP_SIZE  = 7

PROGRAMS = {
    "malaria": {
        "topic_label":   "Malária",
        "code_prefix":   "mal",
        "base_url":      "https://kaiserso.github.io/INS_delphi/malaria",
        "catalog_sheet": "Malária",
        "key_field":     "intervencao",   # internal key used as label
        "col_map": {
            "Área":                          "area",
            "Programa":                      "programa",
            "Componente":                    "componente",
            "Intervenção":                   "intervencao",
            "Nível":                         "nivel",
            "Descricao (o que inclui)":      "descricao",
            "Objectivo(s)":                  "objectivos",
            "Alcance geográfico da intervenção":                "alcance",
            "Recursos necessários para a implementação":        "recursos",
            "Etapas chave para a implementação da intervenção": "etapas",
            "Descrição dos riscos e limitações que compromentem a implementação da intervenção": "riscos",
            "Possiveis factores associados aos riscos e limitações descritas": "factores",
            "Ano de início":                 "ano_inicio",
            "Gastos em 2024 (MZN)":          "gastos_2024",
            "Fonte(s) de financiamento":     "fontes",
            "Pop elegivel":                  "pop_elegivel",
            "Numero alcancado":              "num_alcancado",
            "Cobertura":                     "cobertura",
            "Custo por unidade":             "custo_unidade",
            "Notas":                         "notas",
        },
        "output_headers": [
            "Código","URL da Ficha","Grupo",
            "Área","Programa","Componente","Intervenção",
            "Nível","Descrição (o que inclui)","Objectivo(s)",
            "Alcance geográfico","Recursos necessários","Etapas chave",
            "Riscos e limitações","Possíveis factores associados",
            "Ano de início","Gastos em 2024 (MZN)","Fonte(s) de financiamento",
            "Pop. elegível","Número alcançado","Cobertura","Custo por unidade","Notas",
        ],
        "col_widths": [10,45,22,10,18,20,45,14,40,40,30,35,35,35,35,10,16,20,14,14,10,25,35],
        "skip_row_if": lambda row: False,  # no rows to skip
        "row_builder": "standard",
    },
    "tb": {
        "topic_label":   "TB",
        "code_prefix":   "tb",
        "base_url":      _get("BASE_URL", ""),
        "catalog_sheet": "TB",
        "key_field":     "intervencao",
        "col_map": {
            "Área":                      "area",
            "Programa":                  "programa",
            "Componente":                "componente",
            "Intervenção":               "intervencao",
            "Nível":                     "nivel",
            "Descricao (o que inclui)":  "descricao",
            "Objectivo(s)":              "objectivos",
            "Alcance geográfico da intervenção":                "alcance",
            "Recursos necessários para a implementação":        "recursos",
            "Etapas chave para a implementação da intervenção": "etapas",
            "Descrição dos riscos e limitações que compromentem a implementação da intervenção": "riscos",
            "Possiveis factores associados aos riscos e limitações descritas": "factores",
            "Ano de início":             "ano_inicio",
            "Gastos em 2024":            "gastos_2024",
            "Fonte(s) de financiamento": "fontes",
            "Pop elegivel":              "pop_elegivel",
            "Numero alcancado":          "num_alcancado",
            "Cobertura":                 "cobertura",
            "Custo por unidade":         "custo_unidade",
            "Notas":                     "notas",
            # "Notas - revisao" excluded
        },
        "output_headers": [
            "Código","URL da Ficha","Grupo",
            "Área","Programa","Componente","Intervenção",
            "Nível","Descrição (o que inclui)","Objectivo(s)",
            "Alcance geográfico","Recursos necessários","Etapas chave",
            "Riscos e limitações","Possíveis factores associados",
            "Ano de início","Gastos em 2024","Fonte(s) de financiamento",
            "Pop. elegível","Número alcançado","Cobertura","Custo por unidade","Notas",
        ],
        "col_widths": [10,45,22,10,18,30,55,14,45,45,30,35,35,35,35,12,18,22,14,14,12,25,40],
        # Skip the template example row (Área = "Exemplo - HIV")
        "skip_row_if": lambda row: str(row.get("area","")).startswith("Exemplo"),
        "row_builder": "tb",
    },
    "hiv": {
        "topic_label":   "HIV/SIDA",
        "code_prefix":   "hiv",
        "base_url":      "https://kaiserso.github.io/INS_delphi/hiv",
        "catalog_sheet": "HIV SIDA",
        "key_field":     "actividade",   # "Actividade" used as intervention label
        "col_map": {
            # Note: col A is empty; Programa starts at col B
            "Programa":                      "programa",
            "Componente":                    "componente",
            "Actividade":                    "actividade",
            "Actividade/ Descricao":         "actividade",  # variant in new file
            "Implementador":                 "implementador",
            "Nível":                         "nivel",
            "Descricao (o que inclui)":      "descricao",
            "Objectivo(s)":                  "objectivos",
            "Alcance geográfico da intervenção":                "alcance",
            "Recursos necessários para a implementação":        "recursos",
            "Etapas chave para a implementação da intervenção": "etapas",
            "Descrição dos riscos e limitações que compromentem a implementação da intervenção": "riscos",
            "Possiveis factores associados aos riscos e limitações descritas": "factores",
            "Ano de início":                 "ano_inicio",
            "Total Gastos 2024":             "gastos_2024",   # use the total column
            "Pop elegivel":                  "pop_elegivel",
            "fonte de eligibilidade":        "fonte_eligibilidade",
            "Numero alcancado (Dez 2024)":   "num_alcancado",
            "Cobertura":                     "cobertura",
            "Custo por unidade":             "custo_unidade",
            "Num de US com implementatcao (Dez 2024)": "num_us",
            "Notas":                         "notas",
            # "Nota Geral" and "Revisao" excluded
        },
        "output_headers": [
            "Código","URL da Ficha","Grupo",
            "Programa","Componente","Actividade","Implementador",
            "Nível","Descrição (o que inclui)","Objectivo(s)",
            "Alcance geográfico da intervenção",
            "Recursos necessários para a implementação",
            "Etapas chave para a implementação da intervenção",
            "Descrição dos riscos e limitações que comprometem a implementação da intervenção",
            "Possiveis factores associados aos riscos e limitações descritas",
            "Ano de início","Gastos em 2024 (USD)","Pop. elegível",
            "Fonte de elegibilidade","Número alcançado (Dez 2024)","Cobertura",
            "Custo por unidade","Nº US com implementação","Notas",
        ],
        "col_widths": [10,45,22,18,25,55,25,14,45,45,30,35,35,40,40,14,18,16,22,18,12,25,18,40],
        "skip_row_if": lambda row: not row.get("actividade"),
        "row_builder": "hiv",
    },
    "smi": {
        "topic_label":   "SMI",
        "code_prefix":   "smi",
        "base_url":      "https://kaiserso.github.io/INS_delphi/smi",
        "catalog_sheet": "SMI (2)",
        "key_field":     "intervencao",
        "col_map": {
            "Área":                          "area",
            "Programa":                      "programa",
            "Componente":                    "componente",
            "Intervenção":                   "intervencao",
            "Nível":                         "nivel",
            "Descrição (o que inclui)":      "descricao",   # note accented Descrição
            "Descricao (o que inclui)":      "descricao",   # fallback unaccented
            "Objectivo(s)":                  "objectivos",
            "Alcance geográfico da intervenção":                "alcance",
            "Alcance geografico da intervencao":                "alcance",
            "Recursos necessários para a implementação":        "recursos",
            "Recursos necessarios para a implementacao":        "recursos",
            "Etapas chave para a implementação da intervenção": "etapas",
            "Etapas chave para implementação da intervenção":   "etapas",
            "Etapas chave para a implementacao da intervencao": "etapas",
            "Descrição dos riscos e limitações que comprometem a implementação da intervenção": "riscos",
            "Descricao dos riscos e limitacoes que comprometem a implementacao da intervencao": "riscos",
            "Descrição dos riscos e limitações que compromentem a implementação da intervenção": "riscos",
            "Possíveis factores associados aos riscos e limitações descritas": "factores",
            "Possiveis factores associados aos riscos e limitacoes descritas": "factores",
            "Possiveis factores associados aos riscos e limitações descritas": "factores",
            "Ano de início":                 "ano_inicio",
            "Gastos em 2024 ":               "gastos_2024", # trailing space in source
            "Gastos em 2024":                "gastos_2024", # fallback
            "Fonte(s) de financiamento":     "fontes",
            "Pop elegivel":                  "pop_elegivel",
            "Número alcançado":              "num_alcancado",
            "Numero alcancado":              "num_alcancado",  # fallback
            "Cobertura":                     "cobertura",
            "Custo por unidade":             "custo_unidade",
            "Notas":                         "notas",
            # "Revisão" excluded
        },
        "output_headers": [
            "Código","URL da Ficha","Grupo",
            "Área","Programa","Componente","Intervenção",
            "Nível","Descrição (o que inclui)","Objectivo(s)",
            "Alcance geográfico","Recursos necessários","Etapas chave",
            "Riscos e limitações","Possíveis factores associados",
            "Ano de início","Gastos em 2024 (USD)","Fonte(s) de financiamento",
            "Pop. elegível","Número alcançado","Cobertura","Custo por unidade","Notas",
        ],
        "col_widths": [10,45,22,10,22,25,50,14,45,45,30,35,35,35,35,12,18,22,14,14,12,25,40],
        "skip_row_if": lambda row: False,
        "row_builder": "standard",
    },
}

# ─────────────────────────────────────────────────────────────────
# CATALOG READER
# ─────────────────────────────────────────────────────────────────

def load_catalog(path, cfg):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_name = cfg["catalog_sheet"]
    ws = None
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        # Try case-insensitive / prefix match (e.g. "SMI" matches "SMI (2)" and vice versa)
        sheet_name_lower = sheet_name.lower().split("(")[0].strip()
        for s in wb.sheetnames:
            if s.lower().split("(")[0].strip() == sheet_name_lower and "revis" not in s.lower():
                ws = wb[s]
                break
    if ws is None:
        ws = next((wb[s] for s in wb.sheetnames if "revis" not in s.lower()), wb.worksheets[0])

    col_map   = cfg["col_map"]
    normalized_col_map = {
        _normalize_header(source_header): target_key
        for source_header, target_key in col_map.items()
    }
    key_field = cfg["key_field"]

    # Locate header row: contains the key field label
    key_labels = {"Intervenção", "Actividade", "Actividade/ Descricao"}
    header_row_idx, headers = None, []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if any(str(c).strip() in key_labels for c in row if c):
            header_row_idx = i
            headers = [str(c).strip() if c else "" for c in row]
            break
    if header_row_idx is None:
        sys.exit(f"Header row not found in sheet '{sheet_name}' of {path}")

    interventions = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        # Skip completely empty rows
        if not any(c for c in row if c is not None and str(c).strip()):
            continue
        rec = {}
        for ci, h in enumerate(headers):
            # Use the first matching col_map key (handles duplicate cols like two Gastos)
            key = col_map.get(h) or normalized_col_map.get(_normalize_header(h))
            if key and key not in rec:
                rec[key] = row[ci] if ci < len(row) else None
        if not rec.get(key_field):
            continue
        if cfg["skip_row_if"](rec):
            continue
        interventions.append(rec)

    # Fill down: carry forward area, programa, componente, objectivos when blank
    _FILLDOWN = ("area", "programa", "componente", "objectivos")
    prev = {}
    for rec in interventions:
        for field in _FILLDOWN:
            val = rec.get(field)
            if val is None or str(val).strip() == "":
                if field in prev:
                    rec[field] = prev[field]
            else:
                prev[field] = val

    return interventions

# ─────────────────────────────────────────────────────────────────
# SHARED HELPERS
# ─────────────────────────────────────────────────────────────────

def fmt_coverage(val):
    if val is None:
        return None
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return f"{val * 100:.1f}%" if 0 <= val <= 1.5 else str(val)
    s = str(val).strip()
    try:
        f = float(s.replace("%", "").replace(",", "."))
        if 0 <= f <= 1.5:
            return f"{f * 100:.1f}%"
    except ValueError:
        pass
    return s or None

def sv(val):
    """Safe string: convert to str or return None."""
    return str(val) if val is not None else None

def _normalize_header(value):
    if value is None:
        return ""
    text = unicodedata.normalize("NFKD", str(value))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return " ".join(text.strip().lower().split())

def group_label(idx, interventions, cfg):
    """Build a descriptive group label based on Programa values within the batch."""
    gn    = idx // GROUP_SIZE + 1
    start = (gn - 1) * GROUP_SIZE
    end   = min(start + GROUP_SIZE, len(interventions))
    seen, parts = set(), []
    for i in range(start, end):
        p = str(interventions[i].get("programa", "")).strip().lower()
        if p and p not in seen:
            seen.add(p); parts.append(p)
    label = " e ".join(parts[:2]) if parts else f"grupo {gn}"
    return f"{gn}. {label}"

# ─────────────────────────────────────────────────────────────────
# SHARED STYLES
# ─────────────────────────────────────────────────────────────────

def _thin():
    s = Side(style="thin", color="D5CFC6")
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex6):
    return PatternFill("solid", fgColor=hex6)

def style_hdr(ws, row_n, bg="0F1923", fg="FFFFFF"):
    for cell in ws[row_n]:
        if cell.value is not None:
            cell.font      = Font(name="Arial", bold=True, size=9, color=fg)
            cell.fill      = _fill(bg)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.border    = _thin()

def style_data(ws, row_n, even):
    for cell in ws[row_n]:
        cell.font      = Font(name="Arial", size=9, color="0F1923")
        cell.fill      = _fill("FAFAF8" if even else "FFFFFF")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border    = _thin()

# ─────────────────────────────────────────────────────────────────
# ROW BUILDERS  (one per catalog structure variant)
# ─────────────────────────────────────────────────────────────────

def _ficha_url(base_url, code):
    """Return full URL if base_url is set, otherwise a relative path."""
    return f"{base_url}/{code}.html" if base_url else f"{code}.html"

def _row_standard(idx, rec, cfg, interventions):
    """Malária / SMI row layout."""
    code = f"{cfg['code_prefix']}_{idx+1:02d}"
    return [
        code,
        _ficha_url(cfg["base_url"], code),
        group_label(idx, interventions, cfg),
        rec.get("area"), rec.get("programa"), rec.get("componente"),
        rec.get("intervencao"), rec.get("nivel"),
        rec.get("descricao"), rec.get("objectivos"),
        rec.get("alcance"), rec.get("recursos"), rec.get("etapas"),
        rec.get("riscos"), rec.get("factores"),
        sv(rec.get("ano_inicio")), sv(rec.get("gastos_2024")),
        rec.get("fontes"),
        sv(rec.get("pop_elegivel")), sv(rec.get("num_alcancado")),
        fmt_coverage(rec.get("cobertura")),
        rec.get("custo_unidade"), rec.get("notas"),
    ]

def _row_tb(idx, rec, cfg, interventions):
    """TB row layout (same fields as SMI/Malária)."""
    code = f"{cfg['code_prefix']}_{idx+1:02d}"
    return [
        code,
        _ficha_url(cfg["base_url"], code),
        group_label(idx, interventions, cfg),
        rec.get("area"), rec.get("programa"), rec.get("componente"),
        rec.get("intervencao"), rec.get("nivel"),
        rec.get("descricao"), rec.get("objectivos"),
        rec.get("alcance"), rec.get("recursos"), rec.get("etapas"),
        rec.get("riscos"), rec.get("factores"),
        sv(rec.get("ano_inicio")), sv(rec.get("gastos_2024")),
        rec.get("fontes"),
        sv(rec.get("pop_elegivel")), sv(rec.get("num_alcancado")),
        fmt_coverage(rec.get("cobertura")),
        rec.get("custo_unidade"), rec.get("notas"),
    ]

def _row_hiv(idx, rec, cfg, interventions):
    """HIV row layout — Actividade + Implementador, no Área."""
    code = f"{cfg['code_prefix']}_{idx+1:02d}"
    return [
        code,
        _ficha_url(cfg["base_url"], code),
        group_label(idx, interventions, cfg),
        rec.get("programa"), rec.get("componente"),
        rec.get("actividade"), rec.get("implementador"),
        rec.get("nivel"),
        rec.get("descricao"), rec.get("objectivos"),
        rec.get("alcance"), rec.get("recursos"),
        rec.get("etapas"),
        rec.get("riscos"), rec.get("factores"),
        sv(rec.get("ano_inicio")), sv(rec.get("gastos_2024")),
        sv(rec.get("pop_elegivel")), rec.get("fonte_eligibilidade"),
        sv(rec.get("num_alcancado")), fmt_coverage(rec.get("cobertura")),
        rec.get("custo_unidade"), sv(rec.get("num_us")), rec.get("notas"),
    ]

ROW_BUILDERS = {
    "standard": _row_standard,
    "tb":       _row_tb,
    "hiv":      _row_hiv,
}

# ─────────────────────────────────────────────────────────────────
# SHEET 1: Catalogo
# ─────────────────────────────────────────────────────────────────

SECTION_SPANS = {
    "standard": [("A1","C1"),("D1","G1"),("H1","J1"),("K1","O1"),("P1","R1"),("S1","W1")],
    "tb":       [("A1","C1"),("D1","G1"),("H1","J1"),("K1","O1"),("P1","R1"),("S1","W1")],
    "hiv":      [("A1","C1"),("D1","G1"),("H1","O1"),("P1","R1"),("S1","X1")],
}
SECTION_LABELS = {
    "standard": ["Gerado pelo Script",None,None,
                 "Identificação",None,None,None,
                 "Descrição",None,None,
                 "Implementação",None,None,None,None,
                 "Financiamento",None,None,
                 "Cobertura e Custos",None,None,None,None],
    "tb":       ["Gerado pelo Script",None,None,
                 "Identificação",None,None,None,
                 "Descrição",None,None,
                 "Implementação",None,None,None,None,
                 "Financiamento",None,None,
                 "Cobertura e Custos",None,None,None,None],
    "hiv":      ["Gerado pelo Script",None,None,
                 "Identificação",None,None,None,
                 "Descrição e Implementação",None,None,None,None,None,None,None,
                 "Financiamento",None,None,
                 "Cobertura, Implementação e Custos",None,None,None,None,None],
}

def build_catalogo(wb, interventions, cfg):
    prog     = cfg["code_prefix"]
    variant  = cfg["row_builder"]
    ws       = wb.active
    ws.title = f"Catalogo_{cfg['topic_label'].replace('/','_').replace(' ','_')}"

    # Row 1: section super-headers
    labels = SECTION_LABELS.get(variant, SECTION_LABELS["standard"])
    ws.append(labels + [None] * (len(cfg["output_headers"]) - len(labels)))
    for c1, c2 in SECTION_SPANS.get(variant, SECTION_SPANS["standard"]):
        try:
            ws.merge_cells(f"{c1}:{c2}")
        except Exception:
            pass
    style_hdr(ws, 1, bg="0F1923")
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2: column headers
    ws.append(cfg["output_headers"])
    style_hdr(ws, 2, bg="1A5276")

    builder = ROW_BUILDERS[variant]
    for idx, rec in enumerate(interventions):
        row = builder(idx, rec, cfg, interventions)
        ws.append(row)
        style_data(ws, idx + 3, even=(idx % 2 == 0))

    for i, w in enumerate(cfg["col_widths"], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "D3"
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 30

# ─────────────────────────────────────────────────────────────────
# SHEET 2: Questionnaire template  (identical across programs)
# ─────────────────────────────────────────────────────────────────

Q_HEADERS = [
    "ID", "Secção", "Tipo Kobo", "Sufixo Variável\n(+_{CODE})",
    "Etiqueta (Português)", "Dica / Orientação", "Obrigatório",
    "Condição de Visibilidade\n({CODE} = código da intervenção)",
    "Validação", "Mensagem de Erro", "Lista de Opções",
    "Aparência Kobo", "Nota sobre Skip Logic", "Notas para Metodologia",
]

def _questions(topic_label):
    return [
        ["Q00","Identificação","note","intro_note",
         f"Exercício de Optimização — 1ª Oficina | {topic_label}",
         "Será apresentado com TODAS as intervenções do catálogo. Por favor avalie cada uma.",
         None,None,None,None,None,None,
         "Mostrado uma vez no início","Não repetido por intervenção"],
        ["Q01","Identificação","select_one codes","expert_code",
         "Código do Especialista",
         f"Utilize o código atribuído (ex: {EXPERT_CODE_COUNT:03d}{EXPERT_CODE_SUFFIX}). Use o mesmo em todas as oficinas.",
         "yes",None,f"regex(., '^[0-9]{{3}}({EXPERT_CODE_SUFFIX}|{ADMIN_CODE_SUFFIX})$')",
         f"Código deve ter 3 números seguidos por {EXPERT_CODE_SUFFIX} ou {ADMIN_CODE_SUFFIX}.","codes","minimal",
         "Perguntado uma vez","Chave de ligação entre W1, W2, W3"],
        ["S01","Estrutura","begin_group","grp_{CODE}",
         "Intervenção {N}/{TOTAL}: {LABEL}",
         None,None,None,None,None,None,"field-list",
         "Abre grupo. field-list = todas as perguntas no mesmo ecrã.","Repetido para cada intervenção."],
        ["Q03","Cabeçalho","note","hdr_{CODE}",
         "{LABEL}\n\n[↗ Ver ficha completa]({URL})",
         "{OBJECTIVE}",None,None,None,None,None,None,
         "Sempre visível dentro do grupo","URL renderizado como link markdown no label do note"],
        ["Q04","W1-G Expertise","select_one expertise","exp_{CODE}",
         "M1. Nível de expertise — {LABEL}",
         "Reflita:\n• Tenho envolvimento directo nesta intervenção?\n• Já participei no desenho ou implementação?\n• Sou responsável programático?",
         "yes",None,None,None,"expertise",None,
         "Sempre obrigatório, sem skip","Usado para ponderação no Resumo W1"],
        ["Q05","W1-H Gating","select_one optimizability","gate_{CODE}",
         "M2. Esta intervenção precisa de optimização?",
         "Indique se o modelo actual tem espaço relevante para melhoria em eficiência, qualidade, cobertura ou custo.",
         "yes",None,None,None,"optimizability",None,
         "GATE: controla visibilidade de Q06-Q15","Score: Sim def=2, Possiv=1, Não=0"],
        ["Q06","W1-I Duplicação","select_one yes_no","dup_{CODE}",
         "M3. Existe duplicação com outro programa?",
         "A intervenção realiza actividades semelhantes às de outra intervenção?",
         "yes","${gate_{CODE}} != 'nao'",None,None,"yes_no",None,
         "Visível se gate != 'nao'",None],
        ["Q07","W1-I Duplicação","select_multiple intervention_list","which_dup_{CODE}",
         "M4. Quais as intervenções com duplicação?",
         "Seleccione todas as intervenções com actividades semelhantes. Pode seleccionar várias.",
         "yes","${dup_{CODE}} = 'sim' and ${gate_{CODE}} != 'nao'",
         "not(selected(${which_dup_{CODE}}, '{CODE}'))",
         "Não seleccione a intervenção que está actualmente a avaliar ({LABEL}).",
         "intervention_list",None,
         "Visível se dup=sim AND gate!=nao. Constraint impede auto-selecção.",None],
        ["Q08","W1-I Duplicação","text","which_dup_other_{CODE}",
         "M5. Especifique a outra intervenção com duplicação",
         "Descreva a intervenção não listada.","yes",
         "selected(${which_dup_{CODE}}, 'other') and ${gate_{CODE}} != 'nao'",
         None,None,None,None,"Visível apenas se 'Outro' seleccionado em Q07",None],
        ["Q09","W1-J Integração","select_one yes_no","intg_{CODE}",
         "M6. Pode ser integrada com outra intervenção?",
         "É possível combinar actividades para melhorar eficiência?\nEx: supervisão conjunta, insumos partilhados.",
         "yes","${gate_{CODE}} != 'nao'",None,None,"yes_no",None,
         "Visível se gate != 'nao'",None],
        ["Q10","W1-J Integração","select_multiple intervention_list","which_intg_{CODE}",
         "M7. Com que intervenção(ões) deve ser integrada?",
         "Seleccione as intervenções com as quais esta poderia ser integrada.\nNota: supervisão, logística ou dados partilhados contam.",
         "yes","${intg_{CODE}} = 'sim' and ${gate_{CODE}} != 'nao'",
         "not(selected(${which_intg_{CODE}}, '{CODE}'))",
         "Não seleccione a intervenção que está actualmente a avaliar ({LABEL}).",
         "intervention_list",None,
         "Visível se intg=sim AND gate!=nao. Constraint impede auto-selecção.",None],
        ["Q11","W1-J Integração","text","which_intg_other_{CODE}",
         "M8. Especifique a outra intervenção para integração",
         "Descreva a intervenção não listada.","yes",
         "selected(${which_intg_{CODE}}, 'other') and ${gate_{CODE}} != 'nao'",
         None,None,None,None,"Visível apenas se 'Outro' seleccionado em Q10",None],
        ["Q12","W1-K Recursos","select_one yes_no","res_{CODE}",
         "M9. Pode reduzir recursos sem comprometer impacto?",
         "Existem áreas onde é possível racionalizar recursos sem afectar resultados?",
         "yes","${gate_{CODE}} != 'nao'",None,None,"yes_no",None,
         "Visível se gate != 'nao'",None],
        ["Q13","W1-L Outro motivo","select_one yes_no","oth_{CODE}",
         "M10. Existe outro motivo de optimização não identificado?",
         "Outro motivo relevante não capturado nas perguntas anteriores?",
         "yes","${gate_{CODE}} != 'nao'",None,None,"yes_no",None,
         "Visível se gate != 'nao'",None],
        ["Q14","W1-L Outro motivo","text","oth_reason_{CODE}",
         "M11. Descreva o outro motivo de optimização",
         "Descreva o motivo de optimização não capturado nas categorias anteriores.",
         "yes","${oth_{CODE}} = 'sim' and ${gate_{CODE}} != 'nao'",
         None,None,None,None,"Visível se oth=sim AND gate!=nao","Campo de texto livre"],
        ["Q15","W1-M Impacto","select_one impact","impact_{CODE}",
         "M12. Impacto provável da optimização",
         "Se esta intervenção for optimizada, qual o impacto esperado?\n\n1 – Baixo: melhoria marginal\n2 – Médio: melhoria moderada\n3 – Alto: melhoria significativa / potencialmente transformadora",
         "yes","${gate_{CODE}} != 'nao'",None,None,"impact",None,
         "Visível se gate != 'nao'","Score 1-3 para agregação"],
        ["Q16","Comentário","text","cmt_{CODE}",
         "M13. Como acha que a intervenção pode ser optimizada?",
         "Uma descrição breve de como acha que a intervenção pode ser optimizada.",
         "yes","${gate_{CODE}} != 'nao'",None,None,None,None,
         "Visível se gate != 'nao'",None],
        ["S02","Estrutura","end_group","grp_{CODE}",
         None,None,None,None,None,None,None,None,
         "Fecha grupo da intervenção",None],
    ]

def build_dicionario(wb, topic_label):
    ws = wb.create_sheet("Dicionario_Perguntas")
    ws.append(Q_HEADERS)
    style_hdr(ws, 1, bg="1A5276")
    ws.row_dimensions[1].height = 35
    for i, q in enumerate(_questions(topic_label), 2):
        ws.append(q)
        if q[0] and q[0].startswith("S"):
            for cell in ws[i]:
                cell.fill = _fill("FEF9E7")
                cell.font = Font(name="Arial", bold=True, size=9, color="B7860B")
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = _thin()
        else:
            style_data(ws, i, even=(i % 2 == 0))
    for i, w in enumerate([6,16,24,24,50,55,10,45,40,45,18,14,40,35], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

# ─────────────────────────────────────────────────────────────────
# SHEET 3: Option lists
# ─────────────────────────────────────────────────────────────────

def build_listas(wb, interventions, cfg):
    ws   = wb.create_sheet("Listas_de_Opcoes")
    pfx  = cfg["code_prefix"]
    kf   = cfg["key_field"]
    ws.append(["Lista","Código","Etiqueta (PT)","Notas"])
    style_hdr(ws, 1, bg="1A5276")

    rows = [
        ["modality","presencial","Presencial (oficina)",None],
        ["modality","remoto","Remoto (online/assíncrono)",None],
        ["expertise","1","1 – Conhecimento geral","Familiaridade básica; sem envolvimento directo"],
        ["expertise","2","2 – Domínio intermédio","Experiência prática; participou no desenho/implementação"],
        ["expertise","3","3 – Alto domínio técnico","Especialista técnico; responsável programático"],
        ["optimizability","sim_def","Sim, definitivamente","Gate aberto → Q06-Q15 visíveis"],
        ["optimizability","possivelmente","Possivelmente","Gate aberto → Q06-Q15 visíveis"],
        ["optimizability","nao","Não","Gate fechado → Q06-Q15 ocultas"],
        ["yes_no","sim","Sim",None],
        ["yes_no","nao","Não",None],
        ["impact","1","1 – Baixo","Melhoria marginal"],
        ["impact","2","2 – Médio","Melhoria moderada"],
        ["impact","3","3 – Alto","Melhoria significativa / potencialmente transformadora"],
        ["intervention_list","— gerado automaticamente pelo script —",
         "Uma entrada por intervenção + 'other' no final","Não editar manualmente"],
    ]
    for n in range(1, EXPERT_CODE_COUNT + 1):
        c = f"{n:03d}{EXPERT_CODE_SUFFIX}"; rows.append(["codes", c, c, None])
    for n in range(1, ADMIN_CODE_COUNT + 1):
        c = f"{n:03d}{ADMIN_CODE_SUFFIX}"; rows.append(["codes", c, c, None])
    for idx, rec in enumerate(interventions):
        code  = f"{pfx}_{idx+1:02d}"
        label = str(rec.get(kf, code)).strip()
        rows.append(["intervention_list", code, label, None])
    rows.append(["intervention_list","other","Outro (não listado)",None])

    for i, row in enumerate(rows, 2):
        ws.append(row); style_data(ws, i, even=(i % 2 == 0))
    for col, w in zip("ABCD", [20,35,55,45]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 25

# ─────────────────────────────────────────────────────────────────
# SHEET 4: README
# ─────────────────────────────────────────────────────────────────

def build_readme(wb, n, cfg):
    ws       = wb.create_sheet("README")
    n_groups = math.ceil(n / GROUP_SIZE)
    now      = datetime.now().strftime("%Y-%m-%d %H:%M")
    prog     = cfg["code_prefix"]
    label    = cfg["topic_label"]
    pfx      = cfg["code_prefix"]
    BOLD     = ("DICIONÁRIO","CATÁLOGO","CAMPOS","SKIP","USO","GRUPOS","NOTA")
    hiv_note = (
        "\nNOTA HIV: Este catálogo tem estrutura diferente dos outros programas.\n"
        "  • Sem coluna 'Área'; Programa começa na coluna B.\n"
        "  • Usa 'Actividade' em vez de 'Intervenção'.\n"
        "  • Inclui coluna 'Implementador' (financiador/parceiro).\n"
        "  • Tem duas colunas de Gastos; o dicionário usa 'Total Gastos 2024'.\n"
        "  • 177 linhas = actividades × implementadores (não intervenções distintas).\n"
        "  • Considere consolidar linhas por actividade antes de usar no formulário."
    ) if prog == "hiv" else ""

    lines = [
        f"DICIONÁRIO DE DADOS — Delphi W1 | {label}  (gerado {now})",
        "",
        f"CATÁLOGO (Catalogo_{label.replace('/','_').replace(' ','_')})",
        f"  Col 1  Código — gerado automaticamente: {pfx}_01, {pfx}_02, …",
        f"  Col 2  URL da Ficha — gerado a partir de base_url no script.",
        f"  Col 3  Grupo — {n_groups} grupo(s) de ~{GROUP_SIZE} intervenções cada.",
        f"  Restantes = colunas do catálogo original (Revisao excluída).",
        "",
        "CAMPOS DO CATÁLOGO",
        "  Ver cabeçalhos da folha Catalogo para lista completa de campos.",
        "",
        "SKIP LOGIC (resumo)",
        "  Q05 gate = 'nao'    → Q06-Q15 ocultas",
        "  Q06 dup = 'sim'     → Q07 visível",
        "  Q07 inclui 'other'  → Q08 visível",
        "  Q09 intg = 'sim'    → Q10 visível",
        "  Q10 inclui 'other'  → Q11 visível",
        "  Q13 outro = 'sim'   → Q14 visível",
        "  Q07/Q10 constraint impede selecção da própria intervenção",
        "",
        "GRUPOS DE FORMULÁRIO",
        f"  {n_groups} grupo(s) de ~{GROUP_SIZE} intervenções. Cada especialista recebe {n_groups} ficha(s) Kobo.",
        "",
        "USO",
        f"  1. Actualizar catálogo com dados reais.",
        f"  2. Actualizar base_url no script (PROGRAMS['{prog}']['base_url']).",
        f"  3. python generate_dictionaries.py --program {prog} --catalog <catalog.xlsx>",
        f"  4. Saída: dicionario_delphi_w1_{prog}.xlsx",
        hiv_note,
    ]
    for i, line in enumerate(filter(lambda x: x is not None, lines), 1):
        ws.cell(row=i, column=1, value=line)
        ws.cell(row=i, column=1).font = Font(
            name="Arial", size=9, bold=any(line.startswith(p) for p in BOLD))
        ws.cell(row=i, column=1).alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 90

# ─────────────────────────────────────────────────────────────────
# BUILD ONE DICTIONARY
# ─────────────────────────────────────────────────────────────────

def build_dictionary(program_key, catalog_path, output_dir):
    cfg = PROGRAMS[program_key]
    print(f"\n[{cfg['topic_label']}] Reading: {catalog_path}")
    interventions = load_catalog(catalog_path, cfg)
    n = len(interventions)
    n_groups = math.ceil(n / GROUP_SIZE)
    print(f"  {n} rows · {n_groups} group(s)")

    os.makedirs(output_dir, exist_ok=True)
    out = os.path.join(output_dir, f"dicionario_delphi_w1_{program_key}.xlsx")

    wb = openpyxl.Workbook()
    build_catalogo(wb, interventions, cfg)
    build_dicionario(wb, cfg["topic_label"])
    build_listas(wb, interventions, cfg)
    build_readme(wb, n, cfg)
    wb.save(out)
    print(f"  Saved: {out}")
    return out

# ─────────────────────────────────────────────────────────────────
# DEFAULT CATALOG PATHS  (for --all mode)
# ─────────────────────────────────────────────────────────────────

DEFAULT_CATALOGS = {
    "malaria": "Cata_logo_de_intervenc_o_es___Mala_ria__EG.xlsx",
    "tb":      "Catalogo_de_intervencoes_TB_rev1_revPN21012026.xlsx",
    "hiv":     "Catalogo_de_intervencoes_HIVv_30122025_rev1_11032026_1_.xlsx",
    "smi":     "Cata_logo_de_Intervenc_o_es_SMI-06_02_2026_UV_16h.xlsx",
}

# ─────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────

def main():
    program    = None
    catalog    = None
    # Default output dir: directory of INPUT_FILE from config.env (e.g. "dict/"),
    # falling back to current directory if not set.
    _input_file = _get("INPUT_FILE", "")
    output_dir  = str(pathlib.Path(_input_file).parent) if _input_file else "."
    run_all    = False

    i = 1
    while i < len(sys.argv):
        arg = sys.argv[i]
        if arg == "--all":
            run_all = True; i += 1
        elif arg in ("--program", "-p") and i + 1 < len(sys.argv):
            program = sys.argv[i + 1].lower(); i += 2
        elif arg.startswith("--program="):
            program = arg.split("=",1)[1].lower(); i += 1
        elif arg in ("--catalog", "-c") and i + 1 < len(sys.argv):
            catalog = sys.argv[i + 1]; i += 2
        elif arg.startswith("--catalog="):
            catalog = arg.split("=",1)[1]; i += 1
        elif arg in ("--output-dir", "-o") and i + 1 < len(sys.argv):
            output_dir = sys.argv[i + 1]; i += 2
        elif arg.startswith("--output-dir="):
            output_dir = arg.split("=",1)[1]; i += 1
        else:
            if not arg.startswith("-") and catalog is None:
                catalog = arg
            i += 1

    if run_all:
        missing = []
        for prog, default_path in DEFAULT_CATALOGS.items():
            if not os.path.exists(default_path):
                missing.append(f"  {prog}: {default_path}")
        if missing:
            print("Warning: some default catalog files not found:")
            for m in missing: print(m)
            print("Place them in the current directory or use --program / --catalog.\n")
        for prog, default_path in DEFAULT_CATALOGS.items():
            if os.path.exists(default_path):
                build_dictionary(prog, default_path, output_dir)
            else:
                print(f"[{prog}] Skipped — file not found: {default_path}")
        return

    if not program:
        print(__doc__)
        sys.exit("Error: specify --program <name> or --all")
    if program not in PROGRAMS:
        sys.exit(f"Unknown program '{program}'. Choices: {', '.join(PROGRAMS)}")

    if not catalog:
        catalog = DEFAULT_CATALOGS.get(program, f"catalogo_{program}.xlsx")
    if not os.path.exists(catalog):
        print(__doc__)
        sys.exit(f"Error: catalog file not found: {catalog}")

    build_dictionary(program, catalog, output_dir)


if __name__ == "__main__":
    main()
