#!/usr/bin/env python3
"""
generate_kobo_and_pages.py
Reads dicionario_delphi_w1_malaria.xlsx and produces:
  1. delphi_w1_malaria_kobo.xlsx  — KoboToolbox XLSForm
  2. pages/malaria/*.html         — one static page per intervention

Usage:
  python generate_kobo_and_pages.py
  (paths hardcoded below — edit INPUT_FILE / OUTPUT_KOBO / PAGES_DIR as needed)
"""

import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import datetime
import html as html_module

# ── Configuration — loaded from config.env ────────────────────
# Edit config.env in the same directory. Values persist across
# script regenerations. Do not edit the defaults below directly.

import pathlib

def load_config(path="config.env"):
    """Parse a simple key=value config file. Returns a dict of strings."""
    cfg = {}
    config_path = "." / path
    if not config_path.exists():
        config_path = pathlib.Path(__file__).parent / path
        if not config_path.exists():
            raise FileNotFoundError(
                f"Config file not found: {config_path}\n"
                f"Create config.env next to the script. See config.env.example."
            )
    
    with open(config_path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" not in line:
                continue
            key, _, value = line.partition("=")
            cfg[key.strip()] = value.strip()
    return cfg

_cfg = load_config()

def _get(key, default=""):
    return _cfg.get(key, default)

def _bool(key, default=True):
    return _get(key, str(default)).lower() not in ("false", "0", "no", "off")

INPUT_FILE    = _get("INPUT_FILE",   "dicionario_delphi_w1_malaria.xlsx")
OUTPUT_KOBO   = _get("OUTPUT_KOBO",  "delphi_w1_malaria_kobo.xlsx")
PAGES_DIR     = _get("PAGES_DIR",    "pages/malaria")
TOPIC_LABEL   = _get("TOPIC_LABEL",  "Malária")
TOPIC_CODE    = _get("TOPIC_CODE",   "malaria")
BASE_URL      = _get("BASE_URL",     "https://delphi-catalogo.example.org/malaria")
MAGIC_API_KEY = _get("MAGIC_API_KEY","pk_live_YOUR_KEY_HERE")
GATEWAY_URL   = _get("GATEWAY_URL",  "https://your-site.github.io/gateway.html")
REQUIRE_AUTH  = _bool("REQUIRE_AUTH", True)

SUBFORM_GROUP_BY = _get("SUBFORM_GROUP_BY", "programa").lower().strip()
SUBFORM_MAX_SIZE = int(_get("SUBFORM_MAX_SIZE", "0") or "0")
MASTER_PAGE      = _get("MASTER_PAGE", "master.html")
MASTER_URL       = _get("MASTER_URL",  "https://your-site.github.io/master.html")
RELAY_URL        = _get("RELAY_URL",   "https://your-site.github.io/relay.html")
CLOSING_NOTE     = _get("CLOSING_NOTE",
    "Após submeter, as suas respostas não poderão ser editadas. "
    "O grupo de metodologia irá compilar os resultados antes da 2ª Oficina.")

# Read all SUBFORM_URL_<slug> = <url> entries
SUBFORM_URLS = {
    k[len("SUBFORM_URL_"):].strip(): v
    for k, v in _cfg.items()
    if k.upper().startswith("SUBFORM_URL_")
}

print(f"Config loaded from config.env")
print(f"  Topic:       {TOPIC_LABEL} ({TOPIC_CODE})")
print(f"  Input:       {INPUT_FILE}")
print(f"  Output Kobo: {OUTPUT_KOBO}")
print(f"  Pages dir:   {PAGES_DIR}")
print(f"  Base URL:    {BASE_URL}")
print(f"  Auth guard:  {'enabled' if REQUIRE_AUTH else 'disabled'}")
print(f"  Grouping:    by {SUBFORM_GROUP_BY}"
      + (f", max {SUBFORM_MAX_SIZE} per form" if SUBFORM_MAX_SIZE else ""))

os.makedirs(PAGES_DIR, exist_ok=True)

# ── Auth guard snippet (injected into every HTML page) ────────
# Checks Magic.link session on page load; redirects to gateway if not logged in.
# Rendered as empty string when REQUIRE_AUTH = False.
def auth_guard_html():
    if not REQUIRE_AUTH:
        return ""
    return f"""
  <!-- Auth guard — Magic.link -->
  <script src="https://cdn.jsdelivr.net/npm/magic-sdk/dist/magic.js"></script>
  <script>
    (async function() {{
      document.documentElement.style.visibility = 'hidden';
      let redirecting = false;
      try {{
        const magic = new Magic('{MAGIC_API_KEY}');
        const loggedIn = await magic.user.isLoggedIn();
        if (!loggedIn) {{
          redirecting = true;
          window.location.replace(
            '{GATEWAY_URL}?return=' + encodeURIComponent(window.location.href)
          );
          return;
        }}
        window._magicLogout = async function() {{
          await magic.user.logout();
          window.location.replace('{GATEWAY_URL}');
        }};
      }} catch(e) {{
        console.warn('Auth check failed:', e);
      }} finally {{
        // Only restore visibility if we are NOT redirecting to the gateway.
        // If redirecting, keep hidden to avoid a flash of protected content.
        if (!redirecting) {{
          document.documentElement.style.visibility = '';
        }}
      }}
    }})();
  </script>"""

# ── Style helpers ─────────────────────────────────────────────
def style_header(ws, row_num, bg="1F4E79"):
    for cell in ws[row_num]:
        if cell.value is not None:
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
            cell.alignment = Alignment(wrap_text=True, vertical="center",
                                       horizontal="center")
    ws.row_dimensions[row_num].height = 28

def style_data(ws, row_num, bg="FFFFFF", bold=False, fg="000000"):
    for cell in ws[row_num]:
        if cell.column <= ws.max_column:
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(size=9, name="Arial", bold=bold, color=fg)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── Read data dictionary ──────────────────────────────────────
print("Reading dictionary:", INPUT_FILE)
wb_in = openpyxl.load_workbook(INPUT_FILE, data_only=True)

# --- Catalogue tab ---
cat_ws = wb_in["Catalogo_Malaria"]
# Row 1 = zone headers (merged), row 2 = column headers, row 3+ = data
cat_headers = [str(c.value).strip() if c.value else ""
               for c in cat_ws[2]]

def col_idx(headers, name):
    """Return 0-based index of column by header name."""
    for i, h in enumerate(headers):
        if name.lower() in h.lower():
            return i
    return None

interventions = []
for row in cat_ws.iter_rows(min_row=3, values_only=True):
    if not row[0]:
        continue
    intv = {h: (str(row[i]) if row[i] is not None else "")
            for i, h in enumerate(cat_headers) if h}
    # Override URL from BASE_URL so it stays in sync with config
    intv["URL da Ficha"] = f"{BASE_URL}/{intv['Código']}.html"
    interventions.append(intv)

# ── Normalise numeric columns ─────────────────────────────────
# Columns that should be formatted as integers (thousands sep = ".")
INT_COLS   = ["Pop. elegível", "Número alcançado"]
# Columns that may have decimal places (thousands sep = ".", decimal = ",")
FLOAT_COLS = ["Gastos em 2024 (MZN)"]

def _try_parse_number(raw):
    """
    Try to interpret raw as a number.
    Strips spaces (used as thousands separators in some cells).
    Returns (float_value, is_integer) or None if not clearly a number.
    """
    if not raw or raw in ("None", "nan", "#DIV/0!"):
        return None
    cleaned = raw.strip().replace(" ", "")
    # Reject percentages, strings with letters, or obvious non-numbers
    import re as _re
    if "%" in cleaned:
        return None
    if _re.search(r"[A-Za-zÀ-ÿ]", cleaned):
        return None
    # Normalise: if both "." and "," present, figure out which is decimal
    # European style: 1.234,56 → decimal is ","
    # US/float style: 1234.56  → decimal is "."
    if "," in cleaned and "." in cleaned:
        # Assume European if "." comes before ","
        if cleaned.index(".") < cleaned.index(","):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif "," in cleaned:
        # Comma only — treat as decimal separator
        cleaned = cleaned.replace(",", ".")
    # Remove any remaining spaces or stray chars except digits, ".", "-"
    cleaned = _re.sub(r"[^\d.\-]", "", cleaned)
    if not cleaned:
        return None
    try:
        val = float(cleaned)
        is_int = (val == int(val))
        return (val, is_int)
    except ValueError:
        return None

def _fmt_pt(val, force_int=False):
    """
    Format a number in Portuguese style:
    thousands separator = "."  decimal separator = ","
    """
    if force_int or val == int(val):
        return f"{int(val):,}".replace(",", ".")
    # Up to 2 decimal places, strip trailing zeros
    formatted = f"{val:,.2f}"
    # swap separators: first replace "," with X, then "." with ",", then X with "."
    formatted = formatted.replace(",", "X").replace(".", ",").replace("X", ".")
    formatted = formatted.rstrip("0").rstrip(",")
    return formatted

for intv in interventions:
    for col in INT_COLS:
        raw = intv.get(col, "")
        result = _try_parse_number(raw)
        if result is not None:
            intv[col] = _fmt_pt(result[0], force_int=True)
    for col in FLOAT_COLS:
        raw = intv.get(col, "")
        result = _try_parse_number(raw)
        if result is not None:
            intv[col] = _fmt_pt(result[0], force_int=False)

print(f"  Found {len(interventions)} interventions")

# --- Question dictionary tab ---
qd_ws = wb_in["Dicionario_Perguntas"]
qd_headers = [str(c.value).strip() if c.value else ""
              for c in qd_ws[1]]

q_templates = []
for row in qd_ws.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    q = {h: (str(row[i]) if row[i] is not None else "")
         for i, h in enumerate(qd_headers) if h}
    q_templates.append(q)

# Patch Q00 (intro note) to use {GROUP_LABEL} and {TOPIC_LABEL} placeholders.
# This ensures the title page shows the correct group regardless of what's
# stored in the dictionary. The dictionary may predate this feature.
LABEL_COL = "Etiqueta (Português)"
HINT_COL  = "Dica / Orientação"
for q in q_templates:
    if q.get("ID", "") == "Q00":
        q[LABEL_COL] = ("Exercício de Optimização — 1ª Oficina"
                        " | {TOPIC_LABEL} · {GROUP_LABEL}")
        q[HINT_COL]  = ("Será apresentado com as intervenções do grupo "
                        "'{GROUP_LABEL}'. Por favor avalie cada uma. "
                        "Este exercício não visa eliminar programas.")
        break

print(f"  Found {len(q_templates)} question templates")

# --- Choices tab ---
choices_ws = wb_in["Listas_de_Opcoes"]
base_choices = []
for row in choices_ws.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    # Skip ALL intervention_list rows — the script builds this list
    # entirely from the catalogue, so any row in the dictionary tab
    # (including the placeholder) must be excluded.
    if str(row[0]).strip() == "intervention_list":
        continue
    base_choices.append({
        "list_name": str(row[0]).strip(),
        "name":      str(row[1]).strip() if row[1] else "",
        "label":     str(row[2]).strip() if row[2] else "",
    })

# Build intervention_list choices from catalogue
intv_choices = []
for intv in interventions:
    intv_choices.append({
        "list_name": "intervention_list",
        "name":      intv["Código"],
        "label":     intv["Intervenção"],
    })
# Add 'other' at end
intv_choices.append({
    "list_name": "intervention_list",
    "name":      "other",
    "label":     "Outro (não listado)",
})

all_choices = base_choices + intv_choices
print(f"  Built {len(all_choices)} choice rows "
      f"({len(intv_choices)} in intervention_list)")

# ════════════════════════════════════════════════════════════
# GROUPING — must run before Kobo generation
# ════════════════════════════════════════════════════════════

GROUP_FIELD = {
    "area":       "Área",
    "programa":   "Programa",
    "componente": "Componente",
}.get(SUBFORM_GROUP_BY, "Programa")

def slugify(text):
    import unicodedata
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = text.lower().strip()
    text = re.sub(r"[^\w\s]", "", text)
    text = re.sub(r"\s+", "_", text)
    return text

def split_group(label, items, max_size):
    if max_size <= 0 or len(items) <= max_size:
        return [(label, items)]
    chunks = []
    for idx, start in enumerate(range(0, len(items), max_size), 1):
        chunks.append((f"{label} ({idx})", items[start:start + max_size]))
    return chunks

from collections import OrderedDict
groups_raw = OrderedDict()
for intv in interventions:
    key = intv.get(GROUP_FIELD, "").strip()
    if not key or key in ("None", "nan"):
        key = "Outros"
    groups_raw.setdefault(key, []).append(intv)

subform_groups = []   # list of (label, slug, [intv, ...])
for label, items in groups_raw.items():
    for chunk_label, chunk_items in split_group(label, items, SUBFORM_MAX_SIZE):
        subform_groups.append((chunk_label, slugify(chunk_label), chunk_items))

print(f"\nSub-form groups ({GROUP_FIELD}, max {SUBFORM_MAX_SIZE or 'unlimited'}):")
for label, slug, items in subform_groups:
    print(f"  [{slug}]  {label} — {len(items)} interventions")

# ════════════════════════════════════════════════════════════
# PART 1: GENERATE KOBOTOOLBOX XLSFORMS (one per group)
# ════════════════════════════════════════════════════════════
print("\nGenerating KoboToolbox XLSForms...")

# ── Substitution helpers ──────────────────────────────────────
def substitute(text, intv, n, total, group_label=""):
    """Replace {placeholders} with intervention-specific values."""
    if not text:
        return ""
    text = str(text)
    # Global substitutions — always applied regardless of intv
    text = text.replace("{TOPIC_LABEL}", TOPIC_LABEL)
    text = text.replace("{GROUP_LABEL}", group_label)
    if intv:
        text = text.replace("{CODE}",      intv["Código"])
        text = text.replace("{LABEL}",     intv["Intervenção"])
        text = text.replace("{OBJECTIVE}", intv["Objectivo(s)"])
        text = text.replace("{URL}",       intv["URL da Ficha"])
        text = text.replace("{N}",         str(n))
        text = text.replace("{TOTAL}",     str(total))
    return text

def make_survey_row(q, intv=None, n=None, total=None, group_label=""):
    """Turn a question template into a survey sheet row."""
    def sub(field):
        val = q.get(field, "")
        return substitute(val, intv, n, total, group_label)

    def clean(v):
        return "" if v in ("None", "nan") else v

    return [
        clean(sub("Tipo Kobo")),
        clean(sub("Sufixo Variável\n(+_{CODE})")),
        clean(sub("Etiqueta (Português)")),
        clean(sub("Dica / Orientação")),
        clean(sub("Obrigatório")),
        clean(sub("Condição de Visibilidade\n({CODE} = código da intervenção)")),
        clean(sub("Validação")),
        clean(sub("Mensagem de Erro")),
        "",   # choice_filter
        clean(sub("Aparência Kobo")),
        clean(sub("calculation") if "calculation" in q else ""),
    ]

# ── Row colour maps ───────────────────────────────────────────
TYPE_BG = {
    "note":                             "FFF9C4",
    "text":                             "FFFFFF",
    "begin_group":                      "C55A11",
    "end_group":                        "EEEEEE",
    "calculate":                        "F3F3F3",
    "select_one modality":              "E3F2FD",
    "select_one expertise":             "FCE4D6",
    "select_one optimizability":        "FCE4D6",
    "select_one yes_no":                "F0F7EE",
    "select_multiple intervention_list":"E8EAF6",
    "select_one impact":                "EDE7F6",
}

CH_BG = {
    "modality":          "E3F2FD",
    "expertise":         "FCE4D6",
    "optimizability":    "FCE4D6",
    "yes_no":            "E8F5E9",
    "impact":            "EDE7F6",
    "intervention_list": "F3E5F5",
}

# ── Helper: build one XLSForm for a given list of interventions ──
def generate_xlsform(group_interventions, group_label, group_slug):
    wb_out = openpyxl.Workbook()

    # ── survey sheet ─────────────────────────────────────────
    survey = wb_out.active
    survey.title = "survey"

    s_cols = ["type", "name", "label::Português", "hint::Português",
              "required", "relevant", "constraint", "constraint_message",
              "choice_filter", "appearance", "calculation"]
    survey.append(s_cols)
    style_header(survey, 1)

    group_total = len(group_interventions)

    survey_rows = []

    # 1. Identification questions
    for q in q_templates:
        if q.get("Secção", "") == "Identificação":
            survey_rows.append(("id", make_survey_row(
                q, total=group_total, group_label=group_label)))

    # 2. Per-intervention groups
    for n, intv in enumerate(group_interventions, 1):
        icode = intv["Código"]
        iurl  = intv["URL da Ficha"]
        calc_row = ["calculate", f"url_{icode}", "", "", "", "", "", "",
                    "", "", f"'{iurl}'"]
        survey_rows.append(("calc", calc_row))
        for q in q_templates:
            if q.get("Secção", "") == "Identificação":
                continue
            row = make_survey_row(q, intv, n, group_total, group_label)
            survey_rows.append((q.get("Secção", ""), row))

    # 3. Version stamp note (visible in form — confirms which version is loaded)
    version_stamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    survey_rows.append(("note", [
        "note", "form_version_stamp",
        f"Versão do formulário: {version_stamp}",
        f"Este formulário foi gerado em {version_stamp}. "
        f"Se esta data não corresponder à versão esperada, contacte o administrador.",
        "", "", "", "", "", "", ""
    ]))

    # 4. Closing note
    for q in q_templates:
        if "closing" in q.get("Sufixo Variável\n(+_{CODE})", "").lower():
            survey_rows.append(("closing", make_survey_row(
                q, group_label=group_label)))
            break
    else:
        survey_rows.append(("closing", [
            "note", "closing",
            "Obrigado pela sua contribuição!",
            CLOSING_NOTE,
            "", "", "", "", "", "", ""
        ]))

    # Write survey rows
    for i, (section, row_data) in enumerate(survey_rows):
        survey.append(row_data)
        rnum = i + 2
        rtype = row_data[0] if row_data else ""
        bg = TYPE_BG.get(rtype, "FFFFFF")
        if bg == "FFFFFF" and i % 2 == 1:
            bg = "F9F9F9"
        is_group = rtype == "begin_group"
        for ci in range(1, len(s_cols) + 1):
            cell = survey.cell(row=rnum, column=ci)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(size=9, name="Arial",
                             bold=is_group,
                             color="FFFFFF" if is_group else "000000")
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    set_widths(survey, [34, 26, 50, 65, 10, 55, 45, 42, 14, 12, 45])

    # ── choices sheet — intervention_list scoped to this group ──
    # Build a group-scoped intervention_list (only this group's interventions
    # + other) so experts can only reference interventions in their own form.
    group_intv_choices = [
        {"list_name": "intervention_list",
         "name": intv["Código"], "label": intv["Intervenção"]}
        for intv in group_interventions
    ]
    group_intv_choices.append(
        {"list_name": "intervention_list",
         "name": "other", "label": "Outro (não listado)"}
    )
    group_choices = base_choices + group_intv_choices

    choices_out = wb_out.create_sheet("choices")
    c_cols = ["list_name", "name", "label::Português"]
    choices_out.append(c_cols)
    style_header(choices_out, 1, bg="375623")

    for i, ch in enumerate(group_choices):
        choices_out.append([ch["list_name"], ch["name"], ch["label"]])
        rnum = i + 2
        bg = CH_BG.get(ch["list_name"], "FFFFFF")
        if bg in ("FFFFFF", "F3E5F5") and i % 2 == 1:
            bg = "F0F0F0" if bg == "FFFFFF" else "EAD5EA"
        for ci in range(1, 4):
            cell = choices_out.cell(row=rnum, column=ci)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(size=9, name="Arial")
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    set_widths(choices_out, [22, 16, 55])

    # ── settings sheet ────────────────────────────────────────
    settings_out = wb_out.create_sheet("settings")
    st_cols = ["form_title", "form_id", "version",
               "default_language", "instance_name", "style"]
    settings_out.append(st_cols)
    style_header(settings_out, 1, bg="44546A")
    version = datetime.datetime.now().strftime("%Y%m%d%H")
    safe_slug = re.sub(r"[^\w]", "_", group_slug)
    settings_out.append([
        f"W1 – {group_label} | {TOPIC_LABEL}",
        f"delphi_w1_{TOPIC_CODE}_{safe_slug}",
        version,
        "Português",
        f"concat(${{expert_code}}, '_{TOPIC_CODE}_{safe_slug}_', "
        "format-date(today(), '%Y%m%d'))",
        "pages",
    ])
    for ci in range(1, len(st_cols) + 1):
        cell = settings_out.cell(row=2, column=ci)
        cell.fill = PatternFill("solid", fgColor="D6DCE4")
        cell.font = Font(size=9, name="Arial")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    set_widths(settings_out, [40, 30, 12, 16, 55, 10])

    return wb_out, survey.max_row - 1, len(group_choices)

# ── Generate one XLSForm per group ───────────────────────────
# Derive output filename from OUTPUT_KOBO base + group slug
kobo_base = re.sub(r"\.xlsx$", "", OUTPUT_KOBO, flags=re.IGNORECASE)

for group_label, group_slug, group_interventions in subform_groups:
    wb, n_survey, n_choices = generate_xlsform(
        group_interventions, group_label, group_slug)
    out_path = f"{kobo_base}_{group_slug}.xlsx"
    wb.save(out_path)
    print(f"  {out_path}  ({n_survey} survey rows, {n_choices} choices)")

# ════════════════════════════════════════════════════════════
# PART 2: GENERATE STATIC HTML PAGES
# ════════════════════════════════════════════════════════════
print("\nGenerating static HTML pages...")

def e(text):
    """HTML-escape a string."""
    return html_module.escape(str(text)) if text else ""

def nl2br(text):
    """Convert newlines to <br> tags."""
    return e(text).replace("\n", "<br>")

CSS = """
* { box-sizing: border-box; margin: 0; padding: 0; }
body {
    font-family: 'Segoe UI', Arial, sans-serif;
    font-size: 14px;
    color: #1a1a1a;
    background: #f5f7fa;
    padding: 0;
}
.page-wrap {
    max-width: 860px;
    margin: 32px auto;
    background: #fff;
    border-radius: 6px;
    box-shadow: 0 2px 12px rgba(0,0,0,0.10);
    overflow: hidden;
}
.page-header {
    background: #1F4E79;
    color: #fff;
    padding: 28px 36px 22px 36px;
}
.page-header .topic-tag {
    font-size: 11px;
    font-weight: 600;
    letter-spacing: 1.2px;
    text-transform: uppercase;
    color: #90CAF9;
    margin-bottom: 8px;
}
.page-header h1 {
    font-size: 24px;
    font-weight: 700;
    line-height: 1.3;
    margin-bottom: 10px;
}
.page-header .objective {
    font-size: 14px;
    color: #BBDEFB;
    line-height: 1.6;
    border-left: 3px solid #42A5F5;
    padding-left: 12px;
    margin-top: 10px;
}
.code-badge {
    display: inline-block;
    background: rgba(255,255,255,0.15);
    color: #fff;
    font-size: 11px;
    font-family: monospace;
    padding: 2px 8px;
    border-radius: 3px;
    margin-bottom: 10px;
}
.body-content { padding: 28px 36px 36px 36px; }
.section-title {
    font-size: 11px;
    font-weight: 700;
    letter-spacing: 1px;
    text-transform: uppercase;
    color: #1F4E79;
    border-bottom: 2px solid #BBDEFB;
    padding-bottom: 5px;
    margin: 24px 0 14px 0;
}
.field-grid {
    display: grid;
    grid-template-columns: 220px 1fr;
    gap: 0;
    border: 1px solid #e0e0e0;
    border-radius: 4px;
    overflow: hidden;
    margin-bottom: 12px;
}
.field-grid .field-row:nth-child(odd)  { background: #f8fbff; }
.field-grid .field-row:nth-child(even) { background: #ffffff; }
.field-row {
    display: contents;
}
.field-label {
    padding: 10px 14px;
    font-weight: 600;
    font-size: 12px;
    color: #37474F;
    border-bottom: 1px solid #e8e8e8;
    border-right: 1px solid #e0e0e0;
    background: inherit;
}
.field-value {
    padding: 10px 14px;
    font-size: 13px;
    color: #212121;
    line-height: 1.6;
    border-bottom: 1px solid #e8e8e8;
    background: inherit;
}
.ineff-box {
    background: #FFF8E1;
    border-left: 4px solid #FFB300;
    border-radius: 0 4px 4px 0;
    padding: 14px 16px;
    margin-bottom: 10px;
}
.ineff-box .ineff-title {
    font-weight: 700;
    font-size: 12px;
    color: #E65100;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    margin-bottom: 6px;
}
.ineff-box .ineff-text {
    font-size: 13px;
    color: #4E342E;
    line-height: 1.6;
}
.causes-box {
    background: #FBE9E7;
    border-left: 4px solid #E64A19;
    border-radius: 0 4px 4px 0;
    padding: 14px 16px;
    margin-bottom: 10px;
}
.causes-box .causes-title {
    font-weight: 700;
    font-size: 12px;
    color: #BF360C;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    margin-bottom: 6px;
}
.causes-box .causes-text {
    font-size: 13px;
    color: #4E342E;
    line-height: 1.6;
}
.nav-bar {
    background: #E3F2FD;
    padding: 12px 36px;
    font-size: 12px;
    color: #1565C0;
    border-top: 1px solid #BBDEFB;
    display: flex;
    justify-content: space-between;
    align-items: center;
}
.nav-bar a {
    color: #1565C0;
    text-decoration: none;
    font-weight: 600;
}
.nav-bar a:hover { text-decoration: underline; }
.stub-notice {
    background: #EDE7F6;
    border: 1px dashed #9575CD;
    border-radius: 4px;
    padding: 10px 14px;
    font-size: 12px;
    color: #4527A0;
    margin-bottom: 20px;
    text-align: center;
}
.footer {
    background: #ECEFF1;
    padding: 12px 36px;
    font-size: 11px;
    color: #90A4AE;
    text-align: center;
    border-top: 1px solid #CFD8DC;
}
@media print {
    body { background: #fff; }
    .page-wrap { box-shadow: none; margin: 0; border-radius: 0; }
    .stub-notice { display: none; }
}
"""

def build_html_page(intv, idx, total, prev_url, next_url, index_url):
    code   = intv["Código"]
    url    = intv["URL da Ficha"]
    area   = intv.get("Área", "")
    prog   = intv.get("Programa", "")
    comp   = intv.get("Componente", "")
    label  = intv["Intervenção"]
    level  = intv.get("Nível", "")
    desc   = intv.get("Descrição (o que inclui)", "")
    obj    = intv.get("Objectivo(s)", "")
    geo    = intv.get("Alcance geográfico da intervenção", "")
    res    = intv.get("Recursos necessários para a implementação", "")
    steps  = intv.get("Etapas chave para a implementação", "")
    risks  = intv.get("Riscos e limitações", "")
    causes = intv.get("Possíveis factores associados aos riscos", "")
    year   = intv.get("Ano de início", "")
    spend  = intv.get("Gastos em 2024 (MZN)", "")
    funder = intv.get("Fonte(s) de financiamento", "")
    pop    = intv.get("Pop. elegível", "")
    reached= intv.get("Número alcançado", "")
    cover  = intv.get("Cobertura", "")
    cost   = intv.get("Custo por unidade", "")
    notes  = intv.get("Notas", "")

    nav_prev = (f'<a href="{e(prev_url)}">← Anterior</a>'
                if prev_url else '<span style="color:#aaa">← Anterior</span>')
    nav_next = (f'<a href="{e(next_url)}">Próxima →</a>'
                if next_url else '<span style="color:#aaa">Próxima →</span>')

    def field_row(field_label, value):
        if not value or value in ("None", "nan", "#DIV/0!", ""):
            return ""
        return f"""
      <div class="field-row">
        <div class="field-label">{field_label}</div>
        <div class="field-value">{nl2br(value)}</div>
      </div>"""

    crumbs = [p for p in [area, prog, comp] if p and p not in ("None", "nan", "")]
    breadcrumb = " &rsaquo; ".join(e(p) for p in crumbs) if crumbs else e(TOPIC_LABEL)

    stub = ("" if "example.org" not in url else
            f'<div class="stub-notice">&#9888; Página de demonstração — '
            f'URL provisório: <code>{e(url)}</code>. '
            f'Actualize BASE_URL no script quando as páginas forem publicadas.</div>')

    notes_block = (f"<div class='section-title'>Notas</div>"
                   f"<div class='field-grid'>{field_row('Notas', notes)}</div>"
                   if notes and notes not in ("None", "nan", "") else "")

    return f"""<!DOCTYPE html>
<html lang="pt">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<meta name="robots" content="noindex, nofollow">
<title>{e(label)} — Catálogo Delphi | {e(TOPIC_LABEL)}</title>
{auth_guard_html()}
<style>{CSS}</style>
</head>
<body>
<div class="page-wrap">

  <div class="page-header">
    <div class="topic-tag">{breadcrumb} &nbsp;·&nbsp; Intervenção {idx} de {total}</div>
    <div class="code-badge">{e(code)}</div>
    <h1>{e(label)}</h1>
    <div class="objective">{nl2br(obj)}</div>
  </div>

  <div class="body-content">
    {stub}

    <div class="section-title">Identificação e Âmbito</div>
    <div class="field-grid">
      {field_row("Nível de Implementação", level)}
      {field_row("Alcance Geográfico", geo)}
    </div>

    <div class="section-title">Descrição</div>
    <div class="field-grid">
      {field_row("O que inclui", desc)}
      {field_row("Recursos Necessários", res)}
    </div>

    <div class="section-title">Implementação</div>
    <div class="field-grid">
      {field_row("Etapas Chave", steps)}
    </div>

    <div class="section-title">Riscos, Limitações e Causas</div>
    <div class="ineff-box">
      <div class="ineff-title">Riscos e Limitações</div>
      <div class="ineff-text">{nl2br(risks) if risks and risks not in ("None","nan","") else "—"}</div>
    </div>
    <div class="causes-box">
      <div class="causes-title">Possíveis Factores Associados</div>
      <div class="causes-text">{nl2br(causes) if causes and causes not in ("None","nan","") else "—"}</div>
    </div>

    <div class="section-title">Financiamento e Cobertura</div>
    <div class="field-grid">
      {field_row("Ano de Início", year)}
      {field_row("Gastos em 2024 (MZN)", spend)}
      {field_row("Fonte(s) de Financiamento", funder)}
      {field_row("Pop. Elegível", pop)}
      {field_row("Número Alcançado", reached)}
      {field_row("Cobertura", cover)}
      {field_row("Custo por Unidade", cost)}
    </div>

    {notes_block}

  </div>

  <div class="nav-bar">
    {nav_prev}
    <a href="{e(index_url)}">↑ Índice</a>
    {nav_next}
  </div>
  {"<div style='text-align:right;padding:8px 36px;font-size:11px;border-top:1px solid #eee'><a href=\"#\" onclick=\"window._magicLogout&&window._magicLogout();return false;\" style=\"color:#90A4AE;text-decoration:none\">Terminar sessão</a></div>" if REQUIRE_AUTH else ""}

  <div class="footer">
    Delphi de Optimização de Intervenções &nbsp;·&nbsp; {e(TOPIC_LABEL)}
    &nbsp;·&nbsp; Gerado em {datetime.datetime.now().strftime("%d/%m/%Y")}
  </div>

</div>
</body>
</html>"""


def build_index_page(interventions):
    rows = ""
    for i, intv in enumerate(interventions, 1):
        code  = intv["Código"]
        label = intv["Intervenção"]
        obj   = intv.get("Objectivo(s)", "")
        prog  = intv.get("Programa", "")
        comp  = intv.get("Componente", "")
        crumb = " › ".join(p for p in [prog, comp]
                           if p and p not in ("None", "nan", ""))
        rows += f"""
      <tr>
        <td style="padding:8px 12px;font-family:monospace;font-size:12px;
                   color:#546E7A;border-bottom:1px solid #eee">{e(code)}</td>
        <td style="padding:8px 12px;border-bottom:1px solid #eee">
          {"<div style='font-size:11px;color:#90A4AE;margin-bottom:2px'>" + e(crumb) + "</div>" if crumb else ""}
          <a href="{e(code)}.html" style="color:#1F4E79;font-weight:600;
             text-decoration:none">{e(label)}</a>
          <div style="font-size:12px;color:#78909C;margin-top:3px">{e(obj[:120])}{"…" if len(obj)>120 else ""}</div>
        </td>
      </tr>"""

    return f"""<!DOCTYPE html>
<html lang="pt">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<meta name="robots" content="noindex, nofollow">
<title>Catálogo de Intervenções — {e(TOPIC_LABEL)}</title>
{auth_guard_html()}
<style>{CSS}
  table {{ width:100%; border-collapse:collapse; }}
  tr:hover td {{ background:#f0f7ff !important; }}
</style>
</head>
<body>
<div class="page-wrap">
  <div class="page-header">
    <div class="topic-tag">Delphi de Optimização &nbsp;·&nbsp; 1ª Oficina</div>
    <h1>Catálogo de Intervenções<br>{e(TOPIC_LABEL)}</h1>
    <div class="objective">
      Seleccione uma intervenção para ver a ficha completa com objectivo,
      organização, ineficiências identificadas e causas.
    </div>
  </div>
  <div class="body-content" style="padding:0">
    <table>
      <thead>
        <tr style="background:#E3F2FD">
          <th style="padding:10px 12px;text-align:left;font-size:11px;
                     color:#1565C0;text-transform:uppercase;letter-spacing:1px;
                     border-bottom:2px solid #BBDEFB;width:100px">Código</th>
          <th style="padding:10px 12px;text-align:left;font-size:11px;
                     color:#1565C0;text-transform:uppercase;letter-spacing:1px;
                     border-bottom:2px solid #BBDEFB">Intervenção</th>
        </tr>
      </thead>
      <tbody>{rows}
      </tbody>
    </table>
  </div>
  <div class="footer">
    {len(interventions)} intervenções &nbsp;·&nbsp; {e(TOPIC_LABEL)}
    &nbsp;·&nbsp; Gerado em {datetime.datetime.now().strftime("%d/%m/%Y")}
    &nbsp;·&nbsp;
    <a href="{CATALOGUE_XLSX}" download
       style="color:#1565C0;text-decoration:none;font-weight:600">
      ⬇ Descarregar em Excel</a>
    &nbsp;&nbsp;|&nbsp;&nbsp;
    <a href="{MASTER_URL}"
       style="color:#1565C0;text-decoration:none;font-weight:600">
      ↩ Voltar ao painel de avaliação</a>
  </div>
</div>
</body>
</html>"""


# Generate individual pages
total = len(interventions)
index_url = "index.html"

for i, intv in enumerate(interventions):
    code     = intv["Código"]
    prev_url = f"{interventions[i-1]['Código']}.html" if i > 0 else None
    next_url = f"{interventions[i+1]['Código']}.html" if i < total-1 else None

    html_content = build_html_page(intv, i+1, total,
                                   prev_url, next_url, index_url)
    out_path = os.path.join(PAGES_DIR, f"{code}.html")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html_content)
    print(f"  {out_path}")

# ── Catalogue Excel export ────────────────────────────────────
CATALOGUE_XLSX = f"catalogo_{TOPIC_CODE}.xlsx"

# Columns to include in export (in order), skipping internal-only ones
EXPORT_COLS = [
    "Código", "Área", "Programa", "Componente", "Intervenção", "Nível",
    "Descrição (o que inclui)", "Objectivo(s)",
    "Alcance geográfico da intervenção",
    "Recursos necessários para a implementação",
    "Etapas chave para a implementação",
    "Riscos e limitações", "Possíveis factores associados aos riscos",
    "Ano de início", "Gastos em 2024 (MZN)", "Fonte(s) de financiamento",
    "Pop. elegível", "Número alcançado", "Cobertura", "Custo por unidade",
    "Notas", "URL da Ficha",
]

def generate_catalogue_xlsx(interventions, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Catálogo"

    # Header row
    cols = [c for c in EXPORT_COLS if any(intv.get(c) for intv in interventions)]
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill      = PatternFill("solid", start_color="1A5C8A")
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)

    # Data rows
    for ri, intv in enumerate(interventions, 2):
        for ci, col in enumerate(cols, 1):
            val = intv.get(col, "")
            cell = ws.cell(row=ri, column=ci, value=val if val not in ("None","nan","") else "")
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if ri % 2 == 0:
                cell.fill = PatternFill("solid", start_color="EEF4FB")

    # Column widths
    WIDTH_MAP = {
        "Código": 10, "Área": 16, "Programa": 20, "Componente": 20,
        "Intervenção": 35, "Nível": 12,
        "Descrição (o que inclui)": 45, "Objectivo(s)": 45,
        "Alcance geográfico da intervenção": 30,
        "Recursos necessários para a implementação": 40,
        "Etapas chave para a implementação": 40,
        "Riscos e limitações": 35,
        "Possíveis factores associados aos riscos": 35,
        "Ano de início": 12, "Gastos em 2024 (MZN)": 20,
        "Fonte(s) de financiamento": 25,
        "Pop. elegível": 16, "Número alcançado": 16,
        "Cobertura": 12, "Custo por unidade": 22, "Notas": 35,
        "URL da Ficha": 50,
    }
    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = WIDTH_MAP.get(col, 20)

    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 32
    wb.save(out_path)

catalogue_xlsx_path = os.path.join(PAGES_DIR, CATALOGUE_XLSX)
generate_catalogue_xlsx(interventions, catalogue_xlsx_path)
print(f"  {catalogue_xlsx_path}  (catalogue xlsx)")

# Generate index page
index_content = build_index_page(interventions)
index_path = os.path.join(PAGES_DIR, "index.html")
with open(index_path, "w", encoding="utf-8") as f:
    f.write(index_content)
print(f"  {index_path}  (index)")

# ════════════════════════════════════════════════════════════
# PART 3: MASTER PAGE
# ════════════════════════════════════════════════════════════
print("\nBuilding master page...")


# ── Master page builder ───────────────────────────────────────
def build_master_page(subform_groups):
    """
    Renders a checklist master page. Each sub-form is a card that the expert
    opens in sequence. Completion state is stored in localStorage so the
    checklist survives page refreshes on the same device/browser.

    After submitting a Kobo sub-form the expert is returned here via the
    ?completed=<slug> query parameter, which marks that card done.
    """
    # Build the JS groups array embedded in the page
    import json
    js_groups = []
    for label, slug, items in subform_groups:
        kobo_url = SUBFORM_URLS.get(slug, "")
        if kobo_url:
            # Build the relay target URL, then percent-encode it entirely
            # so Enketo treats the whole thing as a single return_url value.
            # Without this, Enketo splits on the second "?" or "&" in the value.
            from urllib.parse import quote
            relay_target = (f"{RELAY_URL}"
                            f"?slug={quote(slug, safe='')}"
                            f"&dest={quote(MASTER_URL, safe='')}")
            sep = "&" if "?" in kobo_url else "?"
            full_url = f"{kobo_url}{sep}return_url={quote(relay_target, safe='')}"
        else:
            full_url = ""
        js_groups.append({
            "slug":  slug,
            "label": label,
            "count": len(items),
            "codes": [intv["Código"] for intv in items],
            "url":   full_url,
        })

    groups_json = json.dumps(js_groups, ensure_ascii=False, indent=2)
    stub = ("" if "example.org" not in MASTER_URL else
            '<div class="stub-notice" style="margin:0 0 24px">&#9888; '
            'MASTER_URL não configurado — os links de retorno após submissão '
            'não funcionarão até ser definido em config.env.</div>')

    return f"""<!DOCTYPE html>
<html lang="pt">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<meta name="robots" content="noindex, nofollow">
<title>Painel de Avaliação — {e(TOPIC_LABEL)}</title>
{auth_guard_html()}
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Fraunces:wght@300;400;600&family=DM+Sans:wght@300;400;500&display=swap" rel="stylesheet">
<style>
*, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}
:root {{
  --ink:    #0f1923;
  --muted:  #5a6a7a;
  --faint:  #e8edf2;
  --accent: #1a5c8a;
  --green:  #2e7d52;
  --amber:  #b45309;
  --surface:#f7f9fb;
  --white:  #ffffff;
}}
html, body {{
  min-height: 100vh;
  font-family: 'DM Sans', sans-serif;
  background: var(--surface);
  color: var(--ink);
}}
body::before {{
  content: '';
  position: fixed; inset: 0;
  background:
    radial-gradient(ellipse 80% 50% at 10% 15%, rgba(26,92,138,.07) 0%, transparent 60%),
    radial-gradient(ellipse 60% 70% at 90% 85%, rgba(46,125,82,.05)  0%, transparent 60%);
  pointer-events: none;
}}
.page {{
  position: relative;
  max-width: 680px;
  margin: 0 auto;
  padding: 40px 24px 60px;
}}
/* ── Header ── */
.page-header {{
  background: var(--accent);
  border-radius: 12px;
  padding: 28px 32px;
  margin-bottom: 32px;
  position: relative;
  overflow: hidden;
}}
.page-header::after {{
  content: '';
  position: absolute;
  bottom: -30px; right: -30px;
  width: 150px; height: 150px;
  border-radius: 50%;
  background: rgba(255,255,255,.06);
}}
.topic-pill {{
  display: inline-block;
  background: rgba(255,255,255,.18);
  color: rgba(255,255,255,.9);
  font-size: 11px; font-weight: 500;
  letter-spacing: 1px; text-transform: uppercase;
  padding: 3px 10px; border-radius: 20px;
  margin-bottom: 12px;
}}
.page-title {{
  font-family: 'Fraunces', serif;
  font-weight: 300; font-size: 24px;
  color: #fff; line-height: 1.3;
  margin-bottom: 6px;
}}
.page-subtitle {{ font-size: 13px; color: rgba(255,255,255,.65); }}
/* ── Progress bar ── */
.progress-wrap {{
  margin-bottom: 28px;
}}
.progress-label {{
  display: flex; justify-content: space-between;
  font-size: 12px; color: var(--muted);
  margin-bottom: 8px;
}}
.progress-bar-bg {{
  height: 6px; background: var(--faint);
  border-radius: 3px; overflow: hidden;
}}
.progress-bar-fill {{
  height: 100%; background: var(--green);
  border-radius: 3px;
  transition: width .4s ease;
  width: 0%;
}}
/* ── Sub-form cards ── */
.subform-card {{
  background: var(--white);
  border: 1.5px solid var(--faint);
  border-radius: 10px;
  padding: 20px 24px;
  margin-bottom: 14px;
  display: flex;
  align-items: center;
  gap: 16px;
  transition: box-shadow .2s, border-color .2s;
}}
.subform-card.done {{
  border-color: #b7dfcc;
  background: #f6fdf9;
}}
.subform-card.locked {{
  opacity: .55;
  cursor: not-allowed;
}}
.status-icon {{
  flex-shrink: 0;
  width: 36px; height: 36px;
  border-radius: 50%;
  display: flex; align-items: center; justify-content: center;
  font-size: 16px;
  background: var(--faint);
  color: var(--muted);
  font-weight: 600;
  font-family: 'DM Sans', sans-serif;
}}
.subform-card.done .status-icon {{
  background: #d1f0e2; color: var(--green);
}}
.card-body {{ flex: 1; min-width: 0; }}
.card-label {{
  font-weight: 500; font-size: 15px;
  color: var(--ink); margin-bottom: 3px;
}}
.card-meta {{
  font-size: 12px; color: var(--muted);
}}
.card-action {{ flex-shrink: 0; }}
.open-btn {{
  padding: 9px 18px;
  background: var(--accent);
  color: #fff;
  font-family: 'DM Sans', sans-serif;
  font-size: 13px; font-weight: 500;
  border: none; border-radius: 6px;
  cursor: pointer;
  text-decoration: none;
  display: inline-block;
  transition: background .2s;
}}
.open-btn:hover {{ background: #174e78; }}
.open-btn.done-btn {{
  background: transparent;
  border: 1.5px solid #b7dfcc;
  color: var(--green);
  cursor: default;
}}
.open-btn.no-url {{
  background: var(--faint);
  color: var(--muted);
  cursor: not-allowed;
}}
/* ── Catalogue link ── */
.catalogue-link {{
  text-align: center;
  margin-top: 32px;
  font-size: 13px;
  color: var(--muted);
}}
.catalogue-link a {{
  color: var(--accent);
  font-weight: 500;
  text-decoration: none;
}}
.catalogue-link a:hover {{ text-decoration: underline; }}
/* ── Footer ── */
.page-footer {{
  text-align: center;
  font-size: 11px; color: #b0bec5;
  margin-top: 40px;
}}
/* ── Auth logout link ── */
.signout {{
  text-align: right;
  font-size: 11px;
  margin-bottom: 12px;
}}
.signout a {{ color: #b0bec5; text-decoration: none; }}
.signout a:hover {{ color: var(--muted); }}
.stub-notice {{
  background: #EDE7F6;
  border: 1px dashed #9575CD;
  border-radius: 4px;
  padding: 10px 14px;
  font-size: 12px; color: #4527A0;
}}
</style>
</head>
<body>
<div class="page">

  {"<div class='signout'><a href='#' onclick='window._magicLogout&&window._magicLogout();return false;'>Terminar sessão</a></div>" if REQUIRE_AUTH else ""}

  <div class="page-header">
    <div class="topic-pill">{e(TOPIC_LABEL)}</div>
    <div class="page-title">Painel de Avaliação</div>
    <div class="page-subtitle">{e(TOPIC_LABEL)} &nbsp;·&nbsp; {len(subform_groups)} formulários &nbsp;·&nbsp; {sum(len(items) for _,_,items in subform_groups)} intervenções</div>
  </div>

  {stub}

  <div class="progress-wrap">
    <div class="progress-label">
      <span>Progresso</span>
      <span id="progressText">0 de {len(subform_groups)} formulários submetidos</span>
    </div>
    <div class="progress-bar-bg">
      <div class="progress-bar-fill" id="progressFill"></div>
    </div>
  </div>

  <div id="cardList"></div>

  <div class="catalogue-link">
    <a href="index.html">↗ Catálogo de fichas de intervenção</a>
  </div>

  <div class="page-footer">
    Delphi de Optimização &nbsp;·&nbsp; {e(TOPIC_LABEL)}
    &nbsp;·&nbsp; Gerado em {datetime.datetime.now().strftime("%d/%m/%Y")}
    &nbsp;&nbsp;|&nbsp;&nbsp;
    <button onclick="resetProgress()"
      style="background:none;border:1px solid #B0BEC5;border-radius:4px;
             padding:2px 10px;font-size:11px;color:#78909C;cursor:pointer;
             font-family:inherit"
      title="Apagar progresso guardado localmente e recomeçar">
      ↺ Reiniciar progresso
    </button>
    &nbsp;
    <button onclick="showDebug()"
      style="background:none;border:1px solid #B0BEC5;border-radius:4px;
             padding:2px 10px;font-size:11px;color:#78909C;cursor:pointer;
             font-family:inherit"
      title="Mostrar estado de progresso guardado neste dispositivo">
      🔍 Debug
    </button>
  </div>

  <div id="debugPanel" style="display:none;margin:12px auto;max-width:680px;
       background:#263238;color:#80CBC4;border-radius:6px;
       padding:14px 18px;font-size:11px;font-family:monospace;
       white-space:pre-wrap;text-align:left;"></div>

</div>

<script>
const GROUPS = {groups_json};
const STORAGE_KEY = 'delphi_{TOPIC_CODE}_completed';
const TOTAL = GROUPS.length;

// ── Persistence helpers ──────────────────────────────────────
function getCompleted() {{
  try {{ return JSON.parse(localStorage.getItem(STORAGE_KEY) || '[]'); }}
  catch(e) {{ return []; }}
}}
function markCompleted(slug) {{
  const done = getCompleted();
  if (!done.includes(slug)) {{ done.push(slug); }}
  localStorage.setItem(STORAGE_KEY, JSON.stringify(done));
}}
function resetProgress() {{
  if (confirm('Apagar todo o progresso guardado neste dispositivo e recomeçar?')) {{
    localStorage.removeItem(STORAGE_KEY);
    renderCards();
    const panel = document.getElementById('debugPanel');
    if (panel && panel.style.display !== 'none') showDebug();
  }}
}}

function showDebug() {{
  const panel = document.getElementById('debugPanel');
  if (!panel) return;
  if (panel.style.display !== 'none') {{
    panel.style.display = 'none';
    return;
  }}
  const completed = getCompleted();
  const allSlugs = GROUPS.map(g => g.slug);
  const pending = allSlugs.filter(s => !completed.includes(s));
  const lines = [
    '── localStorage key: ' + STORAGE_KEY,
    '── completed (' + completed.length + '/' + allSlugs.length + '): ' +
      (completed.length ? completed.join(', ') : '(none)'),
    '── pending (' + pending.length + '): ' +
      (pending.length ? pending.join(', ') : '(none)'),
    '',
    '── raw value:',
    localStorage.getItem(STORAGE_KEY) || '(not set)',
  ];
  panel.textContent = lines.join('\\n');
  panel.style.display = 'block';
}}

// ── Check for completed=<slug> on page load ──────────────────
// Uses URL hash (#completed=slug) instead of query string so that
// GitHub Pages normalisation redirects don't strip the parameter.
// Also supports legacy ?completed=slug query string as fallback.
(function handleReturn() {{
  let slug = null;

  // Hash-based: master.html#completed=manejo_de_casos_1
  const hash = window.location.hash.slice(1); // strip leading #
  if (hash.startsWith('completed=')) {{
    slug = decodeURIComponent(hash.slice('completed='.length));
    history.replaceState(null, '', window.location.pathname);
  }}

  // Query string fallback: master.html?completed=manejo_de_casos_1
  if (!slug) {{
    const params = new URLSearchParams(window.location.search);
    slug = params.get('completed');
    if (slug) history.replaceState(null, '', window.location.pathname);
  }}

  if (slug) markCompleted(slug);
}})();

// ── Render cards ─────────────────────────────────────────────
function renderCards() {{
  const done = getCompleted();
  const doneCount = GROUPS.filter(g => done.includes(g.slug)).length;
  const pct = TOTAL > 0 ? Math.round(doneCount / TOTAL * 100) : 0;

  document.getElementById('progressFill').style.width = pct + '%';
  document.getElementById('progressText').textContent =
    doneCount + ' de ' + TOTAL + ' formulários submetidos';

  const list = document.getElementById('cardList');
  list.innerHTML = '';

  GROUPS.forEach(function(g, idx) {{
    const isDone  = done.includes(g.slug);
    const hasUrl  = !!g.url;
    const card    = document.createElement('div');
    card.className = 'subform-card' + (isDone ? ' done' : '');

    // Status icon: checkmark if done, number if pending
    const icon = isDone ? '✓' : String(idx + 1);
    const iconBadge = '<div class="status-icon">' + icon + '</div>';

    // Meta: intervention count + codes
    const codesSnippet = g.codes.slice(0, 4).join(', ')
                       + (g.codes.length > 4 ? ', …' : '');
    const meta = g.count + ' intervenção' + (g.count !== 1 ? 'ões' : '')
               + ' &nbsp;·&nbsp; ' + codesSnippet;

    // Action button
    let btn = '';
    if (isDone) {{
      btn = '<span class="open-btn done-btn">✓ Submetido</span>';
    }} else if (!hasUrl) {{
      btn = '<span class="open-btn no-url" title="URL não configurado">Em breve</span>';
    }} else {{
      btn = '<a class="open-btn" href="' + g.url + '">Iniciar →</a>';
    }}

    card.innerHTML =
      iconBadge +
      '<div class="card-body">' +
        '<div class="card-label">' + g.label + '</div>' +
        '<div class="card-meta">' + meta + '</div>' +
      '</div>' +
      '<div class="card-action">' + btn + '</div>';

    list.appendChild(card);
  }});
}}

renderCards();
</script>
</body>
</html>"""


# ── Write master page ─────────────────────────────────────────
master_content = build_master_page(subform_groups)
master_path = os.path.join(PAGES_DIR, MASTER_PAGE)
with open(master_path, "w", encoding="utf-8") as f:
    f.write(master_content)
print(f"\n  {master_path}  (master)")

# ── Write relay page ──────────────────────────────────────────
# Enketo returns to relay.html?slug=<slug>&dest=<master_url>.
# The relay writes the slug to localStorage (same origin as master.html)
# then immediately redirects to master.html — bypassing any GitHub Pages
# redirect that would otherwise strip query parameters from master.html.
relay_content = f"""<!DOCTYPE html>
<html lang="pt">
<head>
<meta charset="UTF-8">
<title>A redirecionar…</title>
<meta name="robots" content="noindex, nofollow">
<style>
  body {{ font-family: sans-serif; display: flex; align-items: center;
         justify-content: center; min-height: 100vh; margin: 0;
         background: #f7f9fb; color: #5a6a7a; }}
</style>
</head>
<body>
<p>A registar submissão…</p>
<script>
(function() {{
  var params  = new URLSearchParams(window.location.search);
  var slug    = params.get('slug');
  var dest    = params.get('dest') || 'master.html';
  var key     = 'delphi_{TOPIC_CODE}_completed';

  if (slug) {{
    try {{
      var done = JSON.parse(localStorage.getItem(key) || '[]');
      if (!done.includes(slug)) done.push(slug);
      localStorage.setItem(key, JSON.stringify(done));
    }} catch(e) {{}}
  }}

  // Small delay so the write completes before navigation
  setTimeout(function() {{ window.location.replace(dest); }}, 100);
}})();
</script>
</body>
</html>"""

relay_path = os.path.join(PAGES_DIR, "relay.html")
with open(relay_path, "w", encoding="utf-8") as f:
    f.write(relay_content)
print(f"  {relay_path}  (relay)")

print(f"\n✅ Generated {total} intervention pages + 1 index + 1 master + 1 relay page")
print(f"   Pages directory: {PAGES_DIR}")
if any(not SUBFORM_URLS.get(slug) for _, slug, _ in subform_groups):
    print("\n⚠️  Some sub-forms have no Kobo URL yet.")
    print("   Deploy one Kobo form per group, then add to config.env:")
    for label, slug, items in subform_groups:
        if not SUBFORM_URLS.get(slug):
            print(f"   SUBFORM_URL_{slug} = https://ee.kobotoolbox.org/x/XXXXXXXX")
print("\nDone.")
