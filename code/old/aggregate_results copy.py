#!/usr/bin/env python3
"""
aggregate_results.py
────────────────────
Aggregates Kobo CSV exports from all sub-form groups into a single
structured Excel workbook, then runs QC checks for completeness and
consistency.

Usage:
    python3 aggregate_results.py                        # auto-detect CSVs in current folder
    python3 aggregate_results.py --input exports/       # folder of CSVs
    python3 aggregate_results.py --fetch                # pull directly from KoboToolbox API
    python3 aggregate_results.py --fetch --assets UID1 UID2  # fetch specific assets only
    python3 aggregate_results.py --qc-only              # re-run QC on existing output
    python3 aggregate_results.py --output results.xlsx  # custom output path

Input:  One CSV export per sub-form group, downloaded from KoboToolbox
        (Project → Data → Downloads → CSV, "XML values and headers")
Output: delphi_w1_malaria_results.xlsx  (aggregated data + QC report)

Sheets produced:
    Raw          — all submissions stacked, one row per submission
    Wide         — one row per expert × intervention (primary analysis sheet)
    QC_Summary   — one row per check, PASS/FAIL/WARN with counts
    QC_Detail    — one row per issue found (expert, intervention, check, detail)
    Coverage     — expert × group submission matrix
"""

import os, sys, re, glob, argparse
from collections import defaultdict

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════
# Config (mirrors config.env)
# ═══════════════════════════════════════════════════════════════
def load_config(path="config.env"):
    cfg = {}
    if not os.path.exists(path):
        return cfg
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, _, v = line.partition("=")
            cfg[k.strip()] = v.strip()
    return cfg

cfg       = load_config()
TOPIC     = cfg.get("TOPIC_CODE", "malaria")
OUT_FILE  = cfg.get("RESULTS_FILE", f"delphi_w1_{TOPIC}_results.xlsx")
KOBO_SERVER  = cfg.get("KOBO_SERVER", "https://eu.kobotoolbox.org")
KOBO_TOKEN   = cfg.get("KOBO_TOKEN",  "")

# ═══════════════════════════════════════════════════════════════
# Question schema — derived from dictionary
# Defines field suffixes, valid values, and skip rules
# ═══════════════════════════════════════════════════════════════

# Fields present once per form (identification section)
ID_FIELDS = ["expert_code", "modality"]

# Per-intervention field suffixes (in order)
INTV_FIELDS = [
    "exp",            # expertise level
    "gate",           # optimizability (gating question)
    "dup",            # duplication: yes/no
    "which_dup",      # which interventions duplicate (select_multiple)
    "which_dup_other",# free text if 'other' selected
    "intg",           # integration: yes/no
    "which_intg",     # which interventions to integrate
    "which_intg_other",
    "res",            # resources: yes/no
    "oth",            # other reason: yes/no
    "oth_reason",     # free text reason
    "impact",         # impact score
    "cmt",            # comment (optional)
]

# Valid values per field (None = any non-empty string is valid)
VALID_VALUES = {
    "modality":  {"presencial", "remoto"},
    "exp":       {"1", "2", "3"},
    "gate":      {"sim_def", "possivelmente", "nao"},
    "dup":       {"sim", "nao"},
    "intg":      {"sim", "nao"},
    "res":       {"sim", "nao"},
    "oth":       {"sim", "nao"},
    "impact":    {"1", "2", "3"},
}

# Fields that should be EMPTY when gate = 'nao'
GATED_FIELDS = ["dup", "which_dup", "which_dup_other",
                "intg", "which_intg", "which_intg_other",
                "res", "oth", "oth_reason", "impact"]

# Skip rules: field → condition under which it should be non-empty
# Expressed as (parent_field, required_value, [additional_condition])
REQUIRED_IF = {
    "which_dup":        [("dup",  "sim"), ("gate", "!nao")],
    "which_intg":       [("intg", "sim"), ("gate", "!nao")],
    "oth_reason":       [("oth",  "sim"), ("gate", "!nao")],
}

# Fields that should be filled whenever gate != 'nao'
REQUIRED_WHEN_NOT_GATED = ["dup", "intg", "res", "oth", "impact"]

# ═══════════════════════════════════════════════════════════════
# Step 1a — Fetch submissions directly from KoboToolbox API
# ═══════════════════════════════════════════════════════════════

def _api_headers():
    if not KOBO_TOKEN:
        print("❌  KOBO_TOKEN not set in config.env")
        print("    Get your token from: Account Settings → Security → API Token")
        sys.exit(1)
    return {"Authorization": f"Token {KOBO_TOKEN}", "Accept": "application/json"}

def list_assets():
    """Return all assets (forms) visible to this token."""
    import requests
    url = f"{KOBO_SERVER.rstrip('/')}/api/v2/assets/"
    assets = []
    while url:
        r = requests.get(url, headers=_api_headers(), params={"limit": 100})
        r.raise_for_status()
        data = r.json()
        assets.extend(data.get("results", []))
        url = data.get("next")
    return assets

def get_configured_assets():
    """
    Read SUBFORM_ASSET_<slug> = <uid> entries from config.env.
    Returns list of dicts with keys: uid, slug.
    """
    assets = []
    for k, v in cfg.items():
        if k.upper().startswith("SUBFORM_ASSET_") and v.strip():
            slug = k[len("SUBFORM_ASSET_"):].strip()
            assets.append({"uid": v.strip(), "slug": slug})
    return assets

def find_delphi_assets():
    """
    Return assets to fetch, in priority order:
    1. SUBFORM_ASSET_* UIDs from config.env  (set by deploy_kobo_forms.py)
    2. Fallback: search account assets by name pattern (requires list_assets())
    Each returned dict has at minimum: uid, name, deployment__active, has_deployment.
    """
    import requests

    configured = get_configured_assets()
    if configured:
        print(f"  Using {len(configured)} asset UID(s) from config.env")
        assets = []
        for entry in configured:
            r = requests.get(
                f"{KOBO_SERVER.rstrip('/')}/api/v2/assets/{entry['uid']}/",
                headers=_api_headers(),
                params={"format": "json"},
            )
            if r.ok:
                a = r.json()
                a.setdefault("name", entry["slug"])
                assets.append(a)
            else:
                print(f"  ⚠️  Could not fetch asset {entry['uid']} "
                      f"(slug={entry['slug']}): {r.status_code}")
        return assets

    # Fallback: search by name pattern
    print(f"  No SUBFORM_ASSET_* entries in config.env — searching account assets by name…")
    all_assets = list_assets()
    print(f"  Found {len(all_assets)} total asset(s) on account. Filtering…")

    # Try progressively looser matches
    topic_word = TOPIC_CODE.lower()   # e.g. "malaria"
    matched = []
    for a in all_assets:
        name = a.get("name", "").lower()
        id_str = ((a.get("settings") or {}).get("id_string") or "").lower()
        if (topic_word in name or
                topic_word in id_str or
                re.search(r"delphi", name)):
            matched.append(a)

    if not matched:
        print(f"\n  ⚠️  No assets matched topic='{TOPIC_CODE}' or keyword 'delphi'.")
        print(f"  All assets on this account:")
        for a in all_assets:
            print(f"    uid={a['uid']}  name={a.get('name','(unnamed)')}")
        print(f"\n  To fix: add SUBFORM_ASSET_<slug> = <uid> lines to config.env")
        print(f"  using the UIDs printed above, then re-run.")

    return matched


def list_all_assets_diagnostic():
    """Print all assets on the account with key fields — for troubleshooting."""
    import requests
    print(f"\nAll assets on {KOBO_SERVER}:")
    print(f"{'─'*80}")
    assets = list_assets()
    if not assets:
        print("  (none found — check KOBO_TOKEN)")
        return
    for a in assets:
        active  = a.get("deployment__active")
        status  = "live" if active is True else ("archived" if active is False else "draft")
        n_subs  = a.get("deployment__submission_count", 0) or 0
        print(f"  uid={a['uid']}  [{status:8}]  subs={n_subs:>4}  {a.get('name','')}")
    print(f"{'─'*80}")
    print(f"  Total: {len(assets)} asset(s)")
    print(f"\n  To use specific assets, add to config.env:")
    print(f"    SUBFORM_ASSET_<slug> = <uid>")
    print()

def fetch_submissions(asset_uid, asset_name):
    """
    Fetch all submissions for one asset via the data endpoint.
    Returns a DataFrame with one row per submission.
    Handles pagination automatically.
    """
    import requests
    base = f"{KOBO_SERVER.rstrip('/')}/api/v2/assets/{asset_uid}/data/"
    rows = []
    url  = base
    page = 1
    while url:
        r = requests.get(url, headers=_api_headers(),
                         params={"format": "json", "limit": 300})
        r.raise_for_status()
        data = r.json()
        results = data.get("results", [])
        rows.extend(results)
        url = data.get("next")
        page += 1

    if not rows:
        print(f"  Fetched: {asset_name}  (0 submissions — no data yet)")
        return pd.DataFrame()

    df = pd.json_normalize(rows)
    df = normalise_columns(df)

    # Derive group slug from asset name or id_string
    slug = re.sub(rf"^delphi_w1_{TOPIC}_", "",
                  asset_name.lower().replace(" ", "_").replace("|","_"))
    slug = re.sub(r"[^a-z0-9_]", "_", slug).strip("_")
    df["_group"]       = slug
    df["_source_file"] = f"API:{asset_uid}"
    print(f"  Fetched: {asset_name}  ({len(df)} submissions, group={slug})")
    return df

def fetch_all(asset_filter=None):
    """
    Find all matching Delphi assets and fetch their submissions.
    Returns stacked DataFrame (same format as load_csvs output).
    asset_filter: optional list of asset UIDs to restrict fetch.
    """
    assets = find_delphi_assets()
    if not assets:
        print(f"❌  No assets found.")
        print(f"    Either add SUBFORM_ASSET_<slug> = <uid> entries to config.env")
        print(f"    (run deploy_kobo_forms.py first), or check KOBO_TOKEN is valid.")
        sys.exit(1)

    print(f"  Found {len(assets)} matching asset(s):")
    for a in assets:
        print(f"    {a['uid']}  {a['name']}")

    if asset_filter:
        assets = [a for a in assets if a["uid"] in asset_filter]

    frames = []
    for a in assets:
        has_deployment = a.get("has_deployment", False)
        active         = a.get("deployment__active")  # True=live, False=archived, None=never deployed

        if active is False:
            print(f"  Skipping {a['name']} (archived — deployment__active=false)")
            continue
        if not has_deployment or active is None:
            print(f"  Skipping {a['name']} (never deployed — no active deployment)")
            continue
        df = fetch_submissions(a["uid"], a["name"])
        if not df.empty:
            frames.append(df)

    if not frames:
        print(f"  ⚠️  No submissions found across any asset — output will be empty.")
        return pd.DataFrame()

    raw = pd.concat(frames, ignore_index=True, sort=False)
    return raw


# ═══════════════════════════════════════════════════════════════
# Step 1b — Load and normalise CSVs
# ═══════════════════════════════════════════════════════════════

def find_csvs(folder="."):
    """Find all CSV files that look like Kobo exports in folder."""
    pattern = os.path.join(folder, "*.csv")
    files = sorted(glob.glob(pattern))
    if not files:
        # Also try common download folder names
        for sub in ("exports", "downloads", "data"):
            files = sorted(glob.glob(os.path.join(folder, sub, "*.csv")))
            if files:
                break
    return files

def normalise_columns(df):
    """
    Strip Kobo group prefixes from column names.
    'grp_mal_01/exp_mal_01' → 'exp_mal_01'
    Also strips leading/trailing whitespace.
    """
    df.columns = [
        re.sub(r"^[^/]+/", "", c).strip()
        for c in df.columns
    ]
    return df

def detect_group_slug(filepath, df):
    """
    Try to identify which sub-form group this CSV belongs to.
    First checks filename, then checks intervention codes present.
    """
    # From filename: delphi_w1_malaria_kobo_manejo_de_casos_1_...csv
    m = re.search(r"kobo_(.+?)(?:_\d{4}|\.csv)", os.path.basename(filepath))
    if m:
        return m.group(1)
    # From intervention codes in columns
    codes = detect_intervention_codes(df)
    if codes:
        return "_".join(sorted(codes)[:2]) + "_etc"
    return os.path.splitext(os.path.basename(filepath))[0]

def detect_intervention_codes(df):
    """Find all intervention codes present as column suffixes."""
    codes = set()
    for col in df.columns:
        m = re.match(r"gate_(.+)", col)
        if m:
            codes.add(m.group(1))
    return sorted(codes)

def load_csvs(csv_files):
    """Load all CSVs, normalise, tag with group slug, return stacked df."""
    frames = []
    for f in csv_files:
        try:
            df = pd.read_csv(f, dtype=str, encoding="utf-8")
        except UnicodeDecodeError:
            df = pd.read_csv(f, dtype=str, encoding="latin-1")

        df = normalise_columns(df)
        slug = detect_group_slug(f, df)
        df["_group"] = slug
        df["_source_file"] = os.path.basename(f)
        frames.append(df)
        print(f"  Loaded: {os.path.basename(f)}  ({len(df)} rows, group={slug})")

    if not frames:
        return pd.DataFrame()

    # Stack — use outer join so different groups' columns are preserved
    raw = pd.concat(frames, ignore_index=True, sort=False)
    return raw

# ═══════════════════════════════════════════════════════════════
# Step 2 — Build wide table (one row per expert × intervention)
# ═══════════════════════════════════════════════════════════════

def build_wide(raw):
    """
    Melt the stacked submissions into one row per (expert, intervention).
    Each row has: expert_code, modality, group, intervention_code,
    plus one column per INTV_FIELDS suffix.
    """
    rows = []
    for _, submission in raw.iterrows():
        expert   = str(submission.get("expert_code", "")).strip()
        modality = str(submission.get("modality",    "")).strip()
        group    = str(submission.get("_group",      "")).strip()
        source   = str(submission.get("_source_file","")).strip()
        codes    = detect_intervention_codes(submission.to_frame().T)

        for code in codes:
            row = {
                "expert_code":      expert,
                "modality":         modality,
                "group":            group,
                "intervention":     code,
                "_source_file":     source,
            }
            for suffix in INTV_FIELDS:
                col = f"{suffix}_{code}"
                val = submission.get(col, pd.NA)
                row[suffix] = "" if pd.isna(val) else str(val).strip()
            rows.append(row)

    if not rows:
        return pd.DataFrame()

    wide = pd.DataFrame(rows)
    # Sort for readability
    wide = wide.sort_values(["expert_code", "intervention"]).reset_index(drop=True)
    return wide

# ═══════════════════════════════════════════════════════════════
# Step 3 — QC checks
# ═══════════════════════════════════════════════════════════════

def _nonempty(val):
    return bool(val and str(val).strip() not in ("", "nan", "None"))

def run_qc(raw, wide):
    """
    Run all QC checks. Returns:
      summary  — list of dicts (check, status, count, description)
      details  — list of dicts (expert, intervention, check, detail)
    """
    summary = []
    details = []

    def add(check, status, count, description):
        summary.append({
            "Check":       check,
            "Status":      status,
            "Issues":      count,
            "Description": description,
        })

    def issue(expert, intv, check, detail):
        details.append({
            "Expert":       expert,
            "Intervention": intv,
            "Check":        check,
            "Detail":       detail,
        })

    # ── C1: Every expert submitted exactly once per group ────────
    if not raw.empty and "expert_code" in raw.columns and "_group" in raw.columns:
        dup_subs = (raw.groupby(["expert_code", "_group"])
                      .size()
                      .reset_index(name="n")
                      .query("n > 1"))
        for _, r in dup_subs.iterrows():
            issue(r["expert_code"], "(all)", "C1_duplicate_submission",
                  f"Submitted {int(r['n'])} times for group '{r['_group']}'")
        add("C1_duplicate_submission", "FAIL" if len(dup_subs) else "PASS",
            len(dup_subs),
            "Each expert should submit exactly once per sub-form group")

    # ── C2: Every expert answered every intervention exactly once ─
    if not wide.empty:
        dup_intv = (wide.groupby(["expert_code", "intervention"])
                        .size()
                        .reset_index(name="n")
                        .query("n > 1"))
        for _, r in dup_intv.iterrows():
            issue(r["expert_code"], r["intervention"], "C2_duplicate_answer",
                  f"Answered {int(r['n'])} times")
        add("C2_duplicate_answer", "FAIL" if len(dup_intv) else "PASS",
            len(dup_intv),
            "Each expert should answer each intervention exactly once")

    # ── C3: No missing gate answers ──────────────────────────────
    if not wide.empty:
        missing_gate = wide[~wide["gate"].apply(_nonempty)]
        for _, r in missing_gate.iterrows():
            issue(r["expert_code"], r["intervention"], "C3_missing_gate",
                  "gate (optimizability) is empty")
        add("C3_missing_gate", "FAIL" if len(missing_gate) else "PASS",
            len(missing_gate),
            "gate (Q05 optimizability) must be answered for every intervention")

    # ── C4: Gated fields filled when gate = 'nao' ────────────────
    if not wide.empty:
        gated_rows = wide[wide["gate"] == "nao"]
        n_violations = 0
        for _, r in gated_rows.iterrows():
            for f in GATED_FIELDS:
                if _nonempty(r.get(f, "")):
                    issue(r["expert_code"], r["intervention"], "C4_gated_field_filled",
                          f"'{f}' is filled but gate='nao' (should be skipped)")
                    n_violations += 1
        add("C4_gated_field_filled", "FAIL" if n_violations else "PASS",
            n_violations,
            "Fields after gate should be empty when gate='nao'")

    # ── C5: Required-when-not-gated fields missing ───────────────
    if not wide.empty:
        active_rows = wide[wide["gate"] != "nao"]
        n_missing = 0
        for _, r in active_rows.iterrows():
            for f in REQUIRED_WHEN_NOT_GATED:
                if not _nonempty(r.get(f, "")):
                    issue(r["expert_code"], r["intervention"], "C5_required_field_missing",
                          f"'{f}' is empty but gate='{r['gate']}' (should be answered)")
                    n_missing += 1
        add("C5_required_field_missing", "FAIL" if n_missing else "PASS",
            n_missing,
            "Required fields (dup, intg, res, oth, impact) must be answered when gate != 'nao'")

    # ── C6: Skip logic — which_dup ───────────────────────────────
    if not wide.empty:
        n = 0
        for _, r in wide.iterrows():
            gate = r.get("gate", "")
            dup  = r.get("dup",  "")
            val  = r.get("which_dup", "")
            # Should be filled if dup=sim and gate!=nao
            should_fill = (dup == "sim" and gate != "nao")
            if should_fill and not _nonempty(val):
                issue(r["expert_code"], r["intervention"], "C6_skip_which_dup",
                      "which_dup is empty but dup='sim' — which interventions duplicate?")
                n += 1
            elif not should_fill and _nonempty(val):
                issue(r["expert_code"], r["intervention"], "C6_skip_which_dup",
                      f"which_dup is filled but dup='{dup}'/gate='{gate}' (should be skipped)")
                n += 1
        add("C6_skip_which_dup", "FAIL" if n else "PASS", n,
            "which_dup should be filled iff dup='sim' and gate!='nao'")

    # ── C7: Skip logic — which_intg ──────────────────────────────
    if not wide.empty:
        n = 0
        for _, r in wide.iterrows():
            gate = r.get("gate", "")
            intg = r.get("intg", "")
            val  = r.get("which_intg", "")
            should_fill = (intg == "sim" and gate != "nao")
            if should_fill and not _nonempty(val):
                issue(r["expert_code"], r["intervention"], "C7_skip_which_intg",
                      "which_intg is empty but intg='sim'")
                n += 1
            elif not should_fill and _nonempty(val):
                issue(r["expert_code"], r["intervention"], "C7_skip_which_intg",
                      f"which_intg filled but intg='{intg}'/gate='{gate}' (should be skipped)")
                n += 1
        add("C7_skip_which_intg", "FAIL" if n else "PASS", n,
            "which_intg should be filled iff intg='sim' and gate!='nao'")

    # ── C8: Skip logic — oth_reason ──────────────────────────────
    if not wide.empty:
        n = 0
        for _, r in wide.iterrows():
            gate = r.get("gate", "")
            oth  = r.get("oth",  "")
            val  = r.get("oth_reason", "")
            should_fill = (oth == "sim" and gate != "nao")
            if should_fill and not _nonempty(val):
                issue(r["expert_code"], r["intervention"], "C8_skip_oth_reason",
                      "oth_reason is empty but oth='sim'")
                n += 1
            elif not should_fill and _nonempty(val):
                issue(r["expert_code"], r["intervention"], "C8_skip_oth_reason",
                      f"oth_reason filled but oth='{oth}'/gate='{gate}' (should be skipped)")
                n += 1
        add("C8_skip_oth_reason", "FAIL" if n else "PASS", n,
            "oth_reason should be filled iff oth='sim' and gate!='nao'")

    # ── C9: Valid values ─────────────────────────────────────────
    if not wide.empty:
        n = 0
        for field, valid in VALID_VALUES.items():
            if field not in wide.columns:
                continue
            col = wide[field]
            # Only check non-empty values
            bad = wide[col.apply(_nonempty) & ~col.isin(valid)]
            for _, r in bad.iterrows():
                issue(r["expert_code"], r.get("intervention", ""),
                      "C9_invalid_value",
                      f"'{field}' = '{r[field]}' — expected one of {sorted(valid)}")
                n += 1
        add("C9_invalid_value", "FAIL" if n else "PASS", n,
            "All select_one responses must be within valid choice list")

    # ── C10: Reciprocal duplication flags ────────────────────────
    # If expert X flags A as duplicate of B, did they also flag B as dup of A?
    if not wide.empty and "which_dup" in wide.columns:
        n = 0
        for expert, grp in wide.groupby("expert_code"):
            dup_flags = {}
            for _, r in grp.iterrows():
                if _nonempty(r.get("which_dup", "")):
                    targets = [t.strip() for t in str(r["which_dup"]).split()
                               if t.strip() and t.strip() != "other"]
                    dup_flags[r["intervention"]] = set(targets)
            for intv_a, targets in dup_flags.items():
                for intv_b in targets:
                    if intv_b in dup_flags:
                        if intv_a not in dup_flags[intv_b]:
                            issue(expert, intv_a, "C10_asymmetric_duplication",
                                  f"Flagged '{intv_b}' as duplicate but '{intv_b}' did not flag '{intv_a}'")
                            n += 1
                    # intv_b not in same group — can't check, skip silently
        add("C10_asymmetric_duplication", "WARN" if n else "PASS", n,
            "If A is a duplicate of B, ideally B should also flag A (cross-group pairs can't be verified)")

    # ── C11: Coverage matrix completeness ────────────────────────
    if not raw.empty and "expert_code" in raw.columns and "_group" in raw.columns:
        experts = sorted(raw["expert_code"].dropna().unique())
        groups  = sorted(raw["_group"].dropna().unique())
        submitted = set(zip(raw["expert_code"], raw["_group"]))
        missing = [(e, g) for e in experts for g in groups
                   if (e, g) not in submitted]
        for e, g in missing:
            issue(e, "(all)", "C11_missing_group_submission",
                  f"No submission found for group '{g}'")
        add("C11_missing_group_submission",
            "FAIL" if missing else "PASS",
            len(missing),
            "Every expert should have submitted every sub-form group")

    return summary, details

# ═══════════════════════════════════════════════════════════════
# Step 4 — Build coverage matrix
# ═══════════════════════════════════════════════════════════════

def build_coverage(raw):
    if raw.empty or "expert_code" not in raw.columns:
        return pd.DataFrame()
    experts = sorted(raw["expert_code"].dropna().unique())
    groups  = sorted(raw["_group"].dropna().unique())
    matrix  = []
    counts  = raw.groupby(["expert_code", "_group"]).size().to_dict()
    for e in experts:
        row = {"Expert": e}
        for g in groups:
            n = counts.get((e, g), 0)
            row[g] = "✓" if n == 1 else (f"×{n}" if n > 1 else "—")
        matrix.append(row)
    return pd.DataFrame(matrix)

# ═══════════════════════════════════════════════════════════════
# Step 5 — Write xlsx
# ═══════════════════════════════════════════════════════════════

# Colour palette
CLR = {
    "header_dark": "1A5C8A",   # accent blue
    "header_mid":  "2E7D52",   # green
    "header_light":"E3F2FD",
    "pass":        "E8F5E9",
    "fail":        "FFEBEE",
    "warn":        "FFF8E1",
    "alt_row":     "F7F9FB",
    "white":       "FFFFFF",
    "border":      "CFD8DC",
}

def _hdr(ws, row, col, value, bg=None, bold=True, color="FFFFFF", wrap=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=color, name="Arial", size=10)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                               wrap_text=wrap)
    return cell

def _border(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin", color=CLR["border"])
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(r, c).border = Border(
                left=thin, right=thin, top=thin, bottom=thin)

def write_sheet_df(wb, name, df, header_bg="1A5C8A", freeze="A2"):
    """Write a DataFrame to a new sheet with basic formatting."""
    ws = wb.create_sheet(name)
    if df.empty:
        ws["A1"] = "No data"
        return ws
    cols = list(df.columns)
    for ci, col in enumerate(cols, 1):
        _hdr(ws, 1, ci, col, bg=header_bg)
    status_col = cols.index("Status") + 1 if "Status" in cols else None
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        bg = CLR["white"] if ri % 2 == 0 else CLR["alt_row"]
        if status_col:
            s = str(row.get("Status", ""))
            bg = CLR["pass"] if s == "PASS" else (
                 CLR["fail"] if s == "FAIL" else
                 CLR["warn"] if s == "WARN" else bg)
        for ci, col in enumerate(cols, 1):
            v = row[col]
            # openpyxl cannot write lists/dicts — flatten to string
            if isinstance(v, (list, dict)):
                v = ", ".join(str(x) for x in v) if isinstance(v, list) else str(v)
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = Font(name="Arial", size=10)
            cell.fill = PatternFill("solid", start_color=bg)
            cell.alignment = Alignment(wrap_text=(col in ("Description","Detail")),
                                       vertical="top")
    # Auto-width
    for ci, col in enumerate(cols, 1):
        max_len = max(
            len(str(col)),
            *[len(str(df.iloc[r][col])) for r in range(min(len(df), 50))]
        )
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 60)
    _border(ws, 1, len(df) + 1, 1, len(cols))
    if freeze:
        ws.freeze_panes = freeze
    return ws

def write_xlsx(out_path, raw, wide, summary, details, coverage):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Raw
    write_sheet_df(wb, "Raw", raw.drop(columns=["_source_file"], errors="ignore"),
                   header_bg=CLR["header_dark"], freeze="C2")

    # Wide
    write_sheet_df(wb, "Wide", wide.drop(columns=["_source_file"], errors="ignore"),
                   header_bg=CLR["header_mid"], freeze="D2")

    # QC Summary
    if summary:
        write_sheet_df(wb, "QC_Summary", pd.DataFrame(summary),
                       header_bg="B45309", freeze="A2")

    # QC Detail
    if details:
        write_sheet_df(wb, "QC_Detail", pd.DataFrame(details),
                       header_bg="C0392B", freeze="C2")
    else:
        ws = wb.create_sheet("QC_Detail")
        ws["A1"] = "No issues found ✓"
        ws["A1"].font = Font(bold=True, color="2E7D52", name="Arial")

    # Coverage
    if not coverage.empty:
        write_sheet_df(wb, "Coverage", coverage,
                       header_bg=CLR["header_dark"], freeze="B2")

    wb.save(out_path)
    print(f"\n  Saved: {out_path}")

# ═══════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Aggregate Delphi W1 Kobo exports and run QC checks")
    parser.add_argument("--input",    default=".",
                        help="Folder containing CSV exports (default: current dir)")
    parser.add_argument("--output",   default=OUT_FILE,
                        help=f"Output xlsx path (default: {OUT_FILE})")
    parser.add_argument("--fetch",    action="store_true",
                        help="Fetch submissions directly from KoboToolbox API "
                             "(requires KOBO_TOKEN in config.env)")
    parser.add_argument("--list-assets", action="store_true",
                        help="List all assets on the KoboToolbox account and exit "
                             "(useful for finding UIDs to add to config.env)")
    parser.add_argument("--assets",   nargs="*", metavar="UID",
                        help="Restrict --fetch to specific asset UIDs")
    parser.add_argument("--qc-only",  action="store_true",
                        help="Re-run QC on existing output without re-reading data")
    args = parser.parse_args()

    print(f"\nDelphi W1 — Results Aggregator & QC")
    print(f"{'═'*45}")

    if args.list_assets:
        try:
            import requests
        except ImportError:
            print("❌  pip install requests")
            sys.exit(1)
        list_all_assets_diagnostic()
        sys.exit(0)

    if args.qc_only and os.path.exists(args.output):
        print(f"Loading existing output for QC re-run: {args.output}")
        raw  = pd.read_excel(args.output, sheet_name="Raw",  dtype=str).fillna("")
        wide = pd.read_excel(args.output, sheet_name="Wide", dtype=str).fillna("")

    elif args.fetch:
        try:
            import requests
        except ImportError:
            print("❌  requests library not installed. Run: pip install requests")
            sys.exit(1)
        print(f"\nFetching from KoboToolbox API: {KOBO_SERVER}")
        raw = fetch_all(asset_filter=args.assets)
        raw = raw.fillna("")
        if raw.empty:
            print("  No submissions yet — writing empty output file.")
        else:
            print(f"\n  Total submissions fetched: {len(raw)}")
        print("\nBuilding wide table (expert × intervention)…")
        wide = build_wide(raw)
        print(f"  Wide table: {len(wide)} rows "
              f"({wide['expert_code'].nunique() if not wide.empty else 0} experts × "
              f"{wide['intervention'].nunique() if not wide.empty else 0} interventions)")

    else:
        csv_files = find_csvs(args.input)
        if not csv_files:
            print(f"❌  No CSV files found in: {os.path.abspath(args.input)}")
            print("    Export CSVs from KoboToolbox (Data → Downloads → CSV)")
            print("    or use --fetch to pull directly from the API.")
            sys.exit(1)
        print(f"\nLoading {len(csv_files)} CSV file(s) from: {os.path.abspath(args.input)}")
        raw  = load_csvs(csv_files)
        raw  = raw.fillna("")
        print(f"\n  Total submissions loaded: {len(raw)}")
        print("\nBuilding wide table (expert × intervention)…")
        wide = build_wide(raw)
        print(f"  Wide table: {len(wide)} rows "
              f"({wide['expert_code'].nunique() if not wide.empty else 0} experts × "
              f"{wide['intervention'].nunique() if not wide.empty else 0} interventions)")

    print("\nRunning QC checks…")
    summary, details = run_qc(raw, wide)

    passes = sum(1 for s in summary if s["Status"] == "PASS")
    fails  = sum(1 for s in summary if s["Status"] == "FAIL")
    warns  = sum(1 for s in summary if s["Status"] == "WARN")

    print(f"\n  QC results: {passes} PASS  {fails} FAIL  {warns} WARN")
    for s in summary:
        icon = "✓" if s["Status"] == "PASS" else ("✗" if s["Status"] == "FAIL" else "⚠")
        print(f"  {icon} {s['Check']:40}  {s['Issues']:>4} issues")

    coverage = build_coverage(raw)

    print(f"\nWriting output…")
    write_xlsx(args.output, raw, wide,
               summary, details, coverage)

    if fails:
        print(f"\n⚠️   {fails} QC check(s) FAILED — review QC_Detail sheet.")
    else:
        print(f"\n✅  All QC checks passed.")
    print()

if __name__ == "__main__":
    main()
