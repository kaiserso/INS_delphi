from __future__ import annotations

import subprocess
import re
from itertools import zip_longest
from pathlib import Path
from typing import Dict, List

import pytest
import openpyxl


def discover_scenarios(repo_root: Path) -> list[str]:
    tests_dir = repo_root / "tests"
    out: list[str] = []
    for item in sorted(tests_dir.iterdir()):
        if item.is_dir() and (item / "config.env").exists():
            out.append(item.name)
    return out


def run_cmd(args: list[str], cwd: Path) -> subprocess.CompletedProcess[str]:
    # capture_output=True keeps pytest output readable while still allowing
    # assertions on command success and selected stdout markers.
    return subprocess.run(
        args,
        cwd=str(cwd),
        check=True,
        text=True,
        capture_output=True,
    )


def _norm(value: object) -> str:
    """Normalise cell values for stable workbook snapshot comparisons."""
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.10f}".rstrip("0").rstrip(".")
    text = str(value)
    # Dynamic version stamp injected in questionnaire hints.
    text = re.sub(
        r"\[Versão:\s*\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}\]",
        "[Versão: <timestamp>]",
        text,
    )
    # Dynamic generation timestamp in README / metadata text.
    text = re.sub(
        r"\(gerado\s+\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}\)",
        "(gerado <timestamp>)",
        text,
    )
    return text


def normalize_html_text(text: str) -> str:
    """Normalise dynamic and whitespace-only HTML differences for snapshots."""
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    # Footer generation date in master page.
    text = re.sub(r"Gerado em\s+\d{2}/\d{2}/\d{4}", "Gerado em <date>", text)
    # Gateway/template generation timestamp if present.
    text = re.sub(r"\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}", "<timestamp>", text)
    # Ignore trailing newline-only differences at EOF.
    return text.rstrip("\n")


def first_diff_line(expected_text: str, generated_text: str) -> str:
    """Return a concise first-difference hint for assertion messages."""
    expected_lines = expected_text.splitlines(keepends=True)
    generated_lines = generated_text.splitlines(keepends=True)
    for idx, (exp, got) in enumerate(
        zip_longest(expected_lines, generated_lines, fillvalue=None),
        start=1,
    ):
        if exp != got:
            return (
                f"first diff at line {idx}:\n"
                f"  expected: {exp!r}\n"
                f"  generated: {got!r}"
            )
    return "content differs (unable to localize first differing line)"


def first_workbook_diff(
    expected_snapshot: Dict[str, List[List[str]]],
    generated_snapshot: Dict[str, List[List[str]]],
) -> str:
    """Return concise first-difference hint across workbook sheets/cells."""
    all_sheets = sorted(set(expected_snapshot.keys()) | set(generated_snapshot.keys()))
    for sheet in all_sheets:
        if sheet not in expected_snapshot:
            return f"extra sheet in generated: {sheet!r}"
        if sheet not in generated_snapshot:
            return f"missing sheet in generated: {sheet!r}"

        exp_rows = expected_snapshot[sheet]
        got_rows = generated_snapshot[sheet]
        for row_idx, (exp_row, got_row) in enumerate(
            zip_longest(exp_rows, got_rows, fillvalue=None),
            start=1,
        ):
            if exp_row is None:
                return f"sheet {sheet!r}: extra generated row {row_idx}"
            if got_row is None:
                return f"sheet {sheet!r}: missing generated row {row_idx}"
            for col_idx, (exp_cell, got_cell) in enumerate(
                zip_longest(exp_row, got_row, fillvalue=""),
                start=1,
            ):
                if exp_cell != got_cell:
                    return (
                        f"sheet {sheet!r} row {row_idx} col {col_idx}: "
                        f"expected={exp_cell!r} generated={got_cell!r}"
                    )
    return "workbook differs (unable to localize first differing cell)"


def workbook_snapshot(path: Path) -> Dict[str, List[List[str]]]:
    """Return workbook content as a sheet->rows snapshot map."""
    wb = openpyxl.load_workbook(path, data_only=True)
    out: Dict[str, List[List[str]]] = {}
    for name in wb.sheetnames:
        rows: List[List[str]] = []
        ws = wb[name]
        for row in ws.iter_rows(values_only=True):
            rows.append([_norm(v) for v in row])
        while rows and all(v == "" for v in rows[-1]):
            rows.pop()

        if name.lower() == "settings" and rows:
            # settings.version is generated from current datetime (YYYYMMDDHH)
            # and is expected to vary between runs. Normalize it.
            header = rows[0]
            try:
                version_col = next(i for i, col in enumerate(header) if col.strip().lower() == "version")
            except StopIteration:
                version_col = None
            if version_col is not None:
                for r in range(1, len(rows)):
                    if version_col < len(rows[r]) and re.fullmatch(r"\d{10,14}", rows[r][version_col]):
                        rows[r][version_col] = "<version>"

        out[name] = rows
    return out


def expected_html_files(repo_root: Path, scenario: str) -> List[Path]:
    return sorted((repo_root / "tests" / scenario / "expected" / "pages").glob("*.html"))


def expected_pages_xlsx_files(repo_root: Path, scenario: str) -> List[Path]:
    return sorted((repo_root / "tests" / scenario / "expected" / "pages").glob("*.xlsx"))


def expected_kobo_files(repo_root: Path, scenario: str) -> List[Path]:
    return sorted((repo_root / "tests" / scenario / "expected" / "kobo").glob("*.xlsx"))


def expected_dictionary_path(repo_root: Path, scenario: str) -> Path:
    candidates = sorted((repo_root / "tests" / scenario / "expected").glob("dicionario_delphi_w1_*.xlsx"))
    assert candidates, f"No expected dictionary snapshot found for scenario: {scenario}"
    return candidates[0]


def generated_dictionary_path(repo_root: Path, scenario: str) -> Path:
    candidates = sorted((repo_root / "tests" / scenario / "out").glob("dicionario_delphi_w1_*.xlsx"))
    assert candidates, f"No generated dictionary found under tests/{scenario}/out"
    return candidates[0]


def assert_workbook_equal(generated: Path, expected: Path) -> None:
    got = workbook_snapshot(generated)
    exp = workbook_snapshot(expected)
    assert got == exp, (
        "Workbook snapshot mismatch\n"
        f"  expected: {expected}\n"
        f"  generated: {generated}\n"
        f"  {first_workbook_diff(exp, got)}"
    )


def assert_html_equal(generated: Path, expected: Path) -> None:
    exp_text = normalize_html_text(expected.read_text(encoding="utf-8"))
    got_text = normalize_html_text(generated.read_text(encoding="utf-8"))
    assert got_text == exp_text, (
        "HTML snapshot mismatch\n"
        f"  expected: {expected}\n"
        f"  generated: {generated}\n"
        f"  {first_diff_line(exp_text, got_text)}"
    )


@pytest.mark.regression
@pytest.mark.parametrize("scenario", ["hiv_teams", "hiv_no_teams"])
def test_base_variant_runs(repo_root: Path, python_exe: str, scenario: str) -> None:
    proc = run_cmd(
        [python_exe, "code/run_tests.py", "run", "--scenario", scenario],
        cwd=repo_root,
    )
    assert f"PASS  {scenario} [base]" in proc.stdout


@pytest.mark.regression
@pytest.mark.parametrize("scenario", ["hiv_teams", "hiv_no_teams"])
def test_base_variant_html_snapshots(repo_root: Path, python_exe: str, scenario: str) -> None:
    # Ensure out/pages is freshly generated for this scenario.
    run_cmd(
        [python_exe, "code/run_tests.py", "run", "--scenario", scenario],
        cwd=repo_root,
    )

    html_files = expected_html_files(repo_root, scenario)
    if not html_files:
        pytest.skip(f"No expected HTML snapshots found for scenario: {scenario}")

    for expected in html_files:
        generated = repo_root / "tests" / scenario / "out" / "pages" / expected.name
        assert generated.exists(), f"Missing generated HTML: {generated}"
        assert_html_equal(generated, expected)


@pytest.mark.regression
@pytest.mark.parametrize("scenario", ["hiv_teams", "hiv_no_teams"])
def test_base_variant_pages_xlsx_snapshots(repo_root: Path, python_exe: str, scenario: str) -> None:
    # Compare any page-side xlsx snapshots present under expected/pages.
    run_cmd(
        [python_exe, "code/run_tests.py", "run", "--scenario", scenario],
        cwd=repo_root,
    )

    expected_files = expected_pages_xlsx_files(repo_root, scenario)
    if not expected_files:
        pytest.skip(f"No expected pages XLSX snapshots found for scenario: {scenario}")

    for expected in expected_files:
        generated = repo_root / "tests" / scenario / "out" / "pages" / expected.name
        assert generated.exists(), f"Missing generated pages XLSX: {generated}"
        assert_workbook_equal(generated, expected)


@pytest.mark.regression
@pytest.mark.parametrize("scenario", ["hiv_teams", "hiv_no_teams"])
def test_base_variant_kobo_xlsx_snapshots(repo_root: Path, python_exe: str, scenario: str) -> None:
    # Compare selected kobo snapshots present under expected/kobo (one or many).
    run_cmd(
        [python_exe, "code/run_tests.py", "run", "--scenario", scenario],
        cwd=repo_root,
    )

    expected_files = expected_kobo_files(repo_root, scenario)
    if not expected_files:
        pytest.skip(f"No expected Kobo XLSX snapshots found for scenario: {scenario}")

    for expected in expected_files:
        generated = repo_root / "tests" / scenario / "out" / "kobo" / expected.name
        assert generated.exists(), f"Missing generated Kobo XLSX: {generated}"
        assert_workbook_equal(generated, expected)


@pytest.mark.integration
@pytest.mark.parametrize("scenario", ["hiv_teams", "hiv_no_teams"])
def test_integration_variant_runs(repo_root: Path, python_exe: str, scenario: str) -> None:
    proc = run_cmd(
        [python_exe, "code/run_tests.py", "run", "--scenario", scenario, "--variant", "integration"],
        cwd=repo_root,
    )
    assert f"PASS  {scenario} [integration]" in proc.stdout


@pytest.mark.integration
@pytest.mark.parametrize("scenario", ["hiv_teams", "hiv_no_teams"])
def test_integration_dictionary_matches_snapshot(repo_root: Path, python_exe: str, scenario: str) -> None:
    # Run the full pipeline first so out/ contains a fresh generated dictionary.
    run_cmd(
        [python_exe, "code/run_tests.py", "run", "--scenario", scenario, "--variant", "integration"],
        cwd=repo_root,
    )

    generated = generated_dictionary_path(repo_root, scenario)
    expected = expected_dictionary_path(repo_root, scenario)

    # Snapshot assertion: workbook structure + values must match expected.
    assert_workbook_equal(generated, expected)


@pytest.mark.regression
def test_hiv_teams_master_contains_team_badges(repo_root: Path, python_exe: str) -> None:
    run_cmd(
        [python_exe, "code/run_tests.py", "run", "--scenario", "hiv_teams"],
        cwd=repo_root,
    )
    master = repo_root / "tests" / "hiv_teams" / "out" / "pages" / "master.html"
    text = master.read_text(encoding="utf-8")
    assert "team-badge" in text
    assert 'const teamBadge = g.team ?' in text
    assert '"team": "' in text


@pytest.mark.regression
def test_hiv_no_teams_master_omits_team_badges(repo_root: Path, python_exe: str) -> None:
    run_cmd(
        [python_exe, "code/run_tests.py", "run", "--scenario", "hiv_no_teams"],
        cwd=repo_root,
    )
    master = repo_root / "tests" / "hiv_no_teams" / "out" / "pages" / "master.html"
    text = master.read_text(encoding="utf-8")
    assert "team-badge" not in text
    assert 'const teamBadge = \'\'' in text
    assert '"team": "' not in text
